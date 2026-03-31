#!/usr/bin/env python3
"""
VE3 Simple - Tạo file Excel mẫu (template)

Tạo file template.xlsx với 3 sheets:
- config: Cấu hình project (project_id, token, ...)
- characters: Danh sách nhân vật/địa điểm
- scenes: Danh sách scenes cần tạo ảnh

Chạy: python create_template.py
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def create_template(output_path: str = None):
    """Tạo file Excel mẫu."""
    if not output_path:
        output_path = str(Path(__file__).parent / "templates" / "template.xlsx")

    wb = Workbook()

    # Colors
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    guide_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    guide_font = Font(name="Arial", size=10, italic=True, color="888888")
    example_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def style_header(ws, headers, col_widths=None):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        if col_widths:
            for col, w in enumerate(col_widths, 1):
                ws.column_dimensions[chr(64 + col) if col <= 26 else 'A'].width = w

    # =========================================================================
    # Sheet 1: CONFIG
    # =========================================================================
    ws_config = wb.active
    ws_config.title = "config"
    style_header(ws_config, ["Key", "Value", "Mô tả"], [25, 40, 50])

    config_rows = [
        ("project_id", "", "Mã dự án (VD: MY-PROJECT-001)"),
        ("flow_bearer_token", "", "Bearer token từ Google Flow (bắt buộc)"),
        ("flow_project_id", "", "Project ID từ Google Flow (bắt buộc)"),
        ("topic", "story", "Chủ đề: story, documentary, education..."),
        ("style", "", "Phong cách ảnh (VD: cinematic, anime, realistic)"),
    ]

    for i, (key, value, desc) in enumerate(config_rows, 2):
        ws_config.cell(row=i, column=1, value=key).border = thin_border
        ws_config.cell(row=i, column=2, value=value).border = thin_border
        cell_desc = ws_config.cell(row=i, column=3, value=desc)
        cell_desc.font = guide_font
        cell_desc.fill = guide_fill
        cell_desc.border = thin_border

    # Hướng dẫn
    guide_row = len(config_rows) + 3
    ws_config.cell(row=guide_row, column=1, value="HƯỚNG DẪN:").font = Font(bold=True, color="E74C3C")
    ws_config.cell(row=guide_row + 1, column=1,
                   value="1. flow_bearer_token: Mở labs.google/fx → F12 → Network → tìm request có 'Bearer' → copy token")
    ws_config.cell(row=guide_row + 2, column=1,
                   value="2. flow_project_id: Trong cùng request, tìm URL có '/projects/XXX/' → XXX là project_id")
    ws_config.cell(row=guide_row + 3, column=1,
                   value="3. project_id: Đặt tên tùy ý cho dự án (dùng làm tên thư mục output)")

    # =========================================================================
    # Sheet 2: CHARACTERS
    # =========================================================================
    ws_chars = wb.create_sheet("characters")
    char_headers = ["id", "role", "name", "english_prompt", "vietnamese_prompt",
                    "character_lock", "image_file", "status", "is_child", "media_id"]
    char_widths = [8, 12, 15, 50, 30, 30, 15, 10, 10, 15]
    style_header(ws_chars, char_headers, char_widths)

    # Ví dụ nhân vật
    examples_char = [
        ("nv1", "main", "John", "A strong 35-year-old man with short dark hair, wearing a brown leather jacket and jeans, serious expression",
         "", "", "nv1.png", "pending", "False", ""),
        ("nv2", "supporting", "Mary", "A graceful 28-year-old woman with long red hair, wearing a blue dress, gentle smile",
         "", "", "nv2.png", "pending", "False", ""),
        ("loc1", "location", "Living Room", "A cozy modern living room with warm lighting, beige sofa, wooden floor, large windows with city view at night",
         "", "", "loc1.png", "pending", "False", ""),
    ]

    for i, row_data in enumerate(examples_char, 2):
        for j, val in enumerate(row_data, 1):
            cell = ws_chars.cell(row=i, column=j, value=val)
            cell.fill = example_fill
            cell.border = thin_border

    # Hướng dẫn
    guide_row = len(examples_char) + 3
    ws_chars.cell(row=guide_row, column=1, value="HƯỚNG DẪN:").font = Font(bold=True, color="E74C3C")
    ws_chars.cell(row=guide_row + 1, column=1,
                  value="- id: nv1, nv2... cho nhân vật | loc1, loc2... cho địa điểm")
    ws_chars.cell(row=guide_row + 2, column=1,
                  value="- role: main (chính), supporting (phụ), location (địa điểm)")
    ws_chars.cell(row=guide_row + 3, column=1,
                  value="- english_prompt: MÔ TẢ CHI TIẾT bằng tiếng Anh (quan trọng nhất!)")
    ws_chars.cell(row=guide_row + 4, column=1,
                  value="- is_child: True nếu là trẻ em (sẽ bỏ qua tạo ảnh reference)")
    ws_chars.cell(row=guide_row + 5, column=1,
                  value="- status: để 'pending' → tool sẽ tự tạo ảnh. Đổi 'skip' để bỏ qua")
    ws_chars.cell(row=guide_row + 6, column=1,
                  value="- Dòng màu xanh là VÍ DỤ, xóa đi rồi nhập data thật")

    # =========================================================================
    # Sheet 3: SCENES
    # =========================================================================
    ws_scenes = wb.create_sheet("scenes")
    scene_headers = ["scene_id", "srt_start", "srt_end", "duration", "planned_duration",
                     "srt_text", "img_prompt", "prompt_json", "video_prompt",
                     "img_path", "video_path", "status_img", "status_vid",
                     "characters_used", "location_used", "reference_files",
                     "media_id", "video_note", "segment_id"]
    scene_widths = [10, 12, 12, 10, 10, 30, 60, 15, 30, 15, 15, 10, 10, 20, 15, 25, 15, 12, 10]
    style_header(ws_scenes, scene_headers, scene_widths)

    # Ví dụ scenes
    examples_scene = [
        (1, "0:00:00,000", "0:00:05,000", 5.0, 5.0,
         "Once upon a time...",
         "A man in a brown leather jacket (nv1.png) stands at the entrance of a cozy living room (loc1.png), looking thoughtful, cinematic lighting",
         "", "", "", "", "pending", "", "nv1", "loc1", '["nv1.png", "loc1.png"]', "", "", 1),
        (2, "0:00:05,000", "0:00:10,000", 5.0, 5.0,
         "He walked to the window...",
         "A man in a brown leather jacket (nv1.png) walking towards a large window in a modern living room (loc1.png), warm evening light, wide shot",
         "", "", "", "", "pending", "", "nv1", "loc1", '["nv1.png", "loc1.png"]', "", "", 1),
        (3, "0:00:10,000", "0:00:15,000", 5.0, 5.0,
         "A woman appeared...",
         "A woman with long red hair in a blue dress (nv2.png) entering a cozy living room (loc1.png), gentle smile, soft lighting, medium shot",
         "", "", "", "", "pending", "", "nv1,nv2", "loc1", '["nv1.png", "nv2.png", "loc1.png"]', "", "", 1),
    ]

    for i, row_data in enumerate(examples_scene, 2):
        for j, val in enumerate(row_data, 1):
            cell = ws_scenes.cell(row=i, column=j, value=val)
            cell.fill = example_fill
            cell.border = thin_border

    # Hướng dẫn
    guide_row = len(examples_scene) + 3
    ws_scenes.cell(row=guide_row, column=1, value="HƯỚNG DẪN:").font = Font(bold=True, color="E74C3C")
    ws_scenes.cell(row=guide_row + 1, column=1,
                  value="- scene_id: Số thứ tự scene (1, 2, 3...)")
    ws_scenes.cell(row=guide_row + 2, column=1,
                  value="- img_prompt: MÔ TẢ CHI TIẾT bằng tiếng Anh, ghi rõ nhân vật (nv1.png) và địa điểm (loc1.png)")
    ws_scenes.cell(row=guide_row + 3, column=1,
                  value="- characters_used: ID nhân vật dùng trong scene (vd: nv1 hoặc nv1,nv2)")
    ws_scenes.cell(row=guide_row + 4, column=1,
                  value="- location_used: ID địa điểm (vd: loc1)")
    ws_scenes.cell(row=guide_row + 5, column=1,
                  value="- reference_files: JSON list file reference (vd: [\"nv1.png\", \"loc1.png\"])")
    ws_scenes.cell(row=guide_row + 6, column=1,
                  value="- status_img: để 'pending' → tool tạo ảnh. 'done' = đã xong. 'skip' = bỏ qua")
    ws_scenes.cell(row=guide_row + 7, column=1,
                  value="- Các cột khác (srt_start, duration...) không bắt buộc, có thể để trống")
    ws_scenes.cell(row=guide_row + 8, column=1,
                  value="- Dòng màu xanh là VÍ DỤ, xóa đi rồi nhập data thật")

    # Save
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"Template created: {output_path}")
    return output_path


if __name__ == "__main__":
    create_template()
