#!/usr/bin/env python3
"""
VE3 Tool - Worker VIDEO BASIC Mode
==================================
Tao anh + video CHI CHO SEGMENT DAU TIEN (tuan thu 8s).

- Tao Excel voi Progressive API
- Chi lay scenes cua segment dau tien
- Tao anh cho segment dau
- Tao video cho segment dau
- Tuan thu rule 8s (moi scene <= 8s)

Usage:
    python run_worker_video_basic.py                     (quet va xu ly tu dong)
    python run_worker_video_basic.py AR47-0028           (chay 1 project cu the)
"""

import sys
import os
import time
import shutil
from pathlib import Path

# Add current directory to path
TOOL_DIR = Path(__file__).parent
sys.path.insert(0, str(TOOL_DIR))

# Import tu run_worker (dung chung logic)
from run_worker import (
    detect_auto_path,
    POSSIBLE_AUTO_PATHS,
    get_channel_from_folder,
    matches_channel,
    is_project_complete_on_master,
    has_excel_with_prompts,
    needs_api_completion,
    copy_from_master,
    copy_to_visual,
    delete_local_project,
    SCAN_INTERVAL,
)

# Detect paths
AUTO_PATH = detect_auto_path()
if AUTO_PATH:
    MASTER_PROJECTS = AUTO_PATH / "ve3-tool-simple" / "PROJECTS"
    MASTER_VISUAL = AUTO_PATH / "VISUAL"
else:
    MASTER_PROJECTS = Path(r"\\tsclient\D\AUTO\ve3-tool-simple\PROJECTS")
    MASTER_VISUAL = Path(r"\\tsclient\D\AUTO\VISUAL")

LOCAL_PROJECTS = TOOL_DIR / "PROJECTS"
WORKER_CHANNEL = get_channel_from_folder()


def get_first_segment_scenes(excel_path: Path) -> list:
    """
    Lay danh sach scenes thuoc segment dau tien.
    Segment dau tien = segment co srt_range_start nho nhat.
    """
    try:
        from modules.excel_manager import PromptWorkbook

        wb = PromptWorkbook(str(excel_path))
        wb.load_or_create()

        # Lay story_segments
        segments = wb.get_story_segments()
        if not segments:
            return []

        # Tim segment dau tien (srt_range_start nho nhat)
        first_segment = min(segments, key=lambda s: s.get('srt_range_start', 999999))
        seg_start = first_segment.get('srt_range_start', 1)
        seg_end = first_segment.get('srt_range_end', seg_start + 10)
        seg_name = first_segment.get('segment_name', 'Segment 1')

        print(f"  [SEGMENT 1] {seg_name}: SRT {seg_start}-{seg_end}")

        # Lay scenes trong range nay
        scenes = wb.get_scenes()
        first_segment_scenes = []

        for scene in scenes:
            scene_srt_start = getattr(scene, 'srt_start', 0)
            if scene_srt_start is None:
                scene_srt_start = 0

            # Scene thuoc segment dau neu srt_start nam trong range
            if seg_start <= scene_srt_start <= seg_end:
                first_segment_scenes.append(scene)

        print(f"  [SEGMENT 1] Found {len(first_segment_scenes)} scenes")
        return first_segment_scenes

    except Exception as e:
        print(f"  Error getting first segment scenes: {e}")
        return []


def filter_excel_for_first_segment(excel_path: Path, code: str) -> bool:
    """
    Tao Excel moi chi chua scenes cua segment dau tien.
    Luu vao {code}_basic_prompts.xlsx
    """
    try:
        from modules.excel_manager import PromptWorkbook
        import openpyxl

        # Load Excel goc
        wb_orig = PromptWorkbook(str(excel_path))
        wb_orig.load_or_create()

        # Lay scenes segment dau
        first_scenes = get_first_segment_scenes(excel_path)
        if not first_scenes:
            print("  No scenes in first segment!")
            return False

        # Update status cua scenes khong thuoc segment 1 thanh "skip"
        first_scene_ids = set(str(s.scene_id) for s in first_scenes)

        # Load openpyxl de edit truc tiep
        wb_xl = openpyxl.load_workbook(str(excel_path))

        if 'scenes' in wb_xl.sheetnames:
            ws = wb_xl['scenes']

            # Tim cot scene_id va status
            headers = [cell.value for cell in ws[1]]
            id_col = headers.index('scene_id') + 1 if 'scene_id' in headers else None
            status_col = headers.index('status') + 1 if 'status' in headers else None

            if id_col and status_col:
                for row in range(2, ws.max_row + 1):
                    scene_id = str(ws.cell(row, id_col).value or '')
                    if scene_id and scene_id not in first_scene_ids:
                        # Skip scenes khong thuoc segment 1
                        ws.cell(row, status_col).value = 'skip_basic'

                wb_xl.save(str(excel_path))
                print(f"  Filtered Excel: {len(first_scene_ids)} scenes for segment 1")
                return True

        return False

    except Exception as e:
        print(f"  Error filtering Excel: {e}")
        return False


def is_local_video_basic_complete(project_dir: Path, name: str) -> bool:
    """Check if local project has videos for first segment."""
    img_dir = project_dir / "img"
    if not img_dir.exists():
        return False

    # Dem videos (.mp4)
    video_files = list(img_dir.glob("*.mp4"))

    if len(video_files) == 0:
        return False

    try:
        excel_path = project_dir / f"{name}_prompts.xlsx"
        if excel_path.exists():
            first_scenes = get_first_segment_scenes(excel_path)
            expected = len(first_scenes)

            # Count videos that match scene IDs
            first_scene_ids = set(str(s.scene_id) for s in first_scenes)
            matching_videos = [v for v in video_files if v.stem in first_scene_ids]

            if len(matching_videos) >= expected:
                print(f"    [{name}] Videos: {len(matching_videos)}/{expected} - COMPLETE")
                return True
            else:
                print(f"    [{name}] Videos: {len(matching_videos)}/{expected} - incomplete")
                return False
    except Exception as e:
        print(f"    [{name}] Warning: {e}")

    return len(video_files) > 0


def process_project_video_basic(code: str, callback=None) -> bool:
    """Process a single project - VIDEO BASIC mode (first segment only)."""

    def log(msg, level="INFO"):
        if callback:
            callback(msg, level)
        else:
            print(msg)

    log(f"\n{'='*60}")
    log(f"[VIDEO BASIC] Processing: {code}")
    log(f"{'='*60}")

    # Step 1: Check if already done on master
    if is_project_complete_on_master(code):
        log(f"  Already in VISUAL folder, skip!")
        return True

    # Step 2: Copy from master
    local_dir = copy_from_master(code)
    if not local_dir:
        return False

    # Step 3: Check/Create Excel
    excel_path = local_dir / f"{code}_prompts.xlsx"
    srt_path = local_dir / f"{code}.srt"

    if not excel_path.exists():
        if srt_path.exists():
            log(f"  No Excel found, creating with API...")
            # Import va tao Excel
            from run_worker_pic_basic import create_excel_with_api_basic
            if not create_excel_with_api_basic(local_dir, code, callback):
                log(f"  Failed to create Excel, skip!", "ERROR")
                return False
        else:
            log(f"  No Excel and no SRT, skip!")
            return False
    elif not has_excel_with_prompts(local_dir, code):
        log(f"  Excel empty/corrupt, recreating...")
        excel_path.unlink()
        from run_worker_pic_basic import create_excel_with_api_basic
        if not create_excel_with_api_basic(local_dir, code, callback):
            log(f"  Failed to recreate Excel, skip!", "ERROR")
            return False

    # Step 4: Filter Excel for first segment only
    log(f"  Filtering for FIRST SEGMENT only...")
    filter_excel_for_first_segment(excel_path, code)

    # Step 5: Create images + videos using SmartEngine
    try:
        from modules.smart_engine import SmartEngine

        engine = SmartEngine(
            worker_id=0,
            total_workers=1
        )

        log(f"  Excel: {excel_path.name}")
        log(f"  Mode: VIDEO BASIC (first segment, images + videos)")

        # Run engine - create images AND videos (skip_video=False)
        result = engine.run(str(excel_path), callback=callback, skip_compose=True, skip_video=False)

        if result.get('error'):
            log(f"  Error: {result.get('error')}", "ERROR")
            return False

    except Exception as e:
        log(f"  Exception: {e}", "ERROR")
        import traceback
        traceback.print_exc()
        return False

    # Step 6: Check completion
    if is_local_video_basic_complete(local_dir, code):
        log(f"  Videos complete for first segment!")
        return True
    else:
        log(f"  Videos incomplete", "WARN")
        return False


def scan_incomplete_local_projects() -> list:
    """Scan local PROJECTS for incomplete projects."""
    incomplete = []

    if not LOCAL_PROJECTS.exists():
        return incomplete

    for item in LOCAL_PROJECTS.iterdir():
        if not item.is_dir():
            continue

        code = item.name

        if not matches_channel(code):
            continue

        if is_project_complete_on_master(code):
            continue

        if is_local_video_basic_complete(item, code):
            continue

        srt_path = item / f"{code}.srt"
        if has_excel_with_prompts(item, code):
            print(f"    - {code}: incomplete (has Excel, no videos)")
            incomplete.append(code)
        elif srt_path.exists():
            print(f"    - {code}: has SRT, no Excel")
            incomplete.append(code)

    return sorted(incomplete)


def scan_master_projects() -> list:
    """Scan master PROJECTS folder for pending projects."""
    pending = []

    if not MASTER_PROJECTS.exists():
        return pending

    for item in MASTER_PROJECTS.iterdir():
        if not item.is_dir():
            continue

        code = item.name

        if not matches_channel(code):
            continue

        if is_project_complete_on_master(code):
            continue

        excel_path = item / f"{code}_prompts.xlsx"
        srt_path = item / f"{code}.srt"

        if has_excel_with_prompts(item, code):
            print(f"    - {code}: ready (has prompts)")
            pending.append(code)
        elif srt_path.exists():
            print(f"    - {code}: has SRT")
            pending.append(code)

    return sorted(pending)


def run_scan_loop():
    """Run continuous scan loop for VIDEO BASIC mode."""
    print(f"\n{'='*60}")
    print(f"  VE3 TOOL - WORKER VIDEO BASIC")
    print(f"{'='*60}")
    print(f"  Worker folder:   {TOOL_DIR.parent.name}")
    print(f"  Channel filter:  {WORKER_CHANNEL or 'ALL'}")
    print(f"  Mode:            VIDEO BASIC (first segment only)")
    print(f"  Output:          Images + Videos for segment 1")
    print(f"{'='*60}")

    cycle = 0

    while True:
        cycle += 1
        print(f"\n[VIDEO BASIC CYCLE {cycle}] Scanning...")

        incomplete_local = scan_incomplete_local_projects()
        pending_master = scan_master_projects()
        pending = list(dict.fromkeys(incomplete_local + pending_master))

        if not pending:
            print(f"  No pending projects")
            print(f"\n  Waiting {SCAN_INTERVAL}s... (Ctrl+C to stop)")
            try:
                time.sleep(SCAN_INTERVAL)
            except KeyboardInterrupt:
                print("\n\nStopped by user.")
                break
        else:
            print(f"  Found: {len(pending)} pending projects")

            for code in pending:
                try:
                    success = process_project_video_basic(code)
                    if not success:
                        print(f"  Skipping {code}, moving to next...")
                        continue
                except KeyboardInterrupt:
                    print("\n\nStopped by user.")
                    return
                except Exception as e:
                    print(f"  Error processing {code}: {e}")
                    continue

            print(f"\n  Processed all pending projects!")
            print(f"  Waiting {SCAN_INTERVAL}s... (Ctrl+C to stop)")
            try:
                time.sleep(SCAN_INTERVAL)
            except KeyboardInterrupt:
                print("\n\nStopped by user.")
                break


def main():
    import argparse
    parser = argparse.ArgumentParser(description='VE3 Worker VIDEO BASIC - First Segment Only')
    parser.add_argument('project', nargs='?', default=None, help='Project code')
    args = parser.parse_args()

    if args.project:
        process_project_video_basic(args.project)
    else:
        run_scan_loop()


if __name__ == "__main__":
    main()
