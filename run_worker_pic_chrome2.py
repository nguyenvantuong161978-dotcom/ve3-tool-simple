#!/usr/bin/env python3
"""
VE3 Tool - Worker PIC Chrome 2 (Image Generation với Chrome 2)
==============================================================
Script riêng cho Chrome 2, dùng DrissionFlowAPI TRỰC TIẾP (giống run_worker_video.py).
Được gọi từ run_worker_pic_basic_2.py như subprocess.

Chrome 2 chỉ tạo SCENES - không tạo characters/locations (Chrome 1 làm việc đó).

Usage:
    python run_worker_pic_chrome2.py --excel <path>
"""

import sys
import os
import time
from pathlib import Path

# Add current directory to path
TOOL_DIR = Path(__file__).parent
sys.path.insert(0, str(TOOL_DIR))


def load_chrome2_path() -> str:
    """Load chrome_portable_2 from settings.yaml."""
    import yaml

    settings_path = TOOL_DIR / "config" / "settings.yaml"
    if not settings_path.exists():
        return None

    with open(settings_path, 'r', encoding='utf-8') as f:
        settings = yaml.safe_load(f) or {}

    chrome2 = settings.get('chrome_portable_2', '')

    # Auto-detect if not configured
    if not chrome2:
        copy_chrome = TOOL_DIR / "GoogleChromePortable - Copy" / "GoogleChromePortable.exe"
        if copy_chrome.exists():
            chrome2 = str(copy_chrome)

    return chrome2


def get_scenes_need_images(excel_path: str) -> list:
    """Get list of scenes that need images (not characters/locations)."""
    from modules.excel_manager import PromptWorkbook

    scenes_need_images = []

    try:
        wb = PromptWorkbook(excel_path)
        scenes = wb.get_scenes()

        for scene in scenes:
            scene_id = str(scene.scene_id)

            # Skip characters (nvc, nv1, nv2...) and locations (loc1, loc2...)
            # Chrome 2 chỉ tạo scenes
            if scene_id.startswith('nv') or scene_id.startswith('loc'):
                continue

            # Check đã có ảnh chưa
            project_dir = Path(excel_path).parent
            img_dir = project_dir / "img"
            img_path = img_dir / f"{scene_id}.png"
            mp4_path = img_dir / f"{scene_id}.mp4"

            if img_path.exists() or mp4_path.exists():
                continue

            # Lấy prompt
            prompt = scene.img_prompt or ""
            if not prompt:
                continue

            scenes_need_images.append({
                'scene_id': scene_id,
                'prompt': prompt
            })

    except Exception as e:
        print(f"[Chrome2-PIC] Error reading Excel: {e}")

    return scenes_need_images


def run_chrome2_pic_worker(excel_path: str):
    """
    Chrome 2 worker for image generation.
    Uses DrissionFlowAPI directly (like run_worker_video.py does).
    """
    # Force UTF-8 encoding và flush output
    sys.stdout.reconfigure(encoding='utf-8', errors='replace', line_buffering=True)
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')

    print(f"[Chrome2-PIC] Starting worker...", flush=True)

    # Đợi Chrome 1 khởi động và tạo project trước
    # Chrome 1 cần thời gian để mở và tạo project URL
    wait_time = 15
    print(f"[Chrome2-PIC] Waiting {wait_time}s for Chrome 1 to start...", flush=True)
    time.sleep(wait_time)

    chrome2 = load_chrome2_path()
    if not chrome2:
        print(f"[Chrome2-PIC] ERROR: chrome_portable_2 not configured!", flush=True)
        return

    print(f"[Chrome2-PIC] Chrome: {chrome2}", flush=True)
    print(f"[Chrome2-PIC] Excel: {excel_path}", flush=True)

    # Safe print function
    def safe_print(msg):
        try:
            print(msg, flush=True)
        except UnicodeEncodeError:
            print(msg.encode('ascii', 'replace').decode('ascii'), flush=True)

    def log_callback(msg, level="INFO"):
        safe_print(f"[Chrome2-PIC] {msg}")

    # Get scenes that need images
    scenes = get_scenes_need_images(excel_path)

    if not scenes:
        safe_print(f"[Chrome2-PIC] No scenes need images")
        return

    # Worker distribution: Chrome 2 takes ODD scenes (1, 3, 5, 7...)
    # Chrome 1 takes EVEN scenes (2, 4, 6, 8...)
    # This way they don't collide
    my_scenes = []
    for idx, scene in enumerate(scenes):
        if idx % 2 == 1:  # Chrome 2 = odd indices
            my_scenes.append(scene)

    if not my_scenes:
        safe_print(f"[Chrome2-PIC] No scenes assigned to Chrome 2")
        return

    safe_print(f"[Chrome2-PIC] {len(my_scenes)} scenes to process (odd indices)")

    try:
        import yaml
        from modules.drission_flow_api import DrissionFlowAPI

        # Load config
        config = {}
        config_path = TOOL_DIR / "config" / "settings.yaml"
        if config_path.exists():
            with open(config_path, 'r', encoding='utf-8') as f:
                config = yaml.safe_load(f) or {}

        # Get project URL from Excel metadata
        project_url = None
        project_dir = Path(excel_path).parent

        # Method 1: Read from Excel sheet 'config'
        try:
            import openpyxl
            wb_xl = openpyxl.load_workbook(excel_path, data_only=True)

            if 'config' in wb_xl.sheetnames:
                ws = wb_xl['config']
            else:
                ws = wb_xl.active

            for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
                if not row:
                    continue

                for cell_val in row:
                    if cell_val and isinstance(cell_val, str):
                        cell_str = str(cell_val).strip()
                        if '/project/' in cell_str and cell_str.startswith('http'):
                            project_url = cell_str
                            break

                if project_url:
                    break

                if len(row) >= 2 and row[0]:
                    key = str(row[0]).strip().lower()
                    val = str(row[1] or '').strip()
                    if key == 'flow_project_url' and '/project/' in val:
                        project_url = val
                        break
                    elif key == 'flow_project_id' and val:
                        project_url = f"https://labs.google/fx/vi/tools/flow/project/{val}"
                        break

            wb_xl.close()
        except Exception as e:
            safe_print(f"[Chrome2-PIC] Error reading Excel config: {e}")

        # Method 2: Read from .media_cache.json
        if not project_url:
            cache_file = project_dir / ".media_cache.json"
            if cache_file.exists():
                try:
                    import json
                    with open(cache_file, 'r', encoding='utf-8') as f:
                        cache_data = json.load(f)
                    project_url = cache_data.get('_project_url', '')
                    if not project_url:
                        project_id = cache_data.get('_project_id', '')
                        if project_id:
                            project_url = f"https://labs.google/fx/vi/tools/flow/project/{project_id}"
                except:
                    pass

        # Poll for project URL if not found (Chrome 1 might still be creating it)
        if not project_url:
            safe_print(f"[Chrome2-PIC] No project URL found yet, waiting for Chrome 1...")
            max_wait = 120  # Wait up to 2 minutes
            waited = 0
            while waited < max_wait:
                time.sleep(10)
                waited += 10
                safe_print(f"[Chrome2-PIC] Waiting for project URL... ({waited}/{max_wait}s)")

                # Retry reading from cache
                cache_file = project_dir / ".media_cache.json"
                if cache_file.exists():
                    try:
                        import json
                        with open(cache_file, 'r', encoding='utf-8') as f:
                            cache_data = json.load(f)
                        project_url = cache_data.get('_project_url', '')
                        if not project_url:
                            project_id = cache_data.get('_project_id', '')
                            if project_id:
                                project_url = f"https://labs.google/fx/vi/tools/flow/project/{project_id}"
                        if project_url:
                            safe_print(f"[Chrome2-PIC] Found project URL from cache!")
                            break
                    except:
                        pass

            if not project_url:
                safe_print(f"[Chrome2-PIC] No project URL found after {max_wait}s! Giving up.")
                return

        safe_print(f"[Chrome2-PIC] Project URL: {project_url[:60]}...")

        # Webshare config
        ws_cfg = config.get('webshare_proxy', {})

        # Create DrissionFlowAPI - SEPARATE PROFILE for Chrome 2
        # Like run_worker_video.py uses "./chrome_profiles/video"
        api = DrissionFlowAPI(
            worker_id=101,  # Different worker_id (100+ range like video)
            headless=False,
            profile_dir="./chrome_profiles/pic2",  # SEPARATE profile for Chrome 2
            chrome_portable=chrome2,  # Use chrome_portable_2
            log_callback=log_callback,
            webshare_enabled=ws_cfg.get('enabled', False),
            machine_id=ws_cfg.get('machine_id', 1) + 100,  # Different machine_id
            total_workers=2  # Chrome 2 = right side
        )

        # Setup Chrome - go to project URL
        safe_print(f"[Chrome2-PIC] Setting up Chrome...")
        if not api.setup(project_url=project_url, skip_mode_selection=True):
            safe_print(f"[Chrome2-PIC] ERROR: Failed to setup Chrome!")
            return

        # Switch to image generation mode
        safe_print(f"[Chrome2-PIC] Switching to image mode...")
        if api.switch_to_image_mode():
            safe_print(f"[Chrome2-PIC] OK - Image mode selected")
        else:
            safe_print(f"[Chrome2-PIC] WARN - Could not switch to image mode")
        time.sleep(1)

        # Generate images
        img_dir = project_dir / "img"
        img_dir.mkdir(exist_ok=True)

        success_count = 0
        consecutive_failures = 0
        MAX_FAILURES = 5
        logout_retry_count = 0
        MAX_LOGOUT_RETRIES = 3

        for scene_info in my_scenes:
            scene_id = scene_info['scene_id']
            prompt = scene_info['prompt']

            # Check if already done
            png_path = img_dir / f"{scene_id}.png"
            if png_path.exists():
                safe_print(f"[Chrome2-PIC] {scene_id}: Already exists, skip")
                continue

            safe_print(f"\n[Chrome2-PIC] Creating image: {scene_id}")
            safe_print(f"   Prompt: {prompt[:60]}...")

            try:
                # Check logout TRƯỚC KHI generate
                if api._is_logged_out():
                    safe_print(f"[Chrome2-PIC] ⚠️ Phát hiện bị LOGOUT!")
                    if logout_retry_count < MAX_LOGOUT_RETRIES:
                        safe_print(f"[Chrome2-PIC] → Đang tự động đăng nhập lại...")
                        if api._auto_login_google():
                            safe_print(f"[Chrome2-PIC] ✓ Đăng nhập thành công!")
                            # Re-setup Chrome
                            api.close()
                            time.sleep(2)
                            if api.setup(project_url=project_url, skip_mode_selection=True):
                                api.switch_to_image_mode()
                                logout_retry_count += 1
                                safe_print(f"[Chrome2-PIC] ✓ Chrome đã sẵn sàng!")
                            else:
                                safe_print(f"[Chrome2-PIC] ✗ Setup lại thất bại!")
                                break
                        else:
                            safe_print(f"[Chrome2-PIC] ✗ Đăng nhập thất bại!")
                            break
                    else:
                        safe_print(f"[Chrome2-PIC] ✗ Đã thử đăng nhập {MAX_LOGOUT_RETRIES} lần, dừng!")
                        break

                # Generate image
                ok, images, error = api.generate_image(
                    prompt=prompt,
                    save_dir=img_dir,
                    filename=scene_id
                )

                if ok and images:
                    success_count += 1
                    consecutive_failures = 0
                    logout_retry_count = 0  # Reset logout counter on success
                    safe_print(f"[Chrome2-PIC] {scene_id}: SUCCESS")
                else:
                    consecutive_failures += 1
                    safe_print(f"[Chrome2-PIC] {scene_id}: FAILED - {error}")

                    # Check if error is due to logout
                    if error and ("logout" in error.lower() or "login" in error.lower() or "unauthorized" in error.lower()):
                        safe_print(f"[Chrome2-PIC] → Có vẻ bị logout, thử đăng nhập lại...")
                        if api._auto_login_google():
                            api.close()
                            time.sleep(2)
                            api.setup(project_url=project_url, skip_mode_selection=True)
                            api.switch_to_image_mode()
                            consecutive_failures = 0

                    elif consecutive_failures >= MAX_FAILURES:
                        safe_print(f"[Chrome2-PIC] {consecutive_failures} failures, restarting Chrome...")
                        try:
                            if api.restart_chrome():
                                consecutive_failures = 0
                        except:
                            pass

            except Exception as e:
                consecutive_failures += 1
                safe_print(f"[Chrome2-PIC] {scene_id}: ERROR - {e}")

        # Cleanup
        try:
            api.close()
        except:
            pass

        safe_print(f"\n[Chrome2-PIC] Done! Created {success_count}/{len(my_scenes)} images")

    except Exception as e:
        safe_print(f"[Chrome2-PIC] ERROR: {e}")
        import traceback
        traceback.print_exc()


def main():
    import argparse
    parser = argparse.ArgumentParser(description='VE3 Worker PIC Chrome 2')
    parser.add_argument('--excel', type=str, required=True, help='Excel path')
    args = parser.parse_args()

    run_chrome2_pic_worker(args.excel)


if __name__ == "__main__":
    main()
