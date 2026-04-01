#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VE3 Tool - Google Login Helper

Đọc thông tin tài khoản từ Google Sheet và đăng nhập vào Chrome.

v1.0.105: Đọc từ sheet THÔNG TIN (thay vì ve3)
- Cột B: Mã kênh (AR35, AR47, KA2...)
- Cột AT: Tài khoản Veo3 (nhiều dòng, format: id|pass|2fa)

Hỗ trợ xoay vòng tài khoản:
- Mỗi kênh có thể có 1-3 tài khoản
- Mỗi khi xong 1 mã, sẽ chuyển sang tài khoản tiếp theo
- Xoay vòng khi hết danh sách

Cách detect mã kênh:
- Từ đường dẫn: Documents\AR35-T1\ve3-tool-simple\ → kênh là AR35
"""

import sys
import os
import json
import time
import re
from pathlib import Path

TOOL_DIR = Path(__file__).parent
sys.path.insert(0, str(TOOL_DIR))


def get_proxy_arg_from_settings(ensure_ready: bool = True) -> str:
    """
    v1.0.572: Doc proxy arg tu settings.yaml.
    Dung cho login de dam bao Chrome login cung dung proxy.

    Khi ensure_ready=True (default):
    - Kiem tra port proxy da san sang chua (SOCKS5 dang chay)
    - Neu chua → tu start SOCKS5 proxy (lay IP tu pool/rotator)
    - Neu khong start duoc → tra ve "" (login khong qua proxy)

    Args:
        ensure_ready: True = tu dong start proxy neu chua chay

    Returns:
        Proxy arg string (vd: "socks5://127.0.0.1:1088") hoac "" neu khong co.
    """
    try:
        import yaml
        settings_path = TOOL_DIR / "config" / "settings.yaml"
        if not settings_path.exists():
            return ""
        with open(settings_path, 'r', encoding='utf-8') as f:
            cfg = yaml.safe_load(f) or {}

        # Check proxy_provider type
        pp_cfg = cfg.get('proxy_provider', {})
        pp_type = pp_cfg.get('type', 'none').lower()

        # Backward compat
        if not pp_cfg:
            ipv6_cfg = cfg.get('ipv6_rotation', {})
            if ipv6_cfg.get('enabled', False):
                pp_type = 'ipv6'

        # Pool API override
        mikrotik_cfg = cfg.get('mikrotik', {})
        pool_url = mikrotik_cfg.get('pool_api_url', '')
        if pool_url and pp_type in ('ipv6', 'none'):
            pp_type = 'ipv6_pool'

        if pp_type in ('ipv6', 'ipv6_pool'):
            # v1.0.613: VM mode dung IPv6 truc tiep + firewall block IPv4
            # ipv6_rotator.set_ipv6() da add IPv6 vao interface
            # Firewall block IPv4 cho Chrome → bat buoc dung IPv6
            # Chi server mode can SOCKS5 proxy (nhieu workers, nhieu IPv6)
            gen_mode = cfg.get('generation_mode', 'api')
            if gen_mode != 'server':
                # VM mode: dam bao IPv6 da tren interface
                if ensure_ready:
                    _ensure_ipv6_on_interface(cfg, pp_type)
                log(f"[PROXY] IPv6 DIRECT mode (firewall block IPv4, khong proxy)", "INFO")
                return ""

            # Server mode: van can SOCKS5 proxy
            port = cfg.get('ipv6_rotation', {}).get('local_proxy_port', 1088)
            proxy_arg = f"socks5://127.0.0.1:{port}"

            if ensure_ready and not _is_port_open(port):
                # Port chua mo → tu start proxy
                log(f"[PROXY] Port {port} chua san sang, dang khoi tao...", "INFO")
                if _ensure_socks5_proxy(cfg, pp_type, port):
                    log(f"[PROXY] SOCKS5 proxy started on port {port}", "INFO")
                else:
                    log(f"[PROXY] Khong start duoc proxy → login KHONG qua proxy", "WARN")
                    return ""

            return proxy_arg

        elif pp_type == 'webshare':
            ws_cfg = pp_cfg.get('webshare', {})
            host = ws_cfg.get('rotating_host', 'p.webshare.io')
            ws_port = ws_cfg.get('rotating_port', 80)
            user = ws_cfg.get('rotating_username', '')
            pwd = ws_cfg.get('rotating_password', '')
            if user and pwd:
                return f"http://{user}:{pwd}@{host}:{ws_port}"
        return ""
    except Exception as e:
        log(f"[PROXY] Error reading settings: {e}", "WARN")
        return ""


def _is_port_open(port: int) -> bool:
    """Check xem port co dang listen khong."""
    import socket
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
        s.settimeout(1)
        s.connect(('127.0.0.1', port))
        s.close()
        return True
    except Exception:
        return False


def _ensure_ipv6_on_interface(cfg: dict, pp_type: str) -> bool:
    """
    v1.0.612: Dam bao IPv6 da tren interface (VM mode, KHONG tao SOCKS5 proxy).
    ipv6_rotator.init_with_working_ipv6() se:
    - Pool mode: lay IP tu API → set_ipv6 (netsh add) → test
    - File mode: doc ipv6.txt → set_ipv6 (netsh add) → test
    """
    try:
        from modules.ipv6_rotator import get_ipv6_rotator
        rotator = get_ipv6_rotator(settings=cfg)

        if not rotator or not rotator.enabled:
            return False

        rotator.set_logger(log)

        # Check xem da co IPv6 hoat dong chua
        if rotator.current_ipv6:
            log(f"[PROXY] IPv6 da san sang: {rotator.current_ipv6}", "INFO")
            return True

        # Chua co → tim va set IPv6 len interface
        working_ipv6 = rotator.init_with_working_ipv6()
        if working_ipv6:
            log(f"[PROXY] IPv6 active: {working_ipv6}", "INFO")
            return True

        log("[PROXY] Khong tim duoc IPv6 hoat dong", "WARN")
        return False
    except Exception as e:
        log(f"[PROXY] Ensure IPv6 error: {e}", "WARN")
        return False


def _ensure_socks5_proxy(cfg: dict, pp_type: str, port: int) -> bool:
    """
    v1.0.574: Start SOCKS5 proxy neu chua chay.
    Dung ipv6_rotator de:
    1. Lay IPv6 tu pool API hoac file
    2. Add IPv6 vao VM interface (netsh)
    3. Start SOCKS5 proxy bind vao IPv6 do

    Returns:
        True neu proxy san sang.
    """
    try:
        from modules.ipv6_rotator import get_ipv6_rotator
        # v1.0.575: Truyen settings de rotator biet pool_api_url
        rotator = get_ipv6_rotator(settings=cfg)

        if not rotator or not rotator.enabled:
            log("[PROXY] IPv6 rotator not enabled", "WARN")
            return False

        rotator.set_logger(log)

        # init_with_working_ipv6() tu dong:
        # - Pool mode: lay IP tu API → set_ipv6 (netsh add) → test
        # - File mode: doc ipv6.txt → set_ipv6 (netsh add) → test
        working_ipv6 = rotator.init_with_working_ipv6()
        if not working_ipv6:
            log("[PROXY] Khong tim duoc IPv6 hoat dong", "WARN")
            return False

        log(f"[PROXY] IPv6 active: {working_ipv6}", "INFO")

        # Start SOCKS5 proxy bind vao IPv6 da add vao interface
        from modules.ipv6_proxy import start_ipv6_proxy
        proxy = start_ipv6_proxy(ipv6_address=working_ipv6, port=port, log_func=log)
        if proxy:
            import time as _time
            _time.sleep(1)
            return _is_port_open(port)
        return False

    except Exception as e:
        log(f"[PROXY] Ensure proxy error: {e}", "WARN")
        return False

CONFIG_FILE = TOOL_DIR / "config" / "config.json"
ACCOUNT_INDEX_FILE = TOOL_DIR / "config" / ".account_index.json"  # Track account rotation
SHEET_NAME = "THÔNG TIN"  # v1.0.105: Sheet mới chứa thông tin tài khoản
CHANNEL_COLUMN = "B"  # Cột B: Mã kênh (AR35, AR47, KA2...)
ACCOUNTS_COLUMN = "AT"  # Cột AT: Tài khoản Veo3 (nhiều dòng, format: id|pass|2fa)
IPV6_COLUMN = "AT"  # Cột chứa danh sách IPv6 (mỗi dòng trong ô = 1 IPv6)


def log(msg: str, level: str = "INFO"):
    """Print log with timestamp."""
    timestamp = time.strftime("%H:%M:%S")
    print(f"[{timestamp}] [{level}] {msg}")


def col_letter_to_index(col: str) -> int:
    """
    Convert column letter to 0-based index.
    A=0, B=1, ..., Z=25, AA=26, ..., AT=45
    """
    col = col.upper()
    result = 0
    for char in col:
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result - 1  # 0-based


def load_account_index(channel: str) -> int:
    """Load current account index for a channel from file."""
    try:
        if ACCOUNT_INDEX_FILE.exists():
            data = json.loads(ACCOUNT_INDEX_FILE.read_text(encoding="utf-8"))
            return data.get(channel, 0)
    except Exception as e:
        log(f"Error loading account index: {e}", "WARN")
    return 0


def save_account_index(channel: str, index: int):
    """Save current account index for a channel to file."""
    try:
        data = {}
        if ACCOUNT_INDEX_FILE.exists():
            data = json.loads(ACCOUNT_INDEX_FILE.read_text(encoding="utf-8"))
        data[channel] = index
        ACCOUNT_INDEX_FILE.parent.mkdir(parents=True, exist_ok=True)
        ACCOUNT_INDEX_FILE.write_text(json.dumps(data, indent=2), encoding="utf-8")
    except Exception as e:
        log(f"Error saving account index: {e}", "WARN")


def rotate_account_index(channel: str, total_accounts: int) -> int:
    """
    Rotate to next account index for a channel.
    Returns new index (wraps around if exceeds total).
    """
    current = load_account_index(channel)
    new_index = (current + 1) % total_accounts
    save_account_index(channel, new_index)
    log(f"Account rotated: {channel} -> index {new_index + 1}/{total_accounts}")
    return new_index


def parse_accounts_cell(cell_value: str) -> list:
    """
    Parse accounts from cell value.
    Format per line: id|pass|2fa

    Returns list of dicts: [{"id": "...", "password": "...", "totp_secret": "..."}, ...]
    """
    accounts = []
    if not cell_value:
        return accounts

    # Split by newlines (cell có thể có nhiều dòng)
    lines = cell_value.strip().split('\n')

    for line in lines:
        line = line.strip()
        if not line:
            continue

        parts = line.split('|')
        if len(parts) >= 2:
            account = {
                "id": parts[0].strip(),
                "password": parts[1].strip(),
                "totp_secret": parts[2].strip() if len(parts) >= 3 else ""
            }
            if account["id"] and account["password"]:
                accounts.append(account)

    return accounts


def get_channel_accounts(channel_code: str, max_retries: int = 3) -> list:
    """
    Lấy danh sách tài khoản Veo3 cho một kênh từ Google Sheet.

    v1.0.110: Tìm cả mã đầy đủ (AR8-T1) và mã channel (AR8)
    - Cột B: Mã máy đầy đủ (AR8-T1, KA2-T2...)
    - Cột AT: Tài khoản Veo3 (format: id|pass|2fa per line)

    Args:
        channel_code: Mã cần tìm (có thể là AR8-T1 hoặc AR8)
        max_retries: Số lần retry khi gặp lỗi

    Returns:
        List of account dicts: [{"id": "...", "password": "...", "totp_secret": "..."}, ...]
    """
    import time

    code_upper = channel_code.upper()
    last_error = None

    # Get column indices
    channel_col_idx = col_letter_to_index(CHANNEL_COLUMN)  # B = 1
    accounts_col_idx = col_letter_to_index(ACCOUNTS_COLUMN)  # AT = 45

    for attempt in range(max_retries):
        if attempt > 0:
            wait_time = 2 * attempt
            log(f"Retry đọc Google Sheet ({attempt + 1}/{max_retries}) sau {wait_time}s...")
            time.sleep(wait_time)

        gc, spreadsheet_name = load_gsheet_client()
        if not gc:
            last_error = "Cannot load gsheet client"
            continue

        try:
            ws = gc.open(spreadsheet_name).worksheet(SHEET_NAME)
            all_data = ws.get_all_values()

            if not all_data:
                log(f"Sheet '{SHEET_NAME}' is empty", "ERROR")
                return []

            # v1.0.110: Tìm theo nhiều cách:
            # 1. Khớp chính xác (AR8-T1 == AR8-T1)
            # 2. Khớp bắt đầu bằng (AR8-T1 starts with AR8)
            # 3. Khớp channel (AR8 == AR8)
            for row_idx, row in enumerate(all_data, start=1):
                if len(row) <= max(channel_col_idx, accounts_col_idx):
                    continue

                row_code = str(row[channel_col_idx]).strip().upper()

                # Kiểm tra khớp: chính xác HOẶC bắt đầu bằng
                is_match = (
                    row_code == code_upper or  # Khớp chính xác
                    row_code.startswith(code_upper + "-") or  # AR8 matches AR8-T1
                    code_upper.startswith(row_code + "-")  # AR8-T1 matches AR8
                )

                if is_match:
                    accounts_cell = str(row[accounts_col_idx]).strip()
                    accounts = parse_accounts_cell(accounts_cell)

                    if accounts:
                        log(f"Found {len(accounts)} accounts for {code_upper} (matched: {row_code})")
                        for i, acc in enumerate(accounts):
                            log(f"  Account {i+1}: {acc['id']} (2FA: {'Yes' if acc['totp_secret'] else 'No'})")
                        return accounts
                    else:
                        log(f"No valid accounts in cell AT for {row_code}", "WARN")
                        return []

            last_error = f"Code '{code_upper}' not found in sheet column B"
            if attempt < max_retries - 1:
                log(f"Không tìm thấy mã {code_upper}, thử lại...", "WARN")
                continue

        except Exception as e:
            last_error = str(e)
            log(f"Error reading sheet (attempt {attempt + 1}): {e}", "WARN")
            continue

    log(f"Code '{code_upper}' not found after {max_retries} attempts: {last_error}", "ERROR")
    return []


def get_channel_ipv6(channel_code: str, max_retries: int = 3) -> list:
    """
    Lấy danh sách IPv6 cho một kênh từ Google Sheet.

    - Sheet: THÔNG TIN
    - Cột B: Mã kênh (KA4-T3, AR8-T1...)
    - Cột IPV6_COLUMN: Danh sách IPv6 (mỗi dòng trong ô = 1 IPv6)

    Returns:
        List of IPv6 strings: ["2001:ee0:b004:3f01::2", ...]
    """
    import time

    code_upper = channel_code.upper()
    last_error = None

    channel_col_idx = col_letter_to_index(CHANNEL_COLUMN)  # B = 1
    ipv6_col_idx = col_letter_to_index(IPV6_COLUMN)

    for attempt in range(max_retries):
        if attempt > 0:
            wait_time = 2 * attempt
            log(f"Retry đọc IPv6 ({attempt + 1}/{max_retries}) sau {wait_time}s...")
            time.sleep(wait_time)

        gc, spreadsheet_name = load_gsheet_client()
        if not gc:
            last_error = "Cannot load gsheet client"
            continue

        try:
            ws = gc.open(spreadsheet_name).worksheet(SHEET_NAME)
            all_data = ws.get_all_values()

            if not all_data:
                log(f"Sheet '{SHEET_NAME}' is empty", "ERROR")
                return []

            log(f"[IPv6] Tìm mã '{code_upper}' trong {len(all_data)} dòng, cột B (idx={channel_col_idx}), cột IPv6 (idx={ipv6_col_idx})")

            # Debug: hiện vài mã đầu trong cột B
            sample_codes = []
            for r in all_data[:10]:
                if len(r) > channel_col_idx:
                    c = str(r[channel_col_idx]).strip()
                    if c:
                        sample_codes.append(c)
            if sample_codes:
                log(f"[IPv6] Mẫu cột B: {sample_codes}")

            for row_idx, row in enumerate(all_data, start=1):
                if len(row) <= max(channel_col_idx, ipv6_col_idx):
                    continue

                row_code = str(row[channel_col_idx]).strip().upper()

                # Match: chính xác hoặc prefix
                is_match = (
                    row_code == code_upper or
                    row_code.startswith(code_upper + "-") or
                    code_upper.startswith(row_code + "-")
                )

                if is_match:
                    ipv6_cell = str(row[ipv6_col_idx]).strip()
                    if not ipv6_cell:
                        log(f"IPv6 cell empty for {row_code} (row {row_idx})", "WARN")
                        return []

                    # Parse: mỗi dòng trong ô = 1 IPv6
                    ipv6_list = []
                    for line in ipv6_cell.split('\n'):
                        line = line.strip()
                        if line and ':' in line:  # IPv6 luôn có dấu ':'
                            ipv6_list.append(line)

                    log(f"Found {len(ipv6_list)} IPv6 for {code_upper} (matched: {row_code}, row {row_idx})")
                    for i, ip in enumerate(ipv6_list):
                        log(f"  IPv6 {i+1}: {ip}")
                    return ipv6_list

            last_error = f"Code '{code_upper}' not found"
            if attempt < max_retries - 1:
                continue

        except Exception as e:
            last_error = str(e)
            log(f"Error reading IPv6 (attempt {attempt + 1}): {e}", "WARN")
            continue

    log(f"IPv6 for '{code_upper}' not found: {last_error}", "ERROR")
    return []


def get_current_account_for_channel(channel_code: str, machine_code: str = None) -> dict:
    """
    Lấy tài khoản hiện tại cho một kênh (theo index rotation).

    v1.0.110: Dùng machine_code (mã đầy đủ) để tìm trong sheet,
    nhưng dùng channel_code để track rotation (tất cả máy cùng kênh dùng chung).

    Args:
        channel_code: Mã kênh (AR8, KA2) - dùng để track rotation
        machine_code: Mã máy đầy đủ (AR8-T1, KA2-T2) - dùng để tìm trong sheet
                      Nếu không có, dùng channel_code

    Returns:
        Account dict: {"id": "...", "password": "...", "totp_secret": "...", "index": N, "total": M}
        or None if no accounts found
    """
    # Dùng machine_code để tìm trong sheet, fallback to channel_code
    search_code = machine_code if machine_code else channel_code
    accounts = get_channel_accounts(search_code)

    if not accounts:
        return None

    # Dùng channel_code để track rotation
    current_index = load_account_index(channel_code)
    # Ensure index is within bounds
    if current_index >= len(accounts):
        current_index = 0
        save_account_index(channel_code, 0)

    account = accounts[current_index].copy()
    account["index"] = current_index
    account["total"] = len(accounts)

    log(f"Using account {current_index + 1}/{len(accounts)} for {channel_code}: {account['id']}")
    return account


def get_account_by_index(machine_code: str, index: int) -> dict:
    """
    v1.0.154: Lấy tài khoản theo index cố định (không rotate).
    Dùng khi cần login lại đúng account đã dùng ban đầu.

    Args:
        machine_code: Mã máy đầy đủ (KA2-T2, AR8-T1) - KHÔNG phải channel code!
        index: Index của account (0-based)

    Returns:
        Account dict: {"id": "...", "password": "...", "totp": "...", "index": N}
        or None if not found
    """
    accounts = get_channel_accounts(machine_code)
    if not accounts:
        log(f"No accounts found for machine: {machine_code}", "WARN")
        return None

    if index < 0 or index >= len(accounts):
        log(f"Invalid index {index} for machine {machine_code} (total: {len(accounts)})", "WARN")
        return None

    account = accounts[index].copy()
    account["index"] = index
    account["total"] = len(accounts)

    log(f"Get account by index {index + 1}/{len(accounts)} for {machine_code}: {account['id']}")
    return account


# ============================================================================
# EXCEL ACCOUNT TRACKING (v1.0.106)
# Lưu/đọc thông tin tài khoản trong Excel để resume đúng account
# ============================================================================

def save_project_account_json(project_dir, channel: str, account_index: int, account_email: str) -> bool:
    """
    v1.0.264: Lưu account vào .account.json trong thư mục project.
    File này độc lập với Excel - không bị ảnh hưởng khi Excel bị xóa/restore.
    """
    try:
        import json
        from pathlib import Path
        p = Path(project_dir) / ".account.json"
        data = {"channel": channel, "index": account_index, "email": account_email}
        p.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
        log(f"Saved account to .account.json: {account_email}")
        return True
    except Exception as e:
        log(f"Error saving .account.json: {e}", "WARN")
        return False


def get_project_account_json(project_dir) -> dict:
    """
    v1.0.264: Đọc account từ .account.json trong thư mục project.
    Ưu tiên hơn Excel vì không bị mất khi Excel restore/xóa.
    Returns {} nếu không tìm thấy.
    """
    try:
        import json
        from pathlib import Path
        p = Path(project_dir) / ".account.json"
        if p.exists():
            data = json.loads(p.read_text(encoding="utf-8"))
            if data.get("email"):
                return data
    except Exception as e:
        log(f"Error reading .account.json: {e}", "WARN")
    return {}


def save_account_to_excel(excel_path: str, channel: str, account_index: int, account_email: str) -> bool:
    """
    Lưu thông tin tài khoản đang sử dụng vào Excel config sheet.

    Args:
        excel_path: Đường dẫn đến file Excel
        channel: Mã kênh (VD: AR35, AR47)
        account_index: Index của tài khoản (0-based)
        account_email: Email của tài khoản

    Returns:
        True nếu lưu thành công
    """
    try:
        from openpyxl import load_workbook
        from pathlib import Path

        path = Path(excel_path)
        if not path.exists():
            log(f"Excel file not found: {excel_path}", "WARN")
            return False

        wb = load_workbook(str(path))

        # Tạo sheet config nếu chưa có
        if 'config' not in wb.sheetnames:
            ws = wb.create_sheet('config')
            ws['A1'] = 'key'
            ws['B1'] = 'value'
        else:
            ws = wb['config']

        # Các key cần lưu
        config_items = {
            'account_channel': channel,
            'account_index': str(account_index),
            'account_email': account_email,
        }

        for key, value in config_items.items():
            # Tìm row có key này để update
            found = False
            for row in range(2, ws.max_row + 1):
                cell_key = ws.cell(row=row, column=1).value
                if cell_key and str(cell_key).strip().lower() == key.lower():
                    ws.cell(row=row, column=2, value=value)
                    found = True
                    break

            # Nếu chưa có, thêm row mới
            if not found:
                next_row = ws.max_row + 1
                ws.cell(row=next_row, column=1, value=key)
                ws.cell(row=next_row, column=2, value=value)

        wb.save(str(path))
        log(f"Saved account info to Excel: {account_email} (index {account_index})")
        return True

    except Exception as e:
        log(f"Error saving account to Excel: {e}", "ERROR")
        return False


def get_account_from_excel(excel_path: str) -> dict:
    """
    Đọc thông tin tài khoản từ Excel config sheet.

    Args:
        excel_path: Đường dẫn đến file Excel

    Returns:
        {"channel": "AR35", "index": 0, "email": "xxx@gmail.com"} hoặc None nếu chưa có
    """
    try:
        from openpyxl import load_workbook
        from pathlib import Path

        path = Path(excel_path)
        if not path.exists():
            return None

        # data_only=True: đọc giá trị thực (không phải công thức)
        # Không dùng read_only=True vì ws.max_row có thể là None
        wb = load_workbook(str(path), data_only=True)

        if 'config' not in wb.sheetnames:
            return None

        ws = wb['config']

        result = {}
        # Dùng ws.iter_rows() thay vì range(max_row) để tránh lỗi max_row=None
        for row_cells in ws.iter_rows(min_row=2, values_only=True):
            cell_key = row_cells[0] if len(row_cells) > 0 else None
            cell_value = row_cells[1] if len(row_cells) > 1 else None

            if not cell_key:
                continue

            key = str(cell_key).strip().lower()
            value = str(cell_value) if cell_value is not None else ""

            if key == 'account_channel':
                result['channel'] = value
            elif key == 'account_index':
                try:
                    result['index'] = int(value)
                except:
                    result['index'] = 0
            elif key == 'account_email':
                result['email'] = value

        # v1.0.266: Chỉ cần channel + index (email optional - Excel cũ có thể thiếu)
        if 'channel' in result and 'index' in result:
            if 'email' not in result:
                result['email'] = ''  # Excel cũ không có email - caller sẽ lookup từ GSheet
            log(f"Read account from Excel: {result.get('email') or 'no-email'} (index {result['index']})")
            return result

        return None

    except Exception as e:
        log(f"Error reading account from Excel: {e}", "WARN")
        return None


def set_account_index_for_resume(excel_path: str, channel: str) -> bool:
    """
    Đọc thông tin account từ Excel và set vào index tracker để resume đúng account.

    Args:
        excel_path: Đường dẫn đến file Excel
        channel: Mã kênh hiện tại

    Returns:
        True nếu đã restore account index từ Excel
    """
    account_info = get_account_from_excel(excel_path)

    if not account_info:
        return False

    # Kiểm tra channel khớp
    if account_info.get('channel', '').upper() != channel.upper():
        log(f"Channel mismatch: Excel has {account_info.get('channel')}, current is {channel}", "WARN")
        return False

    # Set account index
    saved_index = account_info.get('index', 0)
    save_account_index(channel, saved_index)
    log(f"Restored account index {saved_index} for channel {channel} from Excel")
    return True


def detect_machine_code() -> str:
    """
    Detect mã máy từ đường dẫn thư mục tool.

    Ví dụ:
    - C:\\Users\\Admin\\Documents\\AR57-T1\\ve3-tool-simple → AR57-T1
    - D:\\VMs\\AR4-T1\\ve3-tool-simple → AR4-T1
    - C:\\Users\\hoangmai\\Documents\\AR4-T1\\ve3-tool-simple → AR4-T1
    """
    tool_path = TOOL_DIR.resolve()

    # Lấy thư mục cha của ve3-tool-simple
    parent = tool_path.parent

    # Mã máy thường có dạng: XX#-T# hoặc XX##-T# hoặc XX##-####
    # Ví dụ: AR4-T1, AR57-T1, AR47-0028
    # Pattern linh hoạt: 2 chữ cái + 1-3 số + dash + alphanumeric
    code_pattern = re.compile(r'^[A-Z]{2}\d{1,3}-[A-Z0-9]+$', re.IGNORECASE)

    # Kiểm tra parent folder
    if code_pattern.match(parent.name):
        return parent.name.upper()

    # Kiểm tra grandparent (Documents\AR57-T1\ve3-tool-simple)
    grandparent = parent.parent
    if code_pattern.match(grandparent.name):
        return grandparent.name.upper()

    # Thử tìm trong path
    for part in tool_path.parts:
        if code_pattern.match(part):
            return part.upper()

    return ""


def load_gsheet_client():
    """Load Google Sheet client."""
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        log("gspread not installed. Run: pip install gspread google-auth", "ERROR")
        return None, None

    if not CONFIG_FILE.exists():
        log(f"Config file not found: {CONFIG_FILE}", "ERROR")
        return None, None

    try:
        cfg = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))

        sa_path = (
            cfg.get("SERVICE_ACCOUNT_JSON") or
            cfg.get("service_account_json") or
            cfg.get("CREDENTIAL_PATH") or
            cfg.get("credential_path")
        )

        if not sa_path:
            log("Missing SERVICE_ACCOUNT_JSON in config", "ERROR")
            return None, None

        spreadsheet_name = cfg.get("SPREADSHEET_NAME")
        if not spreadsheet_name:
            log("Missing SPREADSHEET_NAME in config", "ERROR")
            return None, None

        scopes = [
            "https://www.googleapis.com/auth/spreadsheets.readonly",
            "https://www.googleapis.com/auth/drive.readonly",
        ]

        sa_file = Path(sa_path)
        if not sa_file.exists():
            sa_file = TOOL_DIR / "config" / sa_path

        if not sa_file.exists():
            log(f"Service account file not found: {sa_path}", "ERROR")
            return None, None

        creds = Credentials.from_service_account_file(str(sa_file), scopes=scopes)
        gc = gspread.authorize(creds)

        return gc, spreadsheet_name

    except Exception as e:
        log(f"Error loading gsheet client: {e}", "ERROR")
        return None, None


def extract_channel_from_machine_code(machine_code: str) -> str:
    """
    Extract channel code from machine code.
    Examples:
        AR35-T1 -> AR35
        AR47-T2 -> AR47
        KA2-T1 -> KA2
    """
    if "-T" in machine_code.upper():
        return machine_code.upper().split("-T")[0]
    if "-" in machine_code:
        return machine_code.rsplit("-", 1)[0].upper()
    return machine_code.upper()


def get_account_info(machine_code: str, max_retries: int = 3) -> dict:
    """
    Lấy thông tin tài khoản từ Google Sheet.

    v1.0.105: Đọc từ sheet THÔNG TIN theo mã kênh (không phải mã máy).
    - Cột B: Mã kênh (AR35, AR47, KA2...)
    - Cột AT: Tài khoản Veo3 (format: id|pass|2fa per line)

    Hỗ trợ xoay vòng tài khoản: mỗi kênh có thể có nhiều tài khoản,
    sẽ lần lượt sử dụng từng tài khoản.

    Args:
        machine_code: Mã máy (VD: AR35-T1) - sẽ extract thành mã kênh (AR35)
        max_retries: Số lần retry khi gặp lỗi

    Returns:
        {"id": "email@gmail.com", "password": "xxx", "totp_secret": "...", "index": N, "total": M}
        or None if not found
    """
    # Extract channel from machine code
    channel_code = extract_channel_from_machine_code(machine_code)
    log(f"Machine code: {machine_code} -> Channel: {channel_code}")

    # Get current account for this channel (with rotation support)
    account = get_current_account_for_channel(channel_code)

    if account:
        log(f"Account for {channel_code}: {account['id']} (index {account['index']+1}/{account['total']})")
        if account.get('totp_secret'):
            log(f"  -> 2FA secret found ({len(account['totp_secret'])} chars)")

    return account


def login_google_chrome(account_info: dict, chrome_portable: str = None, profile_dir: str = None, worker_id: int = 0, proxy_arg: str = "") -> bool:
    """
    Mở Chrome và đăng nhập Google bằng JavaScript.

    Do Google có nhiều biện pháp chống bot, script này sẽ:
    1. Mở Chrome đến trang đăng nhập
    2. Dùng JavaScript để điền email/password và trigger events
    3. Để user xác thực nếu cần

    Args:
        account_info: Dict với 'id' (email) và 'password'
        chrome_portable: Đường dẫn Chrome Portable cụ thể (nếu có).
                        Quan trọng khi có nhiều Chrome chạy song song.
        profile_dir: Đường dẫn profile Chrome cụ thể (nếu có).
                    Dùng cho Chrome 2 với profile riêng (ví dụ: pic2).
        worker_id: Worker ID để dùng port khác nhau cho mỗi Chrome.
        proxy_arg: Proxy argument cho Chrome (vd: "socks5://127.0.0.1:1088").
                   v1.0.571: Dam bao login cung dung proxy nhu khi tao anh.
    """
    try:
        from DrissionPage import ChromiumPage, ChromiumOptions
    except ImportError:
        log("DrissionPage not installed. Run: pip install DrissionPage", "ERROR")
        return False

    email = account_info["id"]
    password = account_info["password"]

    log(f"Opening Chrome for login: {email}")

    try:
        # Setup Chrome options
        options = ChromiumOptions()

        # === Port riêng cho mỗi Chrome (tránh conflict) ===
        # Chrome 1 (worker_id=0): port 9222
        # Chrome 2 (worker_id=1): port 9223
        base_port = 9222 + worker_id
        options.set_local_port(base_port)  # Dùng set_local_port như drission_flow_api.py
        log(f"Using port: {base_port} (worker_id={worker_id})")

        # Ưu tiên chrome_portable được truyền vào (cho Chrome 2 song song)
        chrome_exe = None
        if chrome_portable and Path(chrome_portable).exists():
            chrome_exe = chrome_portable
            log(f"Using specified Chrome: {chrome_exe}")
        else:
            # Fallback: Tìm Chrome Portable mặc định
            chrome_paths = [
                TOOL_DIR / "GoogleChromePortable" / "GoogleChromePortable.exe",
                Path.home() / "Documents" / "GoogleChromePortable" / "GoogleChromePortable.exe",
            ]
            for cp in chrome_paths:
                if cp.exists():
                    chrome_exe = str(cp)
                    break

        if chrome_exe:
            options.set_browser_path(chrome_exe)
            log(f"Using Chrome: {chrome_exe}")

            # === Profile: ưu tiên profile_dir được truyền vào ===
            if profile_dir and Path(profile_dir).exists():
                options.set_user_data_path(str(profile_dir))
                log(f"Using profile: {profile_dir}")
            else:
                # Fallback: dùng profile mặc định của Chrome Portable
                chrome_dir = Path(chrome_exe).parent
                for data_path in [chrome_dir / "Data" / "profile", chrome_dir / "User Data"]:
                    if data_path.exists():
                        options.set_user_data_path(str(data_path))
                        log(f"Using default profile: {data_path}")
                        break

        # v1.0.571: Proxy args - dam bao login dung cung proxy nhu tao anh
        if proxy_arg:
            options.set_argument(f'--proxy-server={proxy_arg}')
            options.set_argument('--proxy-bypass-list=<-loopback>')
            log(f"[NET] Login qua proxy: {proxy_arg}")
        else:
            # v1.0.619: Direct mode - block IPv4 TRUOC khi mo Chrome
            try:
                from modules.drission_flow_api import DrissionFlowAPI
                DrissionFlowAPI._block_ipv4_for_chrome_static(lambda msg: log(msg))
            except Exception as fw_err:
                log(f"[FW] Firewall error: {fw_err}")

        # Mở Chrome mới
        driver = ChromiumPage(options)

        # v1.0.650: Inject fingerprint NGAY SAU khi mo Chrome, TRUOC khi navigate
        # Dam bao login va tao anh dung CUNG fingerprint → Google khong detect thay doi
        try:
            from modules.fingerprint_data import get_unique_seed, build_fingerprint_js
            _fp_seed_file = Path(TOOL_DIR) / "config" / f".fingerprint_seed_{worker_id}"
            _fp_seed = get_unique_seed()
            _fp_js = build_fingerprint_js(_fp_seed)

            # CDP: inject cho TAT CA page loads (TRUOC khi scripts chay)
            try:
                result = driver.run_cdp('Page.addScriptToEvaluateOnNewDocument', source=_fp_js)
                log(f"[SPOOF] CDP fingerprint inject OK (seed={_fp_seed})")
            except Exception as _cdp_err:
                # Fallback: run_js cho page hien tai
                driver.run_js(_fp_js)
                log(f"[SPOOF] Fallback JS fingerprint inject (seed={_fp_seed})")

            # Save seed → file de DrissionFlowAPI doc va dung CUNG seed
            try:
                _fp_seed_file.parent.mkdir(parents=True, exist_ok=True)
                _fp_seed_file.write_text(str(_fp_seed))
                log(f"[SPOOF] Saved seed to {_fp_seed_file.name}")
            except Exception as _sf_err:
                log(f"[SPOOF] Save seed failed: {_sf_err}", "WARN")
        except ImportError:
            log("[SPOOF] fingerprint_data not available, skip fingerprint inject")
        except Exception as _fp_err:
            log(f"[SPOOF] Fingerprint inject error: {_fp_err}", "WARN")

        # Đi đến trang đăng nhập Google
        log("Navigating to Google login...")
        driver.get("https://accounts.google.com/signin")
        time.sleep(3)

        # Kiểm tra xem đã đăng nhập chưa
        if "myaccount.google.com" in driver.url or "google.com/search" in driver.url:
            log("Already logged in!", "OK")
            return True

        # === BƯỚC 1: ĐIỀN EMAIL ===
        log("Finding email input...")
        try:
            # Tìm input email
            email_input = driver.ele('#identifierId', timeout=5)
            if not email_input:
                email_input = driver.ele('input[type="email"]', timeout=3)

            if email_input:
                log(f"Found email input, filling: {email}")
                # Click để focus
                email_input.click()
                time.sleep(0.3)

                # Dùng JavaScript để set value và trigger events
                js_set_email = f'''
                    this.value = "{email}";
                    this.dispatchEvent(new Event('input', {{bubbles: true}}));
                    this.dispatchEvent(new Event('change', {{bubbles: true}}));
                '''
                email_input.run_js(js_set_email)
                log(f"Email filled via JS")
                time.sleep(0.5)

                # Nhấn Enter hoặc click Next
                log("Clicking Next button...")
                try:
                    next_btn = driver.ele('button:contains("Next")', timeout=2) or \
                               driver.ele('button:contains("Tiếp theo")', timeout=2) or \
                               driver.ele('button:contains("Tiếp tục")', timeout=2)
                    if next_btn:
                        next_btn.click()
                        log("Clicked Next button")
                    else:
                        # Fallback: nhấn Enter
                        email_input.input('\n')
                        log("Pressed Enter")
                except:
                    email_input.input('\n')
                    log("Pressed Enter (fallback)")

                # v1.0.623: Doi URL chuyen sang trang password thay vi sleep co dinh
                log("Waiting for password page...")
                for _url_wait in range(20):  # Max 20s cho mang cham
                    current_url = driver.url.lower()
                    if 'challenge/pwd' in current_url or 'challenge/password' in current_url:
                        log(f"Password page loaded (after {_url_wait+1}s)")
                        break
                    # Kiem tra neu bi reject hoac chuyen trang khac
                    if 'signin/rejected' in current_url or 'myaccount' in current_url:
                        log(f"Page redirected: {current_url}")
                        break
                    time.sleep(1)
                else:
                    log("Password page NOT detected after 20s, continuing anyway...", "WARN")
                    time.sleep(2)

                # v1.0.651: Cho page render day du truoc khi tuong tac
                # Du da detect password page, can cho JS/DOM on dinh
                log("Waiting 5s for page to fully render...")
                time.sleep(5)
            else:
                log("Email input not found!", "WARN")
        except Exception as e:
            log(f"Email step error: {e}", "WARN")

        # === BƯỚC 2: ĐIỀN PASSWORD (Ctrl+V) ===
        # v1.0.645: Fix password detection - Google đổi DOM, dùng JS detect + thêm selectors
        log("Waiting for password input...")
        try:
            pw_input = None

            # v1.0.645: Dùng JavaScript detect trước (nhanh + reliable hơn DrissionPage selectors)
            # Google có thể thay đổi attribute nhưng vẫn có input[type=password] trong DOM
            JS_FIND_PW = '''
                // Cách 1: type=password (chuẩn nhất)
                var el = document.querySelector('input[type="password"]');
                if (el) { el.focus(); el.click(); return true; }
                // Cách 2: name=Passwd (Google classic)
                el = document.querySelector('input[name="Passwd"]');
                if (el) { el.focus(); el.click(); return true; }
                // Cách 3: autocomplete=current-password
                el = document.querySelector('input[autocomplete="current-password"]');
                if (el) { el.focus(); el.click(); return true; }
                // Cách 4: aria-label chứa "password" (Google v3 mới)
                el = document.querySelector('input[aria-label*="assword"]');
                if (el) { el.focus(); el.click(); return true; }
                // Cách 5: div#password chứa input
                var div = document.querySelector('#password');
                if (div) { el = div.querySelector('input'); if (el) { el.focus(); el.click(); return true; } }
                // Cách 6: Tìm tất cả input, check type password (shadow DOM fallback)
                var inputs = document.querySelectorAll('input');
                for (var i = 0; i < inputs.length; i++) {
                    if (inputs[i].type === 'password') { inputs[i].focus(); inputs[i].click(); return true; }
                }
                return false;
            '''

            # v1.0.645: Loop 15s thực tế (không phải 70s như trước)
            pw_found_js = False
            for pw_wait in range(15):  # Max 15 giây
                try:
                    result = driver.run_js(JS_FIND_PW)
                    if result:
                        log(f"Password input found via JS (after {pw_wait+1}s)")
                        pw_found_js = True
                        break
                except Exception:
                    pass
                time.sleep(1)

            # Nếu JS không tìm thấy, thử DrissionPage selectors (fallback)
            if not pw_found_js:
                pw_selectors = [
                    'input[type="password"]',
                    'input[name="Passwd"]',
                    'input[autocomplete="current-password"]',
                    'input[aria-label*="assword"]',
                ]
                for sel in pw_selectors:
                    try:
                        pw_input = driver.ele(sel, timeout=1)
                        if pw_input:
                            log(f"Password input found: {sel} (DrissionPage fallback)")
                            break
                    except Exception:
                        pass

            if not pw_found_js and not pw_input:
                log("Password input NOT FOUND after 15s!", "WARN")
                # v1.0.645: Log URL hiện tại để debug
                try:
                    log(f"Current URL: {driver.url[:80]}...", "WARN")
                except Exception:
                    pass
                time.sleep(2)

            # Copy password vào clipboard
            try:
                import pyperclip
                pyperclip.copy(password)
                log("Password copied to clipboard")
            except ImportError:
                import subprocess
                subprocess.run(['clip'], input=password.encode(), check=True)
                log("Password copied to clipboard (via clip)")

            # Click vào password input nếu tìm thấy qua DrissionPage
            if pw_input and not pw_found_js:
                try:
                    pw_input.click()
                    time.sleep(0.5)
                except Exception:
                    pass

            # v1.0.645: Nếu JS đã focus, đợi thêm chút cho ổn định
            if pw_found_js:
                time.sleep(0.3)

            # Gửi Ctrl+V để paste
            from DrissionPage.common import Actions
            actions = Actions(driver)
            actions.key_down('ctrl').key_down('v').key_up('v').key_up('ctrl')
            log("Sent Ctrl+V")
            time.sleep(0.5)

            # Nhấn Enter để submit
            actions.key_down('enter').key_up('enter')
            log("Pressed Enter")
        except Exception as e:
            log(f"Password step error: {e}", "WARN")

        # === BƯỚC 2.5: XỬ LÝ TRANG "PHÊ DUYỆT THIẾT BỊ" (nếu có) ===
        # v1.0.649: Doi URL THAY DOI (khong chi check 'challenge')
        # Bug cu: URL `challenge/pwd` cung chua 'challenge' → break ngay (false positive)
        # Fix: Ghi nhan URL truoc khi Enter, doi URL KHAC hoac chua 'totp'/'az'/'ipp'/'myaccount'
        log("Waiting for password to be accepted...")
        _url_before_enter = driver.url.lower()
        _pwd_accepted = False
        for _2fa_wait in range(15):  # Tang tu 10 len 15s
            current = driver.url.lower()
            # URL da thay doi (khong con trang password)
            if current != _url_before_enter:
                log(f"Page transitioned (after {_2fa_wait+1}s)")
                _pwd_accepted = True
                break
            # URL chua doi nhung co the da chuyen sang challenge khac
            if 'myaccount' in current or 'signin/rejected' in current:
                log(f"Page transitioned (after {_2fa_wait+1}s)")
                _pwd_accepted = True
                break
            time.sleep(1)
        if not _pwd_accepted:
            log(f"Password may not have been accepted (URL unchanged after 15s)", "WARN")
            log(f"Current URL: {driver.url[:80]}...", "WARN")
        time.sleep(1)  # Them 1s cho page render
        try:
            # A) Kiểm tra 2FA option trước (Google Authenticator)
            auth_selectors_25 = [
                'text:Google Authenticator',
                'text:Ứng dụng xác thực',
                'text:Authenticator app',
                'text:Use your authenticator app',
                'text:Dùng ứng dụng xác thực',
            ]
            found_2fa = False
            for selector in auth_selectors_25:
                try:
                    auth_opt = driver.ele(selector, timeout=1)
                    if auth_opt:
                        log(f"Found 2FA option directly: {selector} - no need for 'Try another way'")
                        found_2fa = True
                        break
                except:
                    continue

            # B) Không có 2FA option → thử click "Try another way"
            if not found_2fa:
                try_another_selectors = [
                    'button:contains("Try another way")',
                    'button:contains("Thử cách khác")',
                    'text:Try another way',
                    'text:Thử cách khác',
                ]
                for selector in try_another_selectors:
                    try:
                        btn = driver.ele(selector, timeout=2)
                        if btn:
                            log(f"Found 'Try another way' button - clicking...")
                            btn.click()
                            time.sleep(3)
                            log("Clicked 'Try another way' successfully")
                            break
                    except:
                        continue
            else:
                log("2FA option available - skipping 'Try another way'")
        except Exception as e:
            log(f"Device approval check error: {e}", "WARN")

        # === BƯỚC 3: XỬ LÝ 2FA (nếu có) ===
        # v1.0.621: Doi OTP input xuat hien thay vi sleep co dinh
        totp_secret = account_info.get("totp_secret", "")
        if totp_secret:
            log(f"2FA secret found ({len(totp_secret)} chars)")

            try:
                import pyotp
                from DrissionPage.common import Actions
                actions = Actions(driver)

                # B1: Tim va click "Google Authenticator" option (neu can)
                auth_selectors = [
                    'text:Google Authenticator',
                    'text:Ứng dụng xác thực',
                    'text:Authenticator app',
                    'text:Use your authenticator app',
                    'text:Dùng ứng dụng xác thực',
                ]
                for selector in auth_selectors:
                    try:
                        auth_option = driver.ele(selector, timeout=1)
                        if auth_option:
                            log(f"Found 2FA option: {selector} - clicking...")
                            auth_option.click()
                            time.sleep(1)
                            break
                    except:
                        continue

                # B2: Doi OTP input xuat hien (max 15s cho mang cham)
                otp_input = None
                for otp_wait in range(15):
                    try:
                        # OTP input thuong la input[type="tel"] hoac input[type="text"]
                        otp_input = driver.ele('input[type="tel"]', timeout=1)
                        if not otp_input:
                            otp_input = driver.ele('#totpPin', timeout=1)
                        if otp_input:
                            log(f"OTP input ready (after {otp_wait+1}s)")
                            break
                    except Exception:
                        pass

                if not otp_input:
                    log("OTP input NOT FOUND after 15s, trying blind paste...", "WARN")
                else:
                    # v1.0.651: Cho page render day du truoc khi nhap OTP
                    log("Waiting 5s for 2FA page to fully render...")
                    time.sleep(5)

                # B3: Generate OTP NGAY TRUOC KHI nhap (tranh het han)
                clean_secret = totp_secret.replace(" ", "").replace("-", "").upper()
                totp = pyotp.TOTP(clean_secret)
                otp_code = totp.now()
                log(f"Generated OTP: {otp_code}")

                try:
                    import pyperclip
                    pyperclip.copy(otp_code)
                    log("OTP copied to clipboard")
                except ImportError:
                    import subprocess
                    subprocess.run(['clip'], input=otp_code.encode(), check=True)
                    log("OTP copied to clipboard (via clip)")

                # B4: Click input, paste, enter
                if otp_input:
                    try:
                        otp_input.click()
                        time.sleep(0.3)
                    except Exception:
                        pass

                actions.key_down('ctrl').key_down('v').key_up('v').key_up('ctrl')
                log("Sent Ctrl+V")
                time.sleep(0.5)
                actions.key_down('enter').key_up('enter')
                log("Sent Enter")

            except ImportError:
                log("pyotp not installed! Run: pip install pyotp", "ERROR")
            except Exception as e:
                log(f"2FA step error: {e}", "WARN")

        # === KIỂM TRA LOGIN THỰC SỰ THÀNH CÔNG ===
        # v1.0.649: Tang thoi gian + fix logic check URL
        log("Waiting for login to complete...")
        time.sleep(3)

        # Kiểm tra URL sau khi login
        max_check = 15  # v1.0.649: Tang tu 10 len 15 (30s cho mang cham)
        login_success = False
        for check in range(max_check):
            current_url = driver.url.lower()
            log(f"Check {check+1}/{max_check}: URL = {current_url[:60]}...")

            # Login thành công nếu URL đã rời khỏi accounts.google.com
            if "myaccount.google.com" in current_url or \
               "google.com/search" in current_url or \
               "labs.google" in current_url:
                login_success = True
                log("Login SUCCESS - URL changed!", "OK")
                break

            # Vẫn ở accounts.google.com nhưng KHÔNG còn ở signin/challenge
            if "accounts.google.com" in current_url and \
               "/signin" not in current_url and \
               "/challenge" not in current_url:
                login_success = True
                log("Login SUCCESS - URL changed!", "OK")
                break

            time.sleep(2)

        if not login_success:
            log("Login FAILED - still on login page!", "ERROR")
            # Đóng Chrome
            try:
                driver.quit()
            except:
                pass
            return False

        # v1.0.655: VERIFY LOGIN - check dung tai khoan (giong server/chrome_session.py)
        # Navigate myaccount.google.com → doc email → so sanh voi email dang login
        log("Verifying login account via myaccount.google.com...")
        try:
            driver.get("https://myaccount.google.com")
            time.sleep(3)
            verify_url = (driver.url or '').lower()
            if 'accounts.google.com' in verify_url:
                log("VERIFY FAILED - redirect to login page! Chua dang nhap.", "ERROR")
                try:
                    driver.quit()
                except:
                    pass
                return False

            # Doc email tu page (giong server _check_current_account)
            logged_email = driver.run_js("""
                // Tim email trong page
                var els = document.querySelectorAll('[data-email]');
                if (els.length > 0) return els[0].getAttribute('data-email');

                // Fallback: tim text co @ trong header area
                var all = document.querySelectorAll('header *');
                for (var i = 0; i < all.length; i++) {
                    var t = all[i].textContent.trim();
                    if (t.indexOf('@') > 0 && t.indexOf('.') > 0 && t.length < 60) {
                        return t;
                    }
                }

                // Fallback 2: tim trong aria-label
                var btns = document.querySelectorAll('[aria-label*="@"]');
                if (btns.length > 0) {
                    var label = btns[0].getAttribute('aria-label');
                    var match = label.match(/[\\w.-]+@[\\w.-]+/);
                    if (match) return match[0];
                }

                return '';
            """)
            logged_email = str(logged_email or '').strip().lower()

            if logged_email:
                if logged_email == email.lower():
                    log(f"VERIFY OK - dung tai khoan: {logged_email}", "OK")
                else:
                    log(f"VERIFY WARN - login SAI tai khoan! Expected: {email}, Got: {logged_email}", "WARN")
                    log("Tiep tuc nhung co the gap loi sau...", "WARN")
            else:
                log("VERIFY OK - dang nhap roi nhung khong doc duoc email (non-critical)", "OK")
        except Exception as ve:
            log(f"Verify error (non-critical): {ve}", "WARN")

        # v1.0.158: Tối ưu tốc độ warm-up
        log("Warm-up Flow...")
        flow_url = "https://labs.google/fx/vi/tools/flow"
        try:
            driver.get(flow_url)
            time.sleep(2)  # v1.0.158: Giảm từ 3s

            click_success = False

            for attempt in range(20):  # v1.0.158: Giảm từ 30
                # Tìm button "add_2" - timeout ngắn
                try:
                    btn = driver.ele('tag:button@@text():add_2', timeout=1)
                    if btn:
                        log("[v] Page ready!")
                        click_success = True
                        break
                except:
                    pass

                # Thử click "Create" button
                try:
                    create_btn = driver.ele('tag:button@@text():Create', timeout=1)
                    if create_btn:
                        log("Click 'Create'...")
                        create_btn.click()
                        time.sleep(1)
                        continue
                except:
                    pass

                # Reload page mỗi 5 lần
                if attempt > 0 and attempt % 5 == 0:
                    log(f"Reload Flow ({attempt}/20)...")
                    driver.get(flow_url)
                    time.sleep(2)

                time.sleep(0.5)  # v1.0.158: Giảm từ 1s

            if click_success:
                log("Session warmed up!")
            else:
                log("Button not found", "WARN")

        except Exception as e:
            log(f"Warm up error: {e}", "WARN")

        # v1.0.655: Dong Chrome sau login - DrissionFlowAPI se mo Chrome moi voi cung profile
        # Session login da luu trong cookies/profile → Chrome moi van logged in
        log("Closing Chrome after login (session saved in profile)")
        try:
            driver.quit()
        except:
            pass

        return True

    except Exception as e:
        log(f"Login error: {e}", "ERROR")
        import traceback
        traceback.print_exc()
        return False


def main():
    print("=" * 60)
    print("  VE3 TOOL - GOOGLE LOGIN HELPER")
    print("=" * 60)

    # 1. Detect machine code
    machine_code = detect_machine_code()

    if not machine_code:
        log("Cannot detect machine code from path", "ERROR")
        log(f"Current path: {TOOL_DIR}")
        log("Expected pattern: XX#-T# hoặc XX##-T# (ví dụ: AR4-T1, AR57-T1)")

        # Cho user nhập manual
        machine_code = input("\nEnter machine code (e.g., AR57-T1): ").strip().upper()
        if not machine_code:
            log("No machine code provided", "ERROR")
            return 1

    log(f"Machine code: {machine_code}")

    # 2. Get account info from sheet
    log(f"Reading account info from sheet '{SHEET_NAME}'...")
    account_info = get_account_info(machine_code)

    if not account_info:
        log("Cannot get account info", "ERROR")
        return 1

    # 3. Login to Google
    log("Starting Google login...")
    success = login_google_chrome(account_info)

    if success:
        log("=" * 60)
        log("  LOGIN COMPLETED!")
        log("=" * 60)
        return 0
    else:
        log("Login failed", "ERROR")
        return 1


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print("\nCancelled by user")
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
