#!/usr/bin/env python3
"""
Simple Dashboard - Giao dien don gian
=====================================
- Bam Start de chay
- Xem tien do tuc thi
- Click vao project de xem chi tiet
- Nut HIEN/AN CMD de toggle CMD windows

Usage:
    pythonw vm_manager_gui.py   (an console)
    python vm_manager_gui.py    (co console)
"""

import sys
import os

# AN CONSOLE WINDOW KHI CHAY GUI
if sys.platform == "win32":
    try:
        import ctypes
        # Ẩn console window
        ctypes.windll.user32.ShowWindow(ctypes.windll.kernel32.GetConsoleWindow(), 0)
    except:
        pass

    if sys.stdout:
        try:
            sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        except:
            pass
    os.environ['PYTHONIOENCODING'] = 'utf-8'

import tkinter as tk
from tkinter import ttk
import threading
import time
from pathlib import Path
from typing import Dict, Optional

# PIL for thumbnails
try:
    from PIL import Image, ImageTk
    PIL_AVAILABLE = True
except:
    PIL_AVAILABLE = False

TOOL_DIR = Path(__file__).parent

# Import VM Manager
try:
    from vm_manager import VMManager
    VM_AVAILABLE = True
except:
    VM_AVAILABLE = False

# Import Central Logger
try:
    from modules.central_logger import get_recent_logs, tail_log, LOG_FILE, add_callback, remove_callback
    LOGGER_AVAILABLE = True
except:
    LOGGER_AVAILABLE = False

# Import Excel Status for detailed status checking
try:
    from modules.excel_status import check_project_status as check_excel_status, EXCEL_STEPS
    EXCEL_STATUS_AVAILABLE = True
except:
    EXCEL_STATUS_AVAILABLE = False


# ================================================================================
# SETTINGS WINDOW - Cau hinh va kiem tra tai nguyen
# ================================================================================

class SettingsWindow(tk.Toplevel):
    """Cua so cau hinh - kiem tra va thiet lap tai nguyen."""

    def __init__(self, parent):
        super().__init__(parent)
        self.title("CAU HINH - Settings")
        self.geometry("800x850")  # v1.0.109: Tăng kích thước
        self.configure(bg='#1a1a2e')
        self.resizable(True, True)
        self.minsize(750, 750)

        self._build()
        self._load_settings()
        self._check_resources()

    def _build(self):
        # Header
        header = tk.Frame(self, bg='#e94560', height=50)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text="CAU HINH HE THONG", bg='#e94560', fg='white',
                 font=("Arial", 14, "bold")).pack(pady=12)

        # v1.0.109: Scrollable main content
        container = tk.Frame(self, bg='#1a1a2e')
        container.pack(fill="both", expand=True, padx=10, pady=10)

        canvas = tk.Canvas(container, bg='#1a1a2e', highlightthickness=0)
        scrollbar = tk.Scrollbar(container, orient="vertical", command=canvas.yview)
        main_frame = tk.Frame(canvas, bg='#1a1a2e')

        main_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=main_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        # Mouse wheel scroll
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # === CHU DE NOI DUNG (Topic) ===
        topic_lf = tk.LabelFrame(main_frame, text=" CHU DE NOI DUNG (Topic) ", bg='#16213e', fg='#ff9f43',
                                 font=("Arial", 10, "bold"), padx=10, pady=8)
        topic_lf.pack(fill="x", pady=5)

        self.topic_var = tk.StringVar(value="story")

        topic_row = tk.Frame(topic_lf, bg='#16213e')
        topic_row.pack(fill="x", pady=4)
        tk.Radiobutton(topic_row, text="Story (Phim truyen)", variable=self.topic_var, value="story",
                       bg='#16213e', fg='#ffcc00', selectcolor='#0f3460', font=("Arial", 10)).pack(side="left", padx=15)
        tk.Radiobutton(topic_row, text="Psychology (Tam ly/Giao duc)", variable=self.topic_var, value="psychology",
                       bg='#16213e', fg='#00ff88', selectcolor='#0f3460', font=("Arial", 10)).pack(side="left", padx=15)
        tk.Radiobutton(topic_row, text="Finance History (Tai chinh/Lich su)", variable=self.topic_var, value="finance_history",
                       bg='#16213e', fg='#ffa502', selectcolor='#0f3460', font=("Arial", 10)).pack(side="left", padx=15)

        topic_row2 = tk.Frame(topic_lf, bg='#16213e')
        topic_row2.pack(fill="x", pady=4)
        tk.Radiobutton(topic_row2, text="Finance VN (Tai chinh - Phong cach Viet)", variable=self.topic_var, value="finance_history_vn",
                       bg='#16213e', fg='#ff6b6b', selectcolor='#0f3460', font=("Arial", 10)).pack(side="left", padx=15)

        topic_row3 = tk.Frame(topic_lf, bg='#16213e')
        topic_row3.pack(fill="x", pady=4)
        tk.Label(topic_row3, text="VIDEO ONLY (chi tao video, khong anh):", bg='#16213e', fg='#ff9ff3', font=("Arial", 9, "bold")).pack(side="left", padx=5)
        tk.Radiobutton(topic_row3, text="Tam ly Video", variable=self.topic_var, value="psychology_video",
                       bg='#16213e', fg='#ff9ff3', selectcolor='#0f3460', font=("Arial", 10)).pack(side="left", padx=15)
        tk.Radiobutton(topic_row3, text="Tai chinh Video", variable=self.topic_var, value="finance_video",
                       bg='#16213e', fg='#ff9ff3', selectcolor='#0f3460', font=("Arial", 10)).pack(side="left", padx=15)

        tk.Label(topic_lf, text="VIDEO ONLY: Chi tao 1 anh nhan vat + video T2V truc tiep (nhanh hon, khong can tao anh scene).",
                 bg='#16213e', fg='#666', font=("Arial", 8)).pack(anchor="w")

        # === CHE DO TAO ANH ===
        mode_lf = tk.LabelFrame(main_frame, text=" CHE DO TAO ANH (Mode) ", bg='#16213e', fg='#a29bfe',
                                font=("Arial", 10, "bold"), padx=10, pady=8)
        mode_lf.pack(fill="x", pady=5)

        self.mode_var = tk.StringVar(value="small")

        mode_row = tk.Frame(mode_lf, bg='#16213e')
        mode_row.pack(fill="x", pady=4)
        tk.Radiobutton(mode_row, text="Basic (it anh, nhanh hon)", variable=self.mode_var, value="basic",
                       bg='#16213e', fg='white', selectcolor='#0f3460', font=("Arial", 10)).pack(side="left", padx=15)
        tk.Radiobutton(mode_row, text="Small (mac dinh)", variable=self.mode_var, value="small",
                       bg='#16213e', fg='#ffcc00', selectcolor='#0f3460', font=("Arial", 10)).pack(side="left", padx=15)
        tk.Radiobutton(mode_row, text="Full (nhieu anh nhat)", variable=self.mode_var, value="full",
                       bg='#16213e', fg='#00ff88', selectcolor='#0f3460', font=("Arial", 10)).pack(side="left", padx=15)

        tk.Label(mode_lf, text="(Ap dung cho lan chay tiep theo - luu khi bam LUU CAU HINH)",
                 bg='#16213e', fg='#666', font=("Arial", 8)).pack(anchor="w")

        # === CHE DO TAO ANH - Generation Mode ===
        gen_lf = tk.LabelFrame(main_frame, text=" CHE DO GENERATION (API/Chrome) ", bg='#16213e', fg='#ff6b6b',
                                font=("Arial", 10, "bold"), padx=10, pady=8)
        gen_lf.pack(fill="x", pady=5)

        # Generation mode
        gen_row = tk.Frame(gen_lf, bg='#16213e')
        gen_row.pack(fill="x", pady=4)
        tk.Label(gen_row, text="Generation Mode:", bg='#16213e', fg='white',
                 font=("Arial", 10)).pack(side="left", padx=(0, 10))
        self.gen_mode_var = tk.StringVar(value="api")
        gen_modes = [("API (mac dinh)", "api"), ("Browser", "browser"), ("Chrome UI", "chrome"), ("Server", "server"), ("API + Server", "api+server")]
        for text, val in gen_modes:
            tk.Radiobutton(gen_row, text=text, variable=self.gen_mode_var, value=val,
                           bg='#16213e', fg='white', selectcolor='#0f3460',
                           font=("Arial", 10)).pack(side="left", padx=10)

        # Chrome model selection
        chrome_model_row = tk.Frame(gen_lf, bg='#16213e')
        chrome_model_row.pack(fill="x", pady=4)
        tk.Label(chrome_model_row, text="Chrome Image Model:", bg='#16213e', fg='white',
                 font=("Arial", 10)).pack(side="left", padx=(0, 10))
        self.chrome_model_var = tk.StringVar(value="0")
        chrome_models = [("Nano Banana Pro", "0"), ("Nano Banana 2", "1"), ("Imagen 4", "2")]
        for text, val in chrome_models:
            tk.Radiobutton(chrome_model_row, text=text, variable=self.chrome_model_var, value=val,
                           bg='#16213e', fg='white', selectcolor='#0f3460',
                           font=("Arial", 10)).pack(side="left", padx=10)

        tk.Label(gen_lf, text="Server: Gui anh qua server (khong dung Chrome local). API+Server: API truoc, 403 → tu dong chuyen Server.",
                 bg='#16213e', fg='#666', font=("Arial", 8)).pack(anchor="w")
        tk.Label(gen_lf, text="Chrome UI: Tao anh/video truc tiep qua Chrome (khong dung API). Model chi ap dung cho Chrome mode.",
                 bg='#16213e', fg='#666', font=("Arial", 8)).pack(anchor="w")

        # === LOCAL SERVER (dat o day de de quan sat) ===
        server_frame = tk.LabelFrame(main_frame, text=" LOCAL PROXY SERVER ", bg='#16213e', fg='#ffd93d',
                                      font=("Arial", 10, "bold"), padx=10, pady=10)
        server_frame.pack(fill="x", pady=5)

        tk.Label(server_frame, text="Gui anh qua server thay vi dung Chrome local. Nhieu VM cung nhap chung → tu dong chia tai.",
                 bg='#16213e', fg='#888', font=("Arial", 8), justify="left").pack(anchor="w")

        # v1.0.525: Bo checkbox - chi can nhap URL, mode Server/API+Server tu xu ly
        self.local_server_enabled_var = tk.BooleanVar(value=False)  # Giu var de backward compat

        # Server URLs (multi-line: moi dong 1 server)
        tk.Label(server_frame, text="Server URLs (moi dong 1 server):", bg='#16213e', fg='white',
                 font=("Arial", 10)).pack(anchor="w", pady=(8, 0))
        self.server_urls_text = tk.Text(server_frame, height=3, width=60,
                                         font=("Consolas", 10), bg='#0f3460', fg='white',
                                         insertbackground='white')
        self.server_urls_text.pack(fill="x", pady=2)
        tk.Label(server_frame, text="Vi du: http://192.168.1.100:5000",
                 bg='#16213e', fg='#666', font=("Arial", 8)).pack(anchor="w")

        # Status label
        self.server_status_lbl = tk.Label(server_frame, text="", bg='#16213e', fg='#888',
                                           font=("Consolas", 9))
        self.server_status_lbl.pack(anchor="w", pady=(5, 0))

        # Nut kiem tra ket noi
        tk.Button(server_frame, text="Kiem tra ket noi", command=self._check_server_connection,
                  bg='#0984e3', fg='white', font=("Arial", 9), relief="flat", padx=8).pack(anchor="w", pady=5)

        # === KIEM TRA TAI NGUYEN ===
        check_frame = tk.LabelFrame(main_frame, text=" KIEM TRA TAI NGUYEN ", bg='#16213e', fg='#00ff88',
                                    font=("Arial", 10, "bold"), padx=10, pady=10)
        check_frame.pack(fill="x", pady=5)

        self.resource_labels = {}
        resources = [
            ("chrome", "Chrome Portable", "Trinh duyet de tao anh/video"),
            ("api_key", "API Key (Gemini/DeepSeek)", "De tao noi dung Excel"),
            ("proxy_token", "Proxy API Token", "De tao video qua API"),
            ("projects", "Thu muc PROJECTS", "Chua cac du an"),
        ]

        for res_id, name, desc in resources:
            row = tk.Frame(check_frame, bg='#16213e')
            row.pack(fill="x", pady=3)

            # Status icon
            status_lbl = tk.Label(row, text="?", width=3, bg='#16213e', fg='#ffd93d',
                                  font=("Arial", 12, "bold"))
            status_lbl.pack(side="left", padx=5)
            self.resource_labels[res_id] = status_lbl

            # Name and desc
            tk.Label(row, text=name, width=25, bg='#16213e', fg='white',
                     font=("Arial", 10), anchor="w").pack(side="left", padx=5)
            tk.Label(row, text=desc, bg='#16213e', fg='#888',
                     font=("Arial", 9), anchor="w").pack(side="left", padx=5)

        # Refresh button
        tk.Button(check_frame, text="Kiem tra lai", command=self._check_resources,
                  bg='#0984e3', fg='white', font=("Arial", 9), relief="flat", padx=10).pack(pady=5)

        # === API KEYS ===
        api_frame = tk.LabelFrame(main_frame, text=" API KEYS ", bg='#16213e', fg='#ffd93d',
                                  font=("Arial", 10, "bold"), padx=10, pady=10)
        api_frame.pack(fill="x", pady=5)

        # DeepSeek API Key
        tk.Label(api_frame, text="DeepSeek API Key:", bg='#16213e', fg='white',
                 font=("Arial", 10)).pack(anchor="w")
        self.deepseek_var = tk.StringVar()
        deepseek_entry = tk.Entry(api_frame, textvariable=self.deepseek_var, width=60,
                                  font=("Consolas", 10), bg='#0f3460', fg='white',
                                  insertbackground='white')
        deepseek_entry.pack(fill="x", pady=2)

        # Gemini API Key
        tk.Label(api_frame, text="Gemini API Key:", bg='#16213e', fg='white',
                 font=("Arial", 10)).pack(anchor="w", pady=(10, 0))
        self.gemini_var = tk.StringVar()
        gemini_entry = tk.Entry(api_frame, textvariable=self.gemini_var, width=60,
                                font=("Consolas", 10), bg='#0f3460', fg='white',
                                insertbackground='white')
        gemini_entry.pack(fill="x", pady=2)

        # Proxy API Token
        tk.Label(api_frame, text="Proxy API Token (Video):", bg='#16213e', fg='white',
                 font=("Arial", 10)).pack(anchor="w", pady=(10, 0))
        self.proxy_token_var = tk.StringVar()
        proxy_entry = tk.Entry(api_frame, textvariable=self.proxy_token_var, width=60,
                               font=("Consolas", 10), bg='#0f3460', fg='white',
                               insertbackground='white')
        proxy_entry.pack(fill="x", pady=2)

        # === CHROME ===
        chrome_frame = tk.LabelFrame(main_frame, text=" CHROME PORTABLE ", bg='#16213e', fg='#00d9ff',
                                     font=("Arial", 10, "bold"), padx=10, pady=10)
        chrome_frame.pack(fill="x", pady=5)

        tk.Label(chrome_frame, text="Duong dan Chrome 1:", bg='#16213e', fg='white',
                 font=("Arial", 10)).pack(anchor="w")
        self.chrome1_var = tk.StringVar()
        chrome1_row = tk.Frame(chrome_frame, bg='#16213e')
        chrome1_row.pack(fill="x", pady=2)
        tk.Entry(chrome1_row, textvariable=self.chrome1_var, width=50,
                 font=("Consolas", 9), bg='#0f3460', fg='white',
                 insertbackground='white').pack(side="left", fill="x", expand=True)
        tk.Button(chrome1_row, text="Chon...", command=lambda: self._browse_chrome(1),
                  bg='#6c5ce7', fg='white', font=("Arial", 8), relief="flat").pack(side="left", padx=5)

        tk.Label(chrome_frame, text="Duong dan Chrome 2:", bg='#16213e', fg='white',
                 font=("Arial", 10)).pack(anchor="w", pady=(10, 0))
        self.chrome2_var = tk.StringVar()
        chrome2_row = tk.Frame(chrome_frame, bg='#16213e')
        chrome2_row.pack(fill="x", pady=2)
        tk.Entry(chrome2_row, textvariable=self.chrome2_var, width=50,
                 font=("Consolas", 9), bg='#0f3460', fg='white',
                 insertbackground='white').pack(side="left", fill="x", expand=True)
        tk.Button(chrome2_row, text="Chon...", command=lambda: self._browse_chrome(2),
                  bg='#6c5ce7', fg='white', font=("Arial", 8), relief="flat").pack(side="left", padx=5)

        # === VEO3 ACCOUNTS (v1.0.108) ===
        account_frame = tk.LabelFrame(main_frame, text=" TAI KHOAN VEO3 ", bg='#16213e', fg='#ff6b6b',
                                       font=("Arial", 10, "bold"), padx=10, pady=10)
        account_frame.pack(fill="x", pady=5)

        # Header info
        self.account_header = tk.Label(account_frame, text="", bg='#16213e', fg='#00ff88',
                                        font=("Consolas", 9), anchor="w")
        self.account_header.pack(fill="x")

        # Account listbox with scrollbar (v1.0.108: Có thể chọn để đổi thứ tự)
        list_frame = tk.Frame(account_frame, bg='#16213e')
        list_frame.pack(fill="x", pady=5)

        self.account_listbox = tk.Listbox(list_frame, height=5, width=50,
                                           font=("Consolas", 10), bg='#0f3460', fg='#00ff88',
                                           selectbackground='#e17055', selectforeground='white',
                                           activestyle='none')
        self.account_listbox.pack(side="left", fill="x", expand=True)

        scrollbar = tk.Scrollbar(list_frame, orient="vertical", command=self.account_listbox.yview)
        scrollbar.pack(side="right", fill="y")
        self.account_listbox.config(yscrollcommand=scrollbar.set)

        # Store accounts data for selection
        self._accounts_data = []

        # Buttons row
        account_btn_row = tk.Frame(account_frame, bg='#16213e')
        account_btn_row.pack(fill="x", pady=5)

        tk.Button(account_btn_row, text="Tai lai", command=self._load_account_info,
                  bg='#0984e3', fg='white', font=("Arial", 9), relief="flat", padx=8).pack(side="left", padx=3)

        tk.Button(account_btn_row, text="Xoay vong", command=self._rotate_account,
                  bg='#6c5ce7', fg='white', font=("Arial", 9), relief="flat", padx=8).pack(side="left", padx=3)

        tk.Button(account_btn_row, text="CHON TAI KHOAN NAY", command=self._select_account,
                  bg='#e17055', fg='white', font=("Arial", 9, "bold"), relief="flat", padx=10).pack(side="left", padx=3)

        # Load account info on init
        self.after(500, self._load_account_info)

        # === BUTTONS ===
        btn_frame = tk.Frame(self, bg='#1a1a2e')
        btn_frame.pack(fill="x", padx=10, pady=10)

        tk.Button(btn_frame, text="LUU CAU HINH", command=self._save_settings,
                  bg='#00ff88', fg='#1a1a2e', font=("Arial", 11, "bold"),
                  relief="flat", padx=20, pady=5).pack(side="left", padx=5)

        tk.Button(btn_frame, text="DONG", command=self.destroy,
                  bg='#e94560', fg='white', font=("Arial", 11, "bold"),
                  relief="flat", padx=20, pady=5).pack(side="right", padx=5)

    def _load_settings(self):
        """Load settings tu file."""
        try:
            import yaml
            settings_path = TOOL_DIR / "config" / "settings.yaml"
            if settings_path.exists():
                with open(settings_path, "r", encoding="utf-8") as f:
                    config = yaml.safe_load(f) or {}

                self.topic_var.set(config.get('topic', 'story'))
                self.mode_var.set(config.get('excel_mode', 'small'))
                self.gen_mode_var.set(config.get('generation_mode', 'api'))
                self.chrome_model_var.set(str(config.get('chrome_model_index', 0)))
                self.deepseek_var.set(config.get('deepseek_api_key', ''))
                gemini_keys = config.get('gemini_api_keys', [''])
                self.gemini_var.set(gemini_keys[0] if gemini_keys else '')
                self.proxy_token_var.set(config.get('proxy_api_token', ''))
                self.chrome1_var.set(config.get('chrome_portable', ''))
                self.chrome2_var.set(config.get('chrome_portable_2', ''))
                # Load server URLs: uu tien list, fallback single URL
                server_list = config.get('local_server_list', [])
                if server_list:
                    urls_text = '\n'.join(
                        s.get('url', s) if isinstance(s, dict) else str(s)
                        for s in server_list
                    )
                else:
                    urls_text = config.get('local_server_url', '')
                self.server_urls_text.delete('1.0', 'end')
                self.server_urls_text.insert('1.0', urls_text)
        except Exception as e:
            print(f"Error loading settings: {e}")

    def _save_settings(self):
        """Luu settings vao file."""
        try:
            import yaml
            settings_path = TOOL_DIR / "config" / "settings.yaml"

            # Doc config hien tai
            config = {}
            if settings_path.exists():
                with open(settings_path, "r", encoding="utf-8") as f:
                    config = yaml.safe_load(f) or {}

            # Update
            config['topic'] = self.topic_var.get()
            config['excel_mode'] = self.mode_var.get()
            config['video_mode'] = self.mode_var.get()
            config['generation_mode'] = self.gen_mode_var.get()
            config['chrome_model_index'] = int(self.chrome_model_var.get())
            config['deepseek_api_key'] = self.deepseek_var.get().strip()
            gemini_key = self.gemini_var.get().strip()
            config['gemini_api_keys'] = [gemini_key] if gemini_key else ['']
            config['proxy_api_token'] = self.proxy_token_var.get().strip()
            config['chrome_portable'] = self.chrome1_var.get().strip()
            config['chrome_portable_2'] = self.chrome2_var.get().strip()
            # v1.0.525: Parse server URLs - local_server_enabled tu bat khi co URL
            raw_urls = self.server_urls_text.get('1.0', 'end').strip()
            server_list = []
            first_url = ''
            for line in raw_urls.splitlines():
                url = line.strip()
                if url and url.startswith('http'):
                    server_list.append({'url': url, 'name': url, 'enabled': True})
                    if not first_url:
                        first_url = url
            config['local_server_list'] = server_list
            config['local_server_url'] = first_url  # backward compat
            config['local_server_enabled'] = bool(first_url)  # Tu bat khi co URL

            # Luu
            with open(settings_path, "w", encoding="utf-8") as f:
                yaml.safe_dump(config, f, default_flow_style=False, allow_unicode=True)

            # Thong bao
            from tkinter import messagebox
            messagebox.showinfo("Thanh cong", "Da luu cau hinh!")
            self._check_resources()

            # v1.0.307: Đồng bộ settings vào manager.settings.config (in-memory)
            # Tránh save_config() ghi đè key cũ khi Start workers
            if hasattr(self.master, 'manager') and hasattr(self.master.manager, 'settings'):
                self.master.manager.settings.config = config.copy()

            # Dong bo mode_var voi parent GUI
            if hasattr(self.master, 'mode_var'):
                self.master.mode_var.set(self.mode_var.get())
            if hasattr(self.master, 'mode_display_lbl'):
                self.master.mode_display_lbl.config(text=f"[{self.mode_var.get()}]")

        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Loi", f"Khong the luu: {e}")

    def _toggle_server_fields(self):
        """v1.0.525: Khong can toggle - text box luon enabled."""
        pass

    def _get_server_urls(self) -> list:
        """Lay danh sach server URLs tu text box."""
        raw = self.server_urls_text.get('1.0', 'end').strip()
        urls = []
        for line in raw.splitlines():
            url = line.strip()
            if url and url.startswith('http'):
                urls.append(url)
        return urls

    def _check_server_connection(self):
        """Kiem tra ket noi den tat ca servers."""
        urls = self._get_server_urls()
        if not urls:
            self.server_status_lbl.config(text="Chua nhap URL server!", fg='#e94560')
            return

        self.server_status_lbl.config(text=f"Dang kiem tra {len(urls)} server...", fg='#ffd93d')
        self.update_idletasks()

        import requests
        results = []
        for url in urls:
            try:
                resp = requests.get(f"{url.rstrip('/')}/api/status", timeout=5)
                if resp.status_code == 200:
                    data = resp.json()
                    chrome_ok = data.get('chrome_ready', False)
                    pending = data.get('pending_tasks', 0)
                    done = data.get('completed_tasks', 0)
                    if chrome_ok:
                        results.append(f"{url}: OK (q={pending}, done={done})")
                    else:
                        results.append(f"{url}: Chrome chua san sang")
                else:
                    results.append(f"{url}: HTTP {resp.status_code}")
            except requests.exceptions.ConnectionError:
                results.append(f"{url}: FAIL - khong ket noi")
            except Exception as e:
                results.append(f"{url}: FAIL - {e}")

        ok_count = sum(1 for r in results if ': OK' in r)
        color = '#00ff88' if ok_count == len(urls) else '#ffd93d' if ok_count > 0 else '#e94560'
        status = f"{ok_count}/{len(urls)} server OK"
        self.server_status_lbl.config(text=status, fg=color)

        # Show detail popup
        from tkinter import messagebox
        detail = '\n'.join(results)
        messagebox.showinfo("Server Status", f"{status}\n\n{detail}")

    def _load_account_info(self):
        """
        v1.0.109: Load va hien thi thong tin tai khoan Veo3 trong Listbox.
        Cho phep chon tai khoan de dat lam hien tai.
        """
        try:
            # Clear listbox first
            self.account_listbox.delete(0, "end")
            self._accounts_data = []
            self.account_header.config(text="Dang tai thong tin...")
            self.update_idletasks()  # Force UI update

            from google_login import (
                detect_machine_code, extract_channel_from_machine_code,
                get_channel_accounts, load_account_index
            )

            # Detect channel
            machine_code = detect_machine_code()
            if not machine_code:
                self.account_header.config(text=f"Loi: Khong detect duoc ma may!")
                self.account_listbox.insert("end", f"Path: {TOOL_DIR}")
                return

            channel = extract_channel_from_machine_code(machine_code)
            self._current_channel = channel
            self._current_machine_code = machine_code  # v1.0.110: Lưu machine_code
            self.account_header.config(text=f"Ma may: {machine_code} | Dang ket noi...")
            self.update_idletasks()

            # Get accounts (có thể mất thời gian do network)
            # v1.0.110: Dùng machine_code để tìm trong sheet (cột B có mã đầy đủ)
            try:
                accounts = get_channel_accounts(machine_code)
            except Exception as net_err:
                self.account_header.config(text=f"Kenh: {channel} | Loi mang: {net_err}")
                self.account_listbox.insert("end", "Khong ket noi duoc Google Sheet")
                self.account_listbox.insert("end", "Nhan 'Tai lai' de thu lai")
                return

            current_index = load_account_index(channel)

            if not accounts:
                self.account_header.config(text=f"Ma may: {machine_code} | Khong tim thay!")
                self.account_listbox.insert("end", f"Khong tim thay ma '{machine_code}' trong cot B")
                self.account_listbox.insert("end", "Kiem tra sheet THONG TIN")
                return

            # Update header
            self.account_header.config(
                text=f"Ma may: {machine_code} | Tong: {len(accounts)} TK | Dang dung: #{current_index + 1}"
            )

            # Store accounts data
            self._accounts_data = accounts

            # Populate listbox
            for i, acc in enumerate(accounts):
                marker = ">>>" if i == current_index else "   "
                has_2fa = " [2FA]" if acc.get('totp_secret') else ""
                self.account_listbox.insert("end", f"{marker} {i+1}. {acc['id']}{has_2fa}")

            # Select current account
            if 0 <= current_index < len(accounts):
                self.account_listbox.selection_set(current_index)
                self.account_listbox.see(current_index)

        except Exception as e:
            import traceback
            traceback.print_exc()
            try:
                self.account_header.config(text=f"Loi: {e}")
                self.account_listbox.insert("end", "Nhan 'Tai lai' de thu lai")
            except:
                pass  # Window may be closed

    def _select_account(self):
        """
        v1.0.108: Chon tai khoan duoc highlight trong listbox lam tai khoan hien tai.
        """
        try:
            from google_login import save_account_index
            from tkinter import messagebox

            # Get selected index
            selection = self.account_listbox.curselection()
            if not selection:
                messagebox.showwarning("Thong bao", "Vui long chon mot tai khoan trong danh sach!")
                return

            selected_idx = selection[0]

            if not self._accounts_data or selected_idx >= len(self._accounts_data):
                messagebox.showerror("Loi", "Du lieu tai khoan khong hop le!")
                return

            # Save selected index
            channel = getattr(self, '_current_channel', None)
            if not channel:
                messagebox.showerror("Loi", "Khong xac dinh duoc kenh!")
                return

            save_account_index(channel, selected_idx)
            selected_account = self._accounts_data[selected_idx]

            messagebox.showinfo(
                "Thanh cong",
                f"Da chon tai khoan #{selected_idx + 1}/{len(self._accounts_data)}\n\n"
                f"Email: {selected_account['id']}\n\n"
                f"Tai khoan nay se duoc su dung cho ma tiep theo."
            )

            # Reload to update display
            self._load_account_info()

        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Loi", f"Khong the chon tai khoan: {e}")

    def _rotate_account(self):
        """
        v1.0.105: Xoay sang tai khoan tiep theo.
        """
        try:
            from google_login import (
                detect_machine_code, extract_channel_from_machine_code,
                get_channel_accounts, rotate_account_index
            )
            from tkinter import messagebox

            machine_code = detect_machine_code()
            if not machine_code:
                messagebox.showerror("Loi", "Khong detect duoc ma may!")
                return

            channel = extract_channel_from_machine_code(machine_code)
            # v1.0.110: Dùng machine_code để tìm (cột B có mã đầy đủ)
            accounts = get_channel_accounts(machine_code)

            if not accounts:
                messagebox.showerror("Loi", f"Khong tim thay ma '{machine_code}' trong sheet!")
                return

            if len(accounts) == 1:
                messagebox.showinfo("Thong bao", f"Ma {machine_code} chi co 1 tai khoan")
                return

            # Rotate
            new_idx = rotate_account_index(channel, len(accounts))
            new_account = accounts[new_idx]

            messagebox.showinfo(
                "Thanh cong",
                f"Da chuyen sang tai khoan #{new_idx + 1}/{len(accounts)}\n\n"
                f"Email: {new_account['id']}\n\n"
                f"Luu y: Can khoi dong lai Chrome de dang nhap tai khoan moi."
            )

            # Reload account info
            self._load_account_info()

        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Loi", f"Khong the xoay tai khoan: {e}")

    def _check_resources(self):
        """Kiem tra cac tai nguyen."""
        # Chrome
        chrome1 = self.chrome1_var.get()
        chrome2 = self.chrome2_var.get()
        chrome_ok = (chrome1 and Path(chrome1).exists()) or (chrome2 and Path(chrome2).exists())
        self._set_status("chrome", chrome_ok)

        # API Keys
        deepseek = self.deepseek_var.get().strip()
        gemini = self.gemini_var.get().strip()
        api_ok = bool(deepseek) or bool(gemini)
        self._set_status("api_key", api_ok)

        # Proxy Token
        proxy_token = self.proxy_token_var.get().strip()
        self._set_status("proxy_token", bool(proxy_token))

        # Projects folder
        projects_dir = TOOL_DIR / "PROJECTS"
        self._set_status("projects", projects_dir.exists())

    def _set_status(self, res_id: str, ok: bool):
        """Set status icon."""
        if res_id in self.resource_labels:
            if ok:
                self.resource_labels[res_id].config(text="OK", fg='#00ff88')
            else:
                self.resource_labels[res_id].config(text="X", fg='#e94560')

    def _browse_chrome(self, num: int):
        """Chon file Chrome."""
        from tkinter import filedialog
        path = filedialog.askopenfilename(
            title=f"Chon Chrome Portable {num}",
            filetypes=[("Executable", "*.exe"), ("All files", "*.*")]
        )
        if path:
            if num == 1:
                self.chrome1_var.set(path)
            else:
                self.chrome2_var.set(path)
            self._check_resources()



# ================================================================================
# PROJECT DETAIL - Hiển thị chi tiết từng scene
# ================================================================================

def find_scene_image(img_folder: Path, scene_id: int) -> Path | None:
    """
    Tìm ảnh scene với nhiều format khác nhau:
    - scene_001.png (format chuẩn)
    - 1.png (format đơn giản)
    - scene_1.png (không padding)

    Tìm trong img/ trước, nếu không có thì tìm trong img_backup/
    (Ảnh được di chuyển vào backup sau khi tạo video)
    """
    if not img_folder or not img_folder.exists():
        return None

    # Các format có thể có
    formats = [
        f"scene_{scene_id:03d}.png",  # scene_001.png
        f"{scene_id}.png",             # 1.png
        f"scene_{scene_id}.png",       # scene_1.png
        f"scene_{scene_id:03d}.jpg",  # scene_001.jpg
        f"{scene_id}.jpg",             # 1.jpg
    ]

    # Tìm trong img/ trước
    for fmt in formats:
        path = img_folder / fmt
        if path.exists():
            return path

    # Nếu không có trong img/, tìm trong img_backup/
    backup_folder = img_folder.parent / "img_backup"
    if backup_folder.exists():
        for fmt in formats:
            path = backup_folder / fmt
            if path.exists():
                return path

    return None


def find_scene_video(vid_folder: Path, scene_id: int) -> Path | None:
    """
    Tìm video scene với nhiều format khác nhau.
    """
    if not vid_folder or not vid_folder.exists():
        return None

    candidates = [
        vid_folder / f"scene_{scene_id:03d}.mp4",
        vid_folder / f"{scene_id}.mp4",
        vid_folder / f"scene_{scene_id}.mp4",
    ]

    for path in candidates:
        if path.exists():
            return path
    return None


class ProjectDetail(tk.Toplevel):
    """Xem chi tiết project - Excel steps + từng scene."""

    def __init__(self, parent, code: str):
        super().__init__(parent)
        self.code = code
        self.title(f"{code} - Chi tiết")
        self.geometry("900x700")
        self.configure(bg='#1a1a2e')
        self._auto_refresh = True

        self._build()
        self._load()
        self._start_auto_refresh()

    def _start_auto_refresh(self):
        """Auto refresh mỗi 3 giây."""
        if self._auto_refresh and self.winfo_exists():
            self._load()
            self.after(3000, self._start_auto_refresh)

    def destroy(self):
        """Stop auto refresh khi đóng."""
        self._auto_refresh = False
        super().destroy()

    def _build(self):
        # Header
        tk.Label(
            self, text=f"PROJECT: {self.code}",
            font=("Arial", 16, "bold"),
            bg='#0f3460', fg='white',
            pady=15
        ).pack(fill="x")

        # Buttons
        btn_row = tk.Frame(self, bg='#1a1a2e')
        btn_row.pack(fill="x", padx=10, pady=5)

        tk.Button(btn_row, text="Mở thư mục", command=self._open_folder,
                  bg='#e94560', fg='white', font=("Arial", 10), relief="flat", padx=15).pack(side="left", padx=5)
        tk.Button(btn_row, text="Làm mới", command=self._load,
                  bg='#0f3460', fg='white', font=("Arial", 10), relief="flat", padx=15).pack(side="left", padx=5)

        # === EXCEL STEPS SECTION ===
        excel_frame = tk.LabelFrame(self, text=" EXCEL - 7 BƯỚC ", bg='#16213e', fg='white',
                                    font=("Arial", 10, "bold"), padx=10, pady=5)
        excel_frame.pack(fill="x", padx=10, pady=5)

        # Steps header
        steps_header = tk.Frame(excel_frame, bg='#0f3460')
        steps_header.pack(fill="x")
        for txt, w in [("Bước", 20), ("Trạng thái", 12), ("Thời gian", 15), ("Ghi chú", 30)]:
            tk.Label(steps_header, text=txt, width=w, bg='#0f3460', fg='white',
                     font=("Arial", 9, "bold")).pack(side="left", padx=2, pady=3)

        # Steps container
        self.steps_frame = tk.Frame(excel_frame, bg='#16213e')
        self.steps_frame.pack(fill="x")

        # === SCENES SECTION ===
        # Summary
        self.summary_var = tk.StringVar(value="Đang tải...")
        tk.Label(self, textvariable=self.summary_var, bg='#1a1a2e', fg='#00d9ff',
                 font=("Consolas", 12, "bold")).pack(pady=5)

        # Scenes list
        list_frame = tk.LabelFrame(self, text=" SCENES - ẢNH & VIDEO ", bg='#16213e', fg='white',
                                   font=("Arial", 10, "bold"), padx=5, pady=5)
        list_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        # Header
        header = tk.Frame(list_frame, bg='#0f3460')
        header.pack(fill="x")
        for txt, w in [("Scene", 6), ("Prompt", 35), ("Ảnh", 6), ("Video", 6), ("Status", 10)]:
            tk.Label(header, text=txt, width=w, bg='#0f3460', fg='white',
                     font=("Arial", 9, "bold")).pack(side="left", padx=2, pady=5)

        # Scrollable
        canvas = tk.Canvas(list_frame, bg='#16213e', highlightthickness=0)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        self.scenes_frame = tk.Frame(canvas, bg='#16213e')

        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        canvas.create_window((0, 0), window=self.scenes_frame, anchor="nw")
        self.scenes_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        # Mouse wheel
        canvas.bind_all("<MouseWheel>", lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), "units"))

    def _load(self):
        """Load Excel steps và scenes."""
        # Clear
        for w in self.steps_frame.winfo_children():
            w.destroy()
        for w in self.scenes_frame.winfo_children():
            w.destroy()

        project_dir = self._find_dir()
        if not project_dir:
            self.summary_var.set("Không tìm thấy project!")
            return

        excel_path = project_dir / f"{self.code}_prompts.xlsx"
        if not excel_path.exists():
            self.summary_var.set("Chưa có Excel!")
            self._show_no_excel_steps()
            return

        try:
            from modules.excel_manager import PromptWorkbook
            wb = PromptWorkbook(str(excel_path))
            wb.load_or_create()  # PHẢI gọi load trước khi dùng

            # Load Excel steps status
            self._load_excel_steps(wb)

            scenes = wb.get_scenes()

            if not scenes:
                self.summary_var.set("Excel chưa có scenes!")
                return

            img_folder = project_dir / "img"
            vid_folder = project_dir / "vid"

            img_ok = vid_ok = vid_need = 0

            for i, scene in enumerate(scenes):
                # Handle both Scene objects and dicts
                if hasattr(scene, 'scene_id'):
                    sid = scene.scene_id
                    prompt = (scene.img_prompt or "")[:35]
                    vid_enabled = scene.video_enabled if hasattr(scene, 'video_enabled') else False
                else:
                    sid = scene.get('scene_id', i + 1)
                    prompt = (scene.get('img_prompt') or "")[:35]
                    vid_enabled = scene.get('video_enabled', False)

                img_path = find_scene_image(img_folder, sid)
                vid_path = find_scene_video(vid_folder, sid)
                has_img = img_path is not None
                has_vid = vid_path is not None

                if has_img:
                    img_ok += 1
                if vid_enabled:
                    vid_need += 1
                    if has_vid:
                        vid_ok += 1

                # Row
                bg = '#1a1a2e' if i % 2 == 0 else '#16213e'
                row = tk.Frame(self.scenes_frame, bg=bg)
                row.pack(fill="x", pady=1)

                tk.Label(row, text=str(sid), width=4, bg=bg, fg='white', font=("Consolas", 9)).pack(side="left", padx=2)

                # Thumbnail (40x40)
                thumb_label = tk.Label(row, width=5, height=2, bg='#333', cursor="hand2")
                thumb_label.pack(side="left", padx=2)

                if has_img and img_path and PIL_AVAILABLE:
                    try:
                        img = Image.open(str(img_path))
                        img.thumbnail((40, 40))
                        photo = ImageTk.PhotoImage(img)
                        thumb_label.configure(image=photo, width=40, height=40)
                        thumb_label.image = photo  # Keep reference
                        thumb_label.bind("<Button-1>", lambda e, s=sid: self._open_img(s))
                    except:
                        thumb_label.configure(text="IMG", fg='#00ff88')
                        thumb_label.bind("<Button-1>", lambda e, s=sid: self._open_img(s))
                elif has_img:
                    thumb_label.configure(text="IMG", fg='#00ff88')
                    thumb_label.bind("<Button-1>", lambda e, s=sid: self._open_img(s))
                else:
                    thumb_label.configure(text="--", fg='#666')

                tk.Label(row, text=prompt, width=30, bg=bg, fg='#aaa', font=("Consolas", 8), anchor="w").pack(side="left", padx=2)

                # Video status
                if not vid_enabled:
                    vid_txt, vid_clr = "-", '#444'
                elif has_vid:
                    vid_txt, vid_clr = "VID", '#00ff88'
                else:
                    vid_txt, vid_clr = "...", '#666'

                vid_lbl = tk.Label(row, text=vid_txt, width=5, bg=bg, fg=vid_clr, font=("Arial", 9, "bold"), cursor="hand2")
                vid_lbl.pack(side="left", padx=2)
                if has_vid:
                    vid_lbl.bind("<Button-1>", lambda e, s=sid: self._open_vid(s))

                # Status
                if has_img and (has_vid or not vid_enabled):
                    status_txt, status_clr = "OK", '#00ff88'
                elif has_img:
                    status_txt, status_clr = "...", '#ffaa00'
                else:
                    status_txt, status_clr = "--", '#666'

                tk.Label(row, text=status_txt, width=4, bg=bg, fg=status_clr, font=("Arial", 9)).pack(side="left", padx=2)

            self.summary_var.set(f"Anh: {img_ok}/{len(scenes)} | Video: {vid_ok}/{vid_need}")

        except Exception as e:
            self.summary_var.set(f"Loi: {e}")

    def _load_excel_steps(self, wb):
        """Load và hiển thị trạng thái các bước Excel."""
        step_names = [
            ("step_1", "1. Story Analysis"),
            ("step_2", "2. Story Segments"),
            ("step_3", "3. Characters"),
            ("step_4", "4. Locations"),
            ("step_5", "5. Director Plan"),
            ("step_6", "6. Scene Planning"),
            ("step_7", "7. Scene Prompts")
        ]

        try:
            all_status = wb.get_all_step_status() if hasattr(wb, 'get_all_step_status') else []

            # Convert to dict
            status_dict = {}
            for s in all_status:
                if isinstance(s, dict):
                    status_dict[s.get('step_id', '')] = s

            for i, (step_id, step_name) in enumerate(step_names):
                bg = '#1a1a2e' if i % 2 == 0 else '#16213e'
                row = tk.Frame(self.steps_frame, bg=bg)
                row.pack(fill="x")

                step_status = status_dict.get(step_id, {})
                status = step_status.get('status', 'PENDING')
                last_updated = step_status.get('last_updated', '')
                notes_raw = step_status.get('notes', '') or ''

                # Extract duration from notes (format: "Xs - description")
                duration_txt = "--"
                notes_display = notes_raw[:30]
                if notes_raw and 's - ' in notes_raw:
                    parts = notes_raw.split('s - ', 1)
                    if parts[0].isdigit():
                        secs = int(parts[0])
                        if secs >= 60:
                            duration_txt = f"{secs//60}m{secs%60:02d}s"
                        else:
                            duration_txt = f"{secs}s"
                        notes_display = parts[1][:25] if len(parts) > 1 else ""

                # Status color and icon
                if status == 'COMPLETED':
                    status_txt = "OK"
                    status_clr = '#00ff88'
                elif status == 'IN_PROGRESS':
                    status_txt = "Dang chay"
                    status_clr = '#00d9ff'
                elif status == 'ERROR':
                    status_txt = "Loi"
                    status_clr = '#e94560'
                elif status == 'PARTIAL':
                    status_txt = "Mot phan"
                    status_clr = '#ffaa00'
                else:
                    status_txt = "Cho"
                    status_clr = '#666'

                tk.Label(row, text=step_name, width=20, bg=bg, fg='white',
                         font=("Consolas", 9), anchor="w").pack(side="left", padx=2)
                tk.Label(row, text=status_txt, width=10, bg=bg, fg=status_clr,
                         font=("Arial", 9, "bold")).pack(side="left", padx=2)
                tk.Label(row, text=duration_txt, width=8, bg=bg, fg='#ffcc00',
                         font=("Consolas", 9, "bold")).pack(side="left", padx=2)
                tk.Label(row, text=notes_display, width=28, bg=bg, fg='#888',
                         font=("Consolas", 8), anchor="w").pack(side="left", padx=2)

        except Exception as e:
            self._show_no_excel_steps()

    def _show_no_excel_steps(self):
        """Hiển thị steps khi chưa có Excel."""
        step_names = [
            "1. Story Analysis",
            "2. Story Segments",
            "3. Characters",
            "4. Locations",
            "5. Director Plan",
            "6. Scene Planning",
            "7. Scene Prompts"
        ]

        for i, name in enumerate(step_names):
            bg = '#1a1a2e' if i % 2 == 0 else '#16213e'
            row = tk.Frame(self.steps_frame, bg=bg)
            row.pack(fill="x")

            tk.Label(row, text=name, width=20, bg=bg, fg='#666',
                     font=("Consolas", 9), anchor="w").pack(side="left", padx=2)
            tk.Label(row, text="Cho", width=12, bg=bg, fg='#444',
                     font=("Arial", 9)).pack(side="left", padx=2)
            tk.Label(row, text="--", width=15, bg=bg, fg='#444',
                     font=("Consolas", 9)).pack(side="left", padx=2)
            tk.Label(row, text="", width=30, bg=bg, fg='#444',
                     font=("Consolas", 8)).pack(side="left", padx=2)

    def _find_dir(self):
        local = TOOL_DIR / "PROJECTS" / self.code
        if local.exists():
            return local
        for drive in ["Z:", "Y:", "X:"]:
            master = Path(f"{drive}/AUTO/ve3-tool-simple/PROJECTS/{self.code}")
            if master.exists():
                return master
        return None

    def _open_folder(self):
        d = self._find_dir()
        if d:
            os.startfile(str(d))

    def _open_img(self, sid):
        d = self._find_dir()
        if d:
            p = d / "img" / f"scene_{sid:03d}.png"
            if p.exists():
                os.startfile(str(p))

    def _open_vid(self, sid):
        d = self._find_dir()
        if d:
            p = d / "vid" / f"scene_{sid:03d}.mp4"
            if p.exists():
                os.startfile(str(p))


# ================================================================================
# PROJECT DETAIL - Xem chi tiết project + ảnh tham chiếu
# ================================================================================

class ProjectDetail(tk.Toplevel):
    """Popup hiển thị chi tiết project và ảnh tham chiếu."""

    def __init__(self, parent, code: str):
        super().__init__(parent)
        self.code = code
        self.title(f"Chi tiet: {code}")
        self.geometry("800x600")
        self.configure(bg='#1a1a2e')

        self._build()

    def _find_project_dir(self) -> Optional[Path]:
        """Tìm thư mục project."""
        local = TOOL_DIR / "PROJECTS" / self.code
        if local.exists():
            return local
        return None

    def _build(self):
        # Header
        header = tk.Frame(self, bg='#0f3460', height=40)
        header.pack(fill="x")
        tk.Label(header, text=f"Project: {self.code}", bg='#0f3460', fg='#00ff88',
                 font=("Arial", 14, "bold")).pack(side="left", padx=20, pady=8)

        # Main content - 2 panels
        main = tk.PanedWindow(self, orient=tk.HORIZONTAL, bg='#1a1a2e', sashwidth=4)
        main.pack(fill="both", expand=True, padx=5, pady=5)

        # Left - Info
        left = tk.Frame(main, bg='#16213e')
        main.add(left, width=300)

        info_frame = tk.LabelFrame(left, text=" THONG TIN ", bg='#16213e', fg='white',
                                   font=("Arial", 10, "bold"), padx=10, pady=10)
        info_frame.pack(fill="x", padx=5, pady=5)

        project_dir = self._find_project_dir()
        if project_dir:
            # Count images
            img_dir = project_dir / "img"
            img_count = len(list(img_dir.glob("*.png"))) + len(list(img_dir.glob("*.jpg"))) if img_dir.exists() else 0

            # Count videos
            vid_dir = project_dir / "vid"
            vid_count = len(list(vid_dir.glob("*.mp4"))) if vid_dir.exists() else 0

            # Count reference files
            ref_dir = project_dir / "NV"
            ref_count = len(list(ref_dir.glob("*.*"))) if ref_dir.exists() else 0

            info_text = f"""
Thu muc: {project_dir}

So anh: {img_count}
So video: {vid_count}
So anh tham chieu: {ref_count}
"""
            tk.Label(info_frame, text=info_text, bg='#16213e', fg='#c8d6e5',
                     font=("Consolas", 10), justify="left", anchor="w").pack(anchor="w")
        else:
            tk.Label(info_frame, text="Khong tim thay project", bg='#16213e', fg='#ff6b6b',
                     font=("Consolas", 10)).pack()

        # Right - Reference images
        right = tk.Frame(main, bg='#16213e')
        main.add(right, width=480)

        ref_frame = tk.LabelFrame(right, text=" ANH THAM CHIEU (NV/) ", bg='#16213e', fg='white',
                                  font=("Arial", 10, "bold"), padx=5, pady=5)
        ref_frame.pack(fill="both", expand=True, padx=5, pady=5)

        # Canvas with scrollbar for images
        canvas = tk.Canvas(ref_frame, bg='#1a1a2e', highlightthickness=0)
        scrollbar = ttk.Scrollbar(ref_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#1a1a2e')

        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        # Load reference images
        if project_dir:
            nv_dir = project_dir / "NV"
            if nv_dir.exists():
                self._load_reference_images(scrollable_frame, nv_dir)
            else:
                tk.Label(scrollable_frame, text="Khong co thu muc NV/", bg='#1a1a2e', fg='#666',
                         font=("Consolas", 10)).pack(pady=20)

    def _load_reference_images(self, parent, nv_dir: Path):
        """Load và hiển thị ảnh tham chiếu từ thư mục NV."""
        if not PIL_AVAILABLE:
            tk.Label(parent, text="Can cai PIL de xem anh\npip install Pillow", bg='#1a1a2e', fg='#ffd93d',
                     font=("Consolas", 10)).pack(pady=20)
            return

        # Get all image files
        image_files = []
        for ext in ['*.png', '*.jpg', '*.jpeg', '*.webp']:
            image_files.extend(nv_dir.glob(ext))
        image_files.sort()

        if not image_files:
            tk.Label(parent, text="Khong co anh trong NV/", bg='#1a1a2e', fg='#666',
                     font=("Consolas", 10)).pack(pady=20)
            return

        # Display images in grid (3 columns)
        cols = 3
        row_frame = None
        self.photo_refs = []  # Keep references to prevent garbage collection

        for i, img_path in enumerate(image_files[:30]):  # Limit to 30 images
            if i % cols == 0:
                row_frame = tk.Frame(parent, bg='#1a1a2e')
                row_frame.pack(fill="x", pady=2)

            try:
                img = Image.open(img_path)
                img.thumbnail((140, 100))  # Thumbnail size
                photo = ImageTk.PhotoImage(img)
                self.photo_refs.append(photo)  # Keep reference

                # Frame for each image
                img_frame = tk.Frame(row_frame, bg='#0f3460', padx=2, pady=2)
                img_frame.pack(side="left", padx=3, pady=3)

                # Image label
                lbl = tk.Label(img_frame, image=photo, bg='#0f3460')
                lbl.pack()

                # Filename
                name = img_path.stem[:15] + "..." if len(img_path.stem) > 15 else img_path.stem
                tk.Label(img_frame, text=name, bg='#0f3460', fg='#aaa',
                         font=("Consolas", 7)).pack()

                # Click to open full size
                lbl.bind("<Button-1>", lambda e, p=img_path: os.startfile(str(p)))
                lbl.config(cursor="hand2")

            except Exception as e:
                pass

        # Info label
        tk.Label(parent, text=f"Tong: {len(image_files)} anh (click de mo)", bg='#1a1a2e', fg='#666',
                 font=("Consolas", 9)).pack(pady=10)


# ================================================================================
# MAIN GUI - Dashboard với LOG
# ================================================================================

class SimpleGUI(tk.Tk):
    """Dashboard don gian - click project de xem chi tiet."""

    def __init__(self):
        super().__init__()
        self.title("VE3 Dashboard")
        self.geometry("700x780")
        self.configure(bg='#1a1a2e')
        self.minsize(600, 600)

        # Hide current CMD window if running from CMD
        self._hide_current_cmd_window()

        # Khoi tao manager ngay de hien projects
        self.manager = VMManager(num_chrome_workers=2)
        self.running = False
        self.mode_var = tk.StringVar(value="small")  # loaded from settings.yaml
        self.selected_project = None  # Project dang xem chi tiet
        self.scene_photo_refs = []  # Keep references for thumbnails
        self.windows_visible = True  # Mac dinh: hien Chrome
        self._known_chrome_hwnds = set()  # Track Chrome HWNDs de auto-arrange khi Chrome moi mo
        self._known_cmd_hwnds = set()    # Track CMD HWNDs de auto-arrange khi CMD moi mo

        self._build()
        self._load_mode_from_yaml()  # Load mode sau khi build
        self._load_projects_on_startup()  # Load projects ngay khi mo
        self.after(300, self._position_tool_window)  # Dat tool vao vi tri trai man hinh
        self._update_loop()

        # v1.0.346: Đăng ký GUI callbacks cho master commands
        # Master gửi RUN = ấn BẮT ĐẦU, STOP = ấn DỪNG (chạy trên Tkinter main thread)
        self.manager._gui_start_callback = lambda: self.after(0, self._start)
        self.manager._gui_stop_callback = lambda: self.after(0, self._stop)

        # v1.0.344: Start watchdog ngay khi mở tool (không cần ấn BẮT ĐẦU)
        # Để VM nghe lệnh từ master ngay lập tức
        self.manager.start_watchdog()

    def _load_mode_from_yaml(self):
        """Load excel_mode tu settings.yaml va cap nhat mode_var + display label."""
        try:
            import yaml
            p = TOOL_DIR / "config" / "settings.yaml"
            if p.exists():
                with open(p, 'r', encoding='utf-8') as f:
                    c = yaml.safe_load(f) or {}
                mode = c.get('excel_mode', 'small')
                self.mode_var.set(mode)
                if hasattr(self, 'mode_display_lbl'):
                    self.mode_display_lbl.config(text=f"[{mode}]")
        except Exception as e:
            print(f"[GUI] Load mode error: {e}")

    def _hide_current_cmd_window(self):
        """Hide the CMD window that launched this GUI."""
        try:
            import win32gui
            import win32con
            import ctypes

            # Get console window handle
            kernel32 = ctypes.windll.kernel32
            hwnd = kernel32.GetConsoleWindow()

            if hwnd:
                # Hide it
                win32gui.ShowWindow(hwnd, win32con.SW_HIDE)
                print("[GUI] Hidden launch CMD window")
        except Exception as e:
            print(f"[GUI] Could not hide CMD window: {e}")

    def _build(self):
        # === TITLE BAR (compact, 28px) ===
        title = tk.Frame(self, bg='#080818', height=28)
        title.pack(fill="x")
        title.pack_propagate(False)

        git_info = self._get_git_version()
        tk.Label(title, text=f"VE3 Tool  {git_info}", bg='#080818', fg='#555',
                 font=("Consolas", 9)).pack(side="left", padx=10, pady=4)

        # Worker status dots (center)
        self.worker_dots = {}
        dots_frame = tk.Frame(title, bg='#080818')
        dots_frame.pack(side="left", padx=20)
        for wid, dname in [("excel", "EXCEL"), ("chrome_1", "CHR1"), ("chrome_2", "CHR2")]:
            dot = tk.Label(dots_frame, text=f"■ {dname}", bg='#080818', fg='#555',
                          font=("Consolas", 9, "bold"))
            dot.pack(side="left", padx=8)
            self.worker_dots[wid] = dot

        # Current project (right)
        self.title_project_var = tk.StringVar(value="")
        tk.Label(title, textvariable=self.title_project_var, bg='#080818', fg='#00d9ff',
                 font=("Consolas", 9, "bold")).pack(side="right", padx=10)

        # === BUTTON BAR (2 dong) ===
        btns_outer = tk.Frame(self, bg='#0f3460')
        btns_outer.pack(fill="x")

        # -- Dong 1: BAT DAU | DUNG | RESET | AN CHROME | [mode] --
        btns1 = tk.Frame(btns_outer, bg='#0f3460', height=42)
        btns1.pack(fill="x")
        btns1.pack_propagate(False)

        self.start_btn = tk.Button(btns1, text="BAT DAU", command=self._start,
                                   bg='#00ff88', fg='#1a1a2e', font=("Arial", 11, "bold"),
                                   relief="flat", padx=15, pady=3)
        self.start_btn.pack(side="left", padx=(10, 4), pady=6)

        self.stop_btn = tk.Button(btns1, text="DUNG", command=self._stop,
                                  bg='#e94560', fg='white', font=("Arial", 11, "bold"),
                                  relief="flat", padx=15, pady=3)
        self.stop_btn.pack(side="left", padx=4, pady=6)

        self.reset_btn = tk.Button(btns1, text="RESET", command=self._reset_workers,
                                   bg='#ff6348', fg='white', font=("Arial", 11, "bold"),
                                   relief="flat", padx=12, pady=3)
        self.reset_btn.pack(side="left", padx=4, pady=6)

        self.mode_display_lbl = tk.Label(btns1, text="[small]", bg='#0f3460', fg='#ffcc00',
                                         font=("Consolas", 9))
        self.mode_display_lbl.pack(side="left", padx=8)

        self.status_var = tk.StringVar(value="San sang")
        tk.Label(btns1, textvariable=self.status_var, bg='#0f3460', fg='#00d9ff',
                 font=("Consolas", 9, "bold")).pack(side="right", padx=10)

        # -- Dong 2: SETTINGS | UPDATE | SETUP VM | SAP LAI --
        btns2 = tk.Frame(btns_outer, bg='#0d2a4a', height=32)
        btns2.pack(fill="x")
        btns2.pack_propagate(False)

        self.settings_btn = tk.Button(btns2, text="SETTINGS", command=self._open_settings,
                                      bg='#ff9f43', fg='white', font=("Arial", 8, "bold"),
                                      relief="flat", padx=8)
        self.settings_btn.pack(side="left", padx=(10, 4), pady=4)

        self.update_btn = tk.Button(btns2, text="UPDATE", command=self._run_update,
                                    bg='#0984e3', fg='white', font=("Arial", 8, "bold"),
                                    relief="flat", padx=8)
        self.update_btn.pack(side="left", padx=4, pady=4)

        self.setup_vm_btn = tk.Button(btns2, text="SETUP VM", command=self._setup_vm,
                                      bg='#a29bfe', fg='white', font=("Arial", 8, "bold"),
                                      relief="flat", padx=8)
        self.setup_vm_btn.pack(side="left", padx=4, pady=4)

        tk.Button(btns2, text="SAP LAI", command=self._arrange_windows,
                  bg='#00cec9', fg='white', font=("Arial", 8, "bold"),
                  relief="flat", padx=8).pack(side="left", padx=4, pady=4)

        # Server button
        self._server_process = None
        self.server_btn = tk.Button(btns2, text="CHAY SERVER", command=self._toggle_server,
                                     bg='#fd79a8', fg='white', font=("Arial", 8, "bold"),
                                     relief="flat", padx=8)
        self.server_btn.pack(side="left", padx=4, pady=4)

        self.server_status_dot = tk.Label(btns2, text="", bg='#0d2a4a', fg='#888',
                                           font=("Consolas", 8))
        self.server_status_dot.pack(side="left", padx=2)

        # === WORKERS (3 rows) ===
        wf = tk.Frame(self, bg='#16213e', padx=8, pady=5)
        wf.pack(fill="x", padx=5, pady=(5, 0))

        self.worker_vars = {}
        self.worker_labels = {}
        self.worker_rows = {}
        self.worker_progress = {}

        for wid, dname in [("excel", "EXCEL"), ("chrome_1", "CHR 1"), ("chrome_2", "CHR 2")]:
            row = tk.Frame(wf, bg='#16213e')
            row.pack(fill="x", pady=1)

            name_lbl = tk.Label(row, text=dname, width=7, bg='#16213e', fg='#555',
                               font=("Consolas", 10, "bold"), anchor="w")
            name_lbl.pack(side="left", padx=(0, 4))

            badge = tk.Label(row, text="■ IDLE", width=11, bg='#1a1a2e', fg='#555',
                            font=("Consolas", 9, "bold"))
            badge.pack(side="left", padx=4)

            self.worker_vars[f"{wid}_project"] = tk.StringVar(value="")
            proj_lbl = tk.Label(row, textvariable=self.worker_vars[f"{wid}_project"],
                               bg='#16213e', fg='#00d9ff',
                               font=("Consolas", 10, "bold"), width=12, anchor="w")
            proj_lbl.pack(side="left", padx=4)

            tk.Label(row, text="•", bg='#16213e', fg='#444',
                    font=("Consolas", 10)).pack(side="left", padx=2)

            self.worker_vars[f"{wid}_status"] = tk.StringVar(value="")
            detail_lbl = tk.Label(row, textvariable=self.worker_vars[f"{wid}_status"],
                                 bg='#16213e', fg='#ffd93d',
                                 font=("Consolas", 9), anchor="w", width=40)
            detail_lbl.pack(side="left", padx=4)

            pb_canvas = tk.Canvas(row, bg='#16213e', width=110, height=10,
                                 highlightthickness=0, bd=0)
            pb_canvas.pack(side="left", padx=5)

            pct_lbl = tk.Label(row, text="", width=5, bg='#16213e', fg='#888',
                              font=("Consolas", 8))
            pct_lbl.pack(side="left")

            self.worker_labels[wid] = {
                'name': name_lbl, 'badge': badge,
                'project': proj_lbl, 'detail': detail_lbl, 'pct': pct_lbl
            }
            self.worker_rows[wid] = row
            self.worker_progress[wid] = pb_canvas

        # === PROJECT LIST HEADER ===
        proj_hdr = tk.Frame(self, bg='#0d1b2a', height=28)
        proj_hdr.pack(fill="x", padx=5, pady=(8, 0))
        proj_hdr.pack_propagate(False)

        tk.Label(proj_hdr, text="DANH SACH DU AN", bg='#0d1b2a', fg='#00d9ff',
                 font=("Arial", 10, "bold")).pack(side="left", padx=10, pady=4)

        self.total_projects_var = tk.StringVar(value="")
        tk.Label(proj_hdr, textvariable=self.total_projects_var, bg='#0d1b2a', fg='#888',
                 font=("Consolas", 9)).pack(side="right", padx=10)

        # Column headers
        col_hdr = tk.Frame(self, bg='#0f3460')
        col_hdr.pack(fill="x", padx=5)
        for txt, w in [("Ma", 12), ("Excel", 6), ("Tham chieu", 12), ("Scene anh", 10), ("Con lai", 11), ("", 6)]:
            tk.Label(col_hdr, text=txt, width=w, bg='#0f3460', fg='white',
                     font=("Arial", 9, "bold")).pack(side="left", padx=2, pady=4)

        # Scrollable project list
        proj_list = tk.Frame(self, bg='#16213e')
        proj_list.pack(fill="both", expand=True, padx=5)

        proj_canvas = tk.Canvas(proj_list, bg='#16213e', highlightthickness=0)
        proj_sb = ttk.Scrollbar(proj_list, orient="vertical", command=proj_canvas.yview)
        self.projects_frame = tk.Frame(proj_canvas, bg='#16213e')

        proj_canvas.configure(yscrollcommand=proj_sb.set)
        proj_sb.pack(side="right", fill="y")
        proj_canvas.pack(side="left", fill="both", expand=True)
        proj_canvas.create_window((0, 0), window=self.projects_frame, anchor="nw")
        self.projects_frame.bind("<Configure>",
            lambda e: proj_canvas.configure(scrollregion=proj_canvas.bbox("all")))
        proj_canvas.bind_all("<MouseWheel>",
            lambda e: proj_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units"))

        self.project_rows: Dict[str, dict] = {}

        # Dummy vars for compat (old code may reference these)
        self.current_action_var = tk.StringVar(value="")

    def _select_project(self, code: str):
        """Click project - mo popup Excel detail."""
        self.selected_project = code
        self._show_excel_detail(code)

    def _switch_log_tab(self, worker_id: str):
        """Switch log tab to show different worker."""
        self._update_worker_logs(worker_id)

    def _force_complete_project(self):
        """v1.0.66: Force complete current project and move to next."""
        if not self.manager:
            return

        # Get current project from Chrome workers
        current_project = None
        for wid in ["chrome_1", "chrome_2"]:
            status = self.manager.get_worker_status(wid)
            if status and status.get('current_project'):
                current_project = status.get('current_project')
                break

        if not current_project:
            from tkinter import messagebox
            messagebox.showinfo("Thong bao", "Khong co ma dang chay!")
            return

        # Confirm
        from tkinter import messagebox
        if not messagebox.askyesno(
            "Xac nhan",
            f"Ban muon HOAN THANH ma {current_project}?\n\n"
            f"- Copy ket qua ve may chu\n"
            f"- Chuyen sang ma tiep theo"
        ):
            return

        # Force complete
        self.manager.log("=" * 60, "SYSTEM")
        self.manager.log(f"FORCE COMPLETE: {current_project}", "SYSTEM", "WARN")
        self.manager.log("=" * 60, "SYSTEM")

        try:
            # Copy to master
            self.manager.log(f"Copying {current_project} to master...", "SYSTEM")
            copy_ok = self.manager.copy_project_to_master(current_project)

            if not copy_ok:
                messagebox.showerror("Loi", f"Khong the copy ma {current_project} sang may chu!\nKiem tra AUTO path.")
                return

            # Mark as completed
            if not hasattr(self.manager, '_completed_projects'):
                self.manager._completed_projects = set()
            self.manager._completed_projects.add(current_project)

            # Reset project tracking
            self.manager.project_start_time = None
            self.manager.current_project_code = None

            messagebox.showinfo("Thanh cong", f"Da hoan thanh ma {current_project}!\nChuyen sang ma tiep theo...")

        except Exception as e:
            self.manager.log(f"Force complete failed: {e}", "SYSTEM", "ERROR")
            messagebox.showerror("Loi", f"Khong the hoan thanh ma:\n{e}")

    def _force_complete_by_code(self, project_code: str):
        """v1.0.77: Force complete specific project by code."""
        if not self.manager:
            return

        if not project_code:
            return

        # Confirm
        from tkinter import messagebox
        if not messagebox.askyesno(
            "Xac nhan",
            f"Ban muon HOAN THANH ma {project_code}?\n\n"
            f"- Copy ket qua ve may chu\n"
            f"- Danh dau hoan thanh"
        ):
            return

        # Force complete
        self.manager.log("=" * 60, "SYSTEM")
        self.manager.log(f"FORCE COMPLETE: {project_code}", "SYSTEM", "WARN")
        self.manager.log("=" * 60, "SYSTEM")

        try:
            # Copy to master
            self.manager.log(f"Copying {project_code} to master...", "SYSTEM")
            copy_ok = self.manager.copy_project_to_master(project_code)

            if not copy_ok:
                messagebox.showerror("Loi", f"Khong the copy ma {project_code} sang may chu!\nKiem tra AUTO path.")
                return

            # Mark as completed
            if not hasattr(self.manager, '_completed_projects'):
                self.manager._completed_projects = set()
            self.manager._completed_projects.add(project_code)

            # Reset project timer
            if self.manager.current_project_code == project_code:
                self.manager.project_start_time = None
                self.manager.current_project_code = None

            # Update button to show completed
            if project_code in self.project_rows:
                btn = self.project_rows[project_code]['labels'].get('xong_btn')
                if btn:
                    btn.config(text="DA XONG", bg='#00aa00', state='disabled')

            messagebox.showinfo("Thanh cong", f"Da hoan thanh ma {project_code}!")

        except Exception as e:
            self.manager.log(f"Force complete failed: {e}", "SYSTEM", "ERROR")
            messagebox.showerror("Loi", f"Khong the hoan thanh ma:\n{e}")

    def _update_worker_logs(self, worker_id: str = None):
        """v1.0.47: Show structured worker status instead of raw logs."""
        if worker_id is None:
            worker_id = self.log_tab_var.get()

        if not self.manager:
            self.log_text.delete('1.0', tk.END)
            self.log_text.insert('1.0', f"[{worker_id}] Manager not started\n")
            return

        # Get worker status from agent protocol
        status = self.manager.get_worker_status(worker_id)

        self.log_text.delete('1.0', tk.END)

        if not status:
            self.log_text.insert('1.0', f"[{worker_id}] Waiting for status...\n")
            return

        # Build structured status display
        lines = []
        lines.append(f"{'='*50}")
        lines.append(f"  WORKER: {worker_id.upper()}")
        lines.append(f"{'='*50}")
        lines.append("")

        state = status.get('state', 'idle')
        project = status.get('current_project', '')

        # State indicator
        if state == 'working':
            lines.append(f"  Status:  WORKING")
        elif state == 'error':
            lines.append(f"  Status:  ERROR")
        else:
            lines.append(f"  Status:  IDLE")

        # Project
        if project:
            lines.append(f"  Project: {project}")
        else:
            lines.append(f"  Project: (none)")

        lines.append("")

        # Progress details based on worker type
        if worker_id == "excel":
            step = status.get('current_step', 0)
            step_name = status.get('step_name', '')
            if step > 0:
                lines.append(f"  Step:    {step}/7 - {step_name}")
                progress_bar = self._make_progress_bar(step, 7)
                lines.append(f"  Progress: {progress_bar}")
        else:
            # Chrome worker
            current_scene = status.get('current_scene', 0)
            total_scenes = status.get('total_scenes', 0)
            completed = status.get('completed_count', 0)
            failed = status.get('failed_count', 0)
            step_name = status.get('step_name', '')  # v1.0.65: actual scene ID

            # v1.0.65: Show progress and actual scene ID
            if current_scene > 0 and total_scenes > 0:
                # Show progress: 34/445
                lines.append(f"  Progress: {current_scene}/{total_scenes}")
                progress_bar = self._make_progress_bar(current_scene, total_scenes)
                lines.append(f"           {progress_bar}")
                lines.append("")

            # Show actual scene ID being processed
            if step_name:
                lines.append(f"  Processing: {step_name}")
            elif current_scene > 0:
                lines.append(f"  Processing: scene {current_scene}")
            lines.append("")

            # Lifetime stats
            if completed > 0:
                lines.append(f"  Total Done: {completed}")
            if failed > 0:
                lines.append(f"  Total Failed: {failed}")

        # Last action
        last_action = status.get('last_action', '')
        if last_action:
            lines.append("")
            lines.append(f"  Action:  {last_action[:50]}")

        # Last error
        last_error = status.get('last_error', '')
        if last_error:
            lines.append("")
            lines.append(f"  ERROR:   {last_error[:60]}")

        # v1.0.66: Project Elapsed Time and Remaining Time (6 tiếng max)
        project_elapsed = status.get('project_elapsed_seconds', 0)
        PROJECT_TIMEOUT = 6 * 3600  # 6 hours = max time per project

        if project_elapsed > 0:
            # Elapsed time for current project
            elapsed_hours = int(project_elapsed // 3600)
            elapsed_mins = int((project_elapsed % 3600) // 60)
            lines.append("")
            lines.append(f"  Project Time: {elapsed_hours}h {elapsed_mins}m")

            # Remaining time calculation
            remaining = max(0, PROJECT_TIMEOUT - project_elapsed)
            remaining_hours = int(remaining // 3600)
            remaining_mins = int((remaining % 3600) // 60)

            if remaining > 0:
                lines.append(f"  Remaining: {remaining_hours}h {remaining_mins}m")
            else:
                lines.append(f"  Remaining: TIMEOUT!")

            # Time progress bar
            time_progress = self._make_progress_bar(int(project_elapsed), PROJECT_TIMEOUT)
            lines.append(f"  Timeout:  {time_progress}")

        lines.append("")
        lines.append(f"{'='*50}")

        self.log_text.insert('1.0', '\n'.join(lines))

    def _make_progress_bar(self, current: int, total: int, width: int = 20) -> str:
        """Create a simple ASCII progress bar."""
        if total <= 0:
            return "[" + "-" * width + "] 0%"
        pct = min(100, int(current * 100 / total))
        filled = int(width * current / total)
        bar = "=" * filled + "-" * (width - filled)
        return f"[{bar}] {pct}%"

    def _show_excel_detail(self, code: str):
        """Show popup with detailed Excel status."""
        if not self.manager:
            return

        status = self.manager.quality_checker.get_project_status(code)
        if not status:
            return

        popup = tk.Toplevel(self)
        popup.title(f"Excel Detail - {code}")
        popup.geometry("450x380")
        popup.configure(bg='#1a1a2e')
        popup.transient(self)
        popup.grab_set()

        # Header
        tk.Label(popup, text=f"EXCEL STATUS: {code}", bg='#0f3460', fg='#00ff88',
                 font=("Arial", 12, "bold"), pady=10).pack(fill="x")

        # Content frame
        content = tk.Frame(popup, bg='#1a1a2e', padx=20, pady=10)
        content.pack(fill="both", expand=True)

        # Get values from ProjectStatus
        excel_status = getattr(status, 'excel_status', 'none')
        total_scenes = getattr(status, 'total_scenes', 0)
        img_prompts = getattr(status, 'img_prompts_count', 0)
        video_prompts = getattr(status, 'video_prompts_count', 0)
        fallback_prompts = getattr(status, 'fallback_prompts', 0)
        missing_img = getattr(status, 'missing_img_prompts', [])
        characters_count = getattr(status, 'characters_count', 0)
        characters_with_ref = getattr(status, 'characters_with_ref', 0)

        # Status rows
        rows_data = [
            ("SRT file", "OK" if getattr(status, 'srt_exists', False) else "--"),
            ("Excel file", "OK" if getattr(status, 'excel_exists', False) else "--"),
            ("Total Scenes", str(total_scenes)),
            ("Characters", f"{characters_count}" if characters_count else "--"),
            ("Char with NV", f"{characters_with_ref}/{characters_count}" if characters_count else "--"),
            ("Img Prompts", f"{img_prompts}/{total_scenes}" if total_scenes else "--"),
            ("Video Prompts", f"{video_prompts}/{total_scenes}" if total_scenes else "--"),
            ("Fallback", str(fallback_prompts) if fallback_prompts > 0 else "0"),
        ]

        for name, value in rows_data:
            row = tk.Frame(content, bg='#1a1a2e')
            row.pack(fill="x", pady=3)

            # Determine color based on value
            if value == "OK" or (value.endswith(f"/{total_scenes}") and value.startswith(str(total_scenes))):
                color = '#00ff88'
            elif value == "--" or value == "0":
                color = '#666'
            elif "/" in value:  # partial like "5/10"
                color = '#00d9ff'
            else:
                color = 'white'

            tk.Label(row, text=name, width=15, bg='#1a1a2e', fg='white',
                     font=("Consolas", 10), anchor="w").pack(side="left")
            tk.Label(row, text=value, width=12, bg='#1a1a2e', fg=color,
                     font=("Consolas", 10, "bold")).pack(side="left")

        # Summary
        if excel_status == "complete":
            summary = "HOAN THANH" + (" (co Fallback)" if fallback_prompts > 0 else "")
            summary_color = '#00ff88'
        elif excel_status == "partial":
            pct = int(img_prompts * 100 / total_scenes) if total_scenes > 0 else 0
            summary = f"Dang tao: {pct}%"
            summary_color = '#00d9ff'
        elif excel_status == "fallback":
            summary = "Co Fallback - can API"
            summary_color = '#ffd93d'
        else:
            summary = "Chua co Excel"
            summary_color = '#ff6b6b'

        tk.Label(content, text=summary, bg='#1a1a2e', fg=summary_color,
                 font=("Arial", 11, "bold"), pady=10).pack()

        # Close button
        tk.Button(popup, text="Dong", command=popup.destroy, bg='#0f3460', fg='white',
                  font=("Arial", 10), padx=20).pack(pady=10)

    def _show_nv_detail(self, code: str):
        """Show popup with detailed NV (reference images) status."""
        if not self.manager:
            return

        status = self.manager.quality_checker.get_project_status(code)
        project_dir = TOOL_DIR / "PROJECTS" / code
        nv_dir = project_dir / "nv"

        popup = tk.Toplevel(self)
        popup.title(f"NV Detail - {code}")
        popup.geometry("500x400")
        popup.configure(bg='#1a1a2e')
        popup.transient(self)
        popup.grab_set()

        # Header
        tk.Label(popup, text=f"ANH THAM CHIEU (NV): {code}", bg='#0f3460', fg='#ff6b6b',
                 font=("Arial", 12, "bold"), pady=10).pack(fill="x")

        # Content frame with scrollbar
        canvas = tk.Canvas(popup, bg='#1a1a2e', highlightthickness=0)
        scrollbar = ttk.Scrollbar(popup, orient="vertical", command=canvas.yview)
        content = tk.Frame(canvas, bg='#1a1a2e')

        canvas.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True, padx=10, pady=5)
        canvas.create_window((0, 0), window=content, anchor="nw")
        content.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))

        # Get characters from Excel
        excel_path = project_dir / f"{code}_prompts.xlsx"
        characters = []
        if excel_path.exists():
            try:
                from openpyxl import load_workbook
                wb = load_workbook(excel_path, read_only=True)
                if "Characters" in wb.sheetnames:
                    ws = wb["Characters"]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and row[0]:
                            char_id = str(row[0]).strip()
                            if char_id:
                                characters.append(char_id)
                wb.close()
            except:
                pass

        if not characters:
            tk.Label(content, text="Khong co Characters trong Excel", bg='#1a1a2e', fg='#666',
                     font=("Consolas", 10), pady=20).pack()
        else:
            # Header row
            header = tk.Frame(content, bg='#0f3460')
            header.pack(fill="x", pady=(0, 5))
            tk.Label(header, text="ID", width=15, bg='#0f3460', fg='white',
                     font=("Consolas", 9, "bold")).pack(side="left", padx=5)
            tk.Label(header, text="Status", width=10, bg='#0f3460', fg='white',
                     font=("Consolas", 9, "bold")).pack(side="left", padx=5)

            # List each character
            done_count = 0
            for i, char_id in enumerate(characters):
                bg = '#1a1a2e' if i % 2 == 0 else '#16213e'
                row = tk.Frame(content, bg=bg)
                row.pack(fill="x", pady=1)

                # Check if image exists
                has_image = False
                for ext in ['.png', '.jpg', '.jpeg', '.webp']:
                    if (nv_dir / f"{char_id}{ext}").exists():
                        has_image = True
                        done_count += 1
                        break

                tk.Label(row, text=char_id, width=15, bg=bg, fg='white',
                         font=("Consolas", 10), anchor="w").pack(side="left", padx=5)

                status_text = "OK" if has_image else "Thieu"
                status_color = '#00ff88' if has_image else '#ff6b6b'
                tk.Label(row, text=status_text, width=10, bg=bg, fg=status_color,
                         font=("Consolas", 10, "bold")).pack(side="left", padx=5)

            # Summary
            total = len(characters)
            summary_frame = tk.Frame(content, bg='#1a1a2e')
            summary_frame.pack(fill="x", pady=10)

            if done_count >= total:
                summary = f"HOAN THANH: {done_count}/{total}"
                summary_color = '#00ff88'
            else:
                summary = f"CON THIEU: {total - done_count}/{total}"
                summary_color = '#ff6b6b'

            tk.Label(summary_frame, text=summary, bg='#1a1a2e', fg=summary_color,
                     font=("Arial", 11, "bold")).pack()

        # Close button
        tk.Button(popup, text="Dong", command=popup.destroy, bg='#0f3460', fg='white',
                  font=("Arial", 10), padx=20).pack(pady=10)

    def _show_prompt_popup(self, scene_id, prompt_text: str):
        """Show popup with full prompt text and copy button."""
        popup = tk.Toplevel(self)
        popup.title(f"Prompt - Scene {scene_id}")
        popup.geometry("600x400")
        popup.configure(bg='#1a1a2e')
        popup.transient(self)
        popup.grab_set()

        # Header
        tk.Label(popup, text=f"PROMPT - Scene {scene_id}", bg='#0f3460', fg='#ffd93d',
                 font=("Arial", 12, "bold"), pady=10).pack(fill="x")

        # Text widget with scrollbar
        text_frame = tk.Frame(popup, bg='#1a1a2e')
        text_frame.pack(fill="both", expand=True, padx=10, pady=10)

        scrollbar = ttk.Scrollbar(text_frame)
        scrollbar.pack(side="right", fill="y")

        text_widget = tk.Text(text_frame, wrap="word", bg='#16213e', fg='white',
                              font=("Consolas", 10), yscrollcommand=scrollbar.set,
                              padx=10, pady=10)
        text_widget.pack(fill="both", expand=True)
        text_widget.insert("1.0", prompt_text)
        text_widget.config(state="disabled")

        scrollbar.config(command=text_widget.yview)

        # Button frame
        btn_frame = tk.Frame(popup, bg='#1a1a2e')
        btn_frame.pack(pady=10)

        def copy_to_clipboard():
            self.clipboard_clear()
            self.clipboard_append(prompt_text)
            copy_btn.config(text="Da copy!")
            popup.after(1500, lambda: copy_btn.config(text="Copy"))

        copy_btn = tk.Button(btn_frame, text="Copy", command=copy_to_clipboard,
                             bg='#00d9ff', fg='black', font=("Arial", 10), padx=20)
        copy_btn.pack(side="left", padx=5)

        tk.Button(btn_frame, text="Dong", command=popup.destroy,
                  bg='#0f3460', fg='white', font=("Arial", 10), padx=20).pack(side="left", padx=5)

    def _load_project_detail(self, code: str):
        """Load chi tiet project vao panel phai."""
        project_dir = TOOL_DIR / "PROJECTS" / code
        if not project_dir.exists():
            return

        # Load reference images
        self._load_reference_images(code)

        # Load scenes from Excel
        self._load_scenes_list(code)

    def _load_reference_images(self, project_code: str):
        """Load characters tu Excel - hien thi ID, ten file va character lock."""
        for widget in self.ref_images_frame.winfo_children():
            widget.destroy()
        self.ref_photo_refs = []

        project_dir = TOOL_DIR / "PROJECTS" / project_code
        excel_path = project_dir / f"{project_code}_prompts.xlsx"
        nv_dir = project_dir / "nv"

        if not excel_path.exists():
            tk.Label(self.ref_images_frame, text="Chua co Excel", bg='#1a1a2e', fg='#666',
                     font=("Consolas", 10)).pack(pady=5)
            return

        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(excel_path))

            # Tim sheet characters
            char_sheet = None
            for name in wb.sheetnames:
                if name.lower() == 'characters':
                    char_sheet = name
                    break

            if not char_sheet:
                tk.Label(self.ref_images_frame, text="Khong co sheet characters", bg='#1a1a2e', fg='#666',
                         font=("Consolas", 10)).pack(pady=5)
                wb.close()
                return

            ws = wb[char_sheet]
            headers = [cell.value for cell in ws[1]]

            # Tim index cua cac cot quan trong
            id_idx = headers.index('id') if 'id' in headers else 0
            name_idx = headers.index('name') if 'name' in headers else 1
            lock_idx = headers.index('character_lock') if 'character_lock' in headers else 2
            file_idx = headers.index('image_file') if 'image_file' in headers else -1

            characters = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[id_idx] is None:
                    continue
                characters.append({
                    'id': row[id_idx],
                    'name': row[name_idx] if name_idx < len(row) else '',
                    'lock': row[lock_idx] if lock_idx < len(row) else '',
                    'file': row[file_idx] if file_idx >= 0 and file_idx < len(row) else ''
                })

            wb.close()

            if not characters:
                tk.Label(self.ref_images_frame, text="Chua co characters", bg='#1a1a2e', fg='#666',
                         font=("Consolas", 10)).pack(pady=5)
                return

            # Hien thi header info
            tk.Label(self.ref_images_frame, text=f"Tim thay {len(characters)} characters",
                     bg='#1a1a2e', fg='#00ff88', font=("Consolas", 9, "bold")).pack(pady=2)

            ROW_WIDTH = 700

            for i, char in enumerate(characters[:20]):  # Max 20
                bg_color = '#1a1a2e' if i % 2 == 0 else '#16213e'
                row = tk.Frame(self.ref_images_frame, bg=bg_color, height=40, width=ROW_WIDTH)
                row.pack(fill="x", pady=1)
                row.pack_propagate(False)

                # ID
                tk.Label(row, text=str(char['id']), width=6, bg=bg_color, fg='#00d9ff',
                         font=("Consolas", 9, "bold")).pack(side="left", padx=3)

                # Thumbnail neu co file
                thumb_frame = tk.Frame(row, bg=bg_color, width=40, height=35)
                thumb_frame.pack(side="left", padx=2)
                thumb_frame.pack_propagate(False)

                img_file = char.get('file', '')
                if img_file and nv_dir.exists() and PIL_AVAILABLE:
                    img_path = nv_dir / img_file
                    if img_path.exists():
                        try:
                            img = Image.open(img_path)
                            img.thumbnail((35, 35))
                            photo = ImageTk.PhotoImage(img)
                            self.ref_photo_refs.append(photo)
                            thumb_lbl = tk.Label(thumb_frame, image=photo, bg=bg_color, cursor="hand2")
                            thumb_lbl.pack(expand=True)
                            thumb_lbl.bind("<Button-1>", lambda e, p=img_path: os.startfile(str(p)))
                        except:
                            tk.Label(thumb_frame, text="--", bg=bg_color, fg='#444',
                                     font=("Consolas", 8)).pack(expand=True)
                    else:
                        tk.Label(thumb_frame, text="--", bg=bg_color, fg='#444',
                                 font=("Consolas", 8)).pack(expand=True)
                else:
                    tk.Label(thumb_frame, text="--", bg=bg_color, fg='#444',
                             font=("Consolas", 8)).pack(expand=True)

                # Character lock (truncated) - noi dung chinh
                lock_text = char.get('lock', '') or ''
                lock_display = lock_text[:50] + "..." if len(lock_text) > 50 else lock_text or "--"
                tk.Label(row, text=lock_display, width=50, bg=bg_color, fg='#c8d6e5',
                         font=("Consolas", 8), anchor="w").pack(side="left", padx=3)

        except Exception as e:
            tk.Label(self.ref_images_frame, text=f"Loi: {str(e)[:30]}", bg='#1a1a2e', fg='#ff6b6b',
                     font=("Consolas", 9)).pack(pady=5)
            print(f"[ERROR] Load characters: {e}")

    def _load_scenes_list(self, project_code: str):
        """Load danh sach scenes tu Excel - voi thumbnail va SRT time."""
        for widget in self.scenes_list_frame.winfo_children():
            widget.destroy()
        self.scene_photo_refs = []

        project_dir = TOOL_DIR / "PROJECTS" / project_code
        excel_path = project_dir / f"{project_code}_prompts.xlsx"

        if not excel_path.exists():
            tk.Label(self.scenes_list_frame, text="Chua co Excel", bg='#1a1a2e', fg='#666',
                     font=("Consolas", 12)).pack(pady=20)
            return

        try:
            from modules.excel_manager import PromptWorkbook
            wb = PromptWorkbook(str(excel_path))
            wb.load_or_create()
            scenes = wb.get_scenes()

            # Hien thi so luong scenes
            if not scenes:
                tk.Label(self.scenes_list_frame, text="Khong tim thay scenes trong Excel",
                         bg='#1a1a2e', fg='#ffd93d', font=("Consolas", 10)).pack(pady=10)
                return

            info_label = tk.Label(self.scenes_list_frame, text=f"Tim thay {len(scenes)} scenes",
                                  bg='#1a1a2e', fg='#00ff88', font=("Consolas", 10, "bold"))
            info_label.pack(pady=3)

            img_dir = project_dir / "img"
            vid_dir = project_dir / "vid"

            # Tinh toan width cho row: Status(6) + ID(4) + Thumb(6) + SRT(18) + Prompt(45) + Img(5) + Vid(5)
            ROW_WIDTH = 800

            # v1.0.47: Get current running scene from worker status
            current_running_scene = 0
            if self.manager:
                for wid in ["chrome_1", "chrome_2"]:
                    try:
                        status = self.manager.get_worker_status(wid)
                        if status and status.get('current_project') == project_code:
                            current_running_scene = status.get('current_scene', 0)
                            if current_running_scene > 0:
                                break
                    except:
                        pass

            # Store row references for later updates
            self.scene_rows = {}

            for i, scene in enumerate(scenes[:150]):  # Max 150 scenes
                scene_id = scene.scene_id

                # v1.0.47: Determine scene status and colors
                has_image = find_scene_image(img_dir, scene_id) is not None
                is_running = (scene_id == current_running_scene)

                if has_image:
                    status_text = "DONE"
                    status_fg = '#00ff88'  # Green
                    bg = '#0a2e0a' if i % 2 == 0 else '#0d380d'  # Dark green bg
                elif is_running:
                    status_text = "RUN"
                    status_fg = '#ffd93d'  # Yellow
                    bg = '#2e2e0a' if i % 2 == 0 else '#38380d'  # Dark yellow bg
                else:
                    status_text = "--"
                    status_fg = '#666'
                    bg = '#1a1a2e' if i % 2 == 0 else '#16213e'

                row = tk.Frame(self.scenes_list_frame, bg=bg, height=50, width=ROW_WIDTH)
                row.pack(fill="x", pady=1)
                row.pack_propagate(False)

                # Store reference for updates
                self.scene_rows[scene_id] = {'frame': row, 'bg_even': i % 2 == 0}

                # Status column (NEW)
                status_label = tk.Label(row, text=status_text, width=5, bg=bg, fg=status_fg,
                         font=("Consolas", 9, "bold"))
                status_label.pack(side="left", padx=2)
                self.scene_rows[scene_id]['status_label'] = status_label

                # Scene ID
                tk.Label(row, text=str(scene_id), width=4, bg=bg, fg='#00d9ff',
                         font=("Consolas", 11, "bold")).pack(side="left", padx=3)

                # Thumbnail (45x45)
                img_path = find_scene_image(img_dir, scene_id)
                thumb_frame = tk.Frame(row, bg=bg, width=50, height=45)
                thumb_frame.pack(side="left", padx=3)
                thumb_frame.pack_propagate(False)

                if img_path and PIL_AVAILABLE:
                    try:
                        img = Image.open(str(img_path))
                        img.thumbnail((45, 45))
                        photo = ImageTk.PhotoImage(img)
                        self.scene_photo_refs.append(photo)

                        thumb_lbl = tk.Label(thumb_frame, image=photo, bg=bg, cursor="hand2")
                        thumb_lbl.pack(expand=True)
                        thumb_lbl.bind("<Button-1>", lambda e, p=img_path: os.startfile(str(p)))
                    except:
                        tk.Label(thumb_frame, text="IMG", bg=bg, fg='#1dd1a1',
                                 font=("Consolas", 9)).pack(expand=True)
                else:
                    tk.Label(thumb_frame, text="--", bg=bg, fg='#444',
                             font=("Consolas", 9)).pack(expand=True)

                # SRT time - lay tu srt_start va srt_end
                srt_start = getattr(scene, 'srt_start', '') or ''
                srt_end = getattr(scene, 'srt_end', '') or ''

                # Debug 3 scenes dau
                if i < 3:
                    print(f"[DEBUG SRT] Scene {scene.scene_id}: start={repr(srt_start)}, end={repr(srt_end)}")

                # Convert to string neu la datetime.time hoac timedelta
                if hasattr(srt_start, 'strftime'):
                    srt_start = srt_start.strftime('%H:%M:%S')
                elif hasattr(srt_start, 'total_seconds'):  # timedelta
                    total = int(srt_start.total_seconds())
                    h, m, s = total // 3600, (total % 3600) // 60, total % 60
                    srt_start = f"{h:02d}:{m:02d}:{s:02d}"

                if hasattr(srt_end, 'strftime'):
                    srt_end = srt_end.strftime('%H:%M:%S')
                elif hasattr(srt_end, 'total_seconds'):  # timedelta
                    total = int(srt_end.total_seconds())
                    h, m, s = total // 3600, (total % 3600) // 60, total % 60
                    srt_end = f"{h:02d}:{m:02d}:{s:02d}"

                srt_start = str(srt_start) if srt_start else ''
                srt_end = str(srt_end) if srt_end else ''

                if srt_start and srt_end:
                    # Format: "00:01:23" thay vi "00:01:23,456"
                    start_short = srt_start.split(',')[0] if ',' in str(srt_start) else str(srt_start)
                    end_short = srt_end.split(',')[0] if ',' in str(srt_end) else str(srt_end)
                    srt_text = f"{start_short} - {end_short}"
                else:
                    srt_text = "--"

                tk.Label(row, text=srt_text, width=18, bg=bg, fg='#ffd93d',
                         font=("Consolas", 9)).pack(side="left", padx=3)

                # Prompt (truncated) - click de xem day du
                prompt_text = scene.img_prompt or ""
                prompt_display = prompt_text[:50] + "..." if len(prompt_text) > 50 else prompt_text or "--"
                prompt_label = tk.Label(row, text=prompt_display, width=45, bg=bg, fg='#c8d6e5',
                         font=("Consolas", 9), anchor="w", cursor="hand2" if prompt_text else "")
                prompt_label.pack(side="left", padx=3)
                if prompt_text:
                    prompt_label.bind("<Button-1>", lambda e, p=prompt_text, sid=scene_id: self._show_prompt_popup(sid, p))

                # Image status (simplified - status column already shows this)
                if img_path:
                    img_status = tk.Label(row, text="IMG", width=4, bg=bg, fg='#1dd1a1',
                             font=("Consolas", 9), cursor="hand2")
                    img_status.pack(side="left", padx=2)
                    img_status.bind("<Button-1>", lambda e, p=img_path: os.startfile(str(p)))
                else:
                    tk.Label(row, text="--", width=4, bg=bg, fg='#666',
                             font=("Consolas", 9)).pack(side="left", padx=2)

                # Video status
                vid_path = find_scene_video(vid_dir, scene_id)
                if vid_path:
                    vid_status = tk.Label(row, text="OK", width=5, bg=bg, fg='#1dd1a1',
                             font=("Consolas", 10, "bold"), cursor="hand2")
                    vid_status.pack(side="left", padx=2)
                    # Copy vid_path to avoid closure issue
                    vid_status.bind("<Button-1>", lambda e, p=str(vid_path): os.startfile(p))
                else:
                    tk.Label(row, text="--", width=5, bg=bg, fg='#666',
                             font=("Consolas", 10)).pack(side="left", padx=2)

        except Exception as e:
            tk.Label(self.scenes_list_frame, text=f"Loi: {e}", bg='#1a1a2e', fg='#ff6b6b',
                     font=("Consolas", 10)).pack(pady=20)

    def _update_scene_status(self):
        """v1.0.47: Update scene status colors without reloading entire list."""
        if not self.selected_project or not hasattr(self, 'scene_rows') or not self.scene_rows:
            return

        project_code = self.selected_project
        project_dir = TOOL_DIR / "PROJECTS" / project_code
        img_dir = project_dir / "img"

        # Get current running scene from workers
        current_running_scene = 0
        if self.manager:
            for wid in ["chrome_1", "chrome_2"]:
                try:
                    status = self.manager.get_worker_status(wid)
                    if status and status.get('current_project') == project_code:
                        current_running_scene = status.get('current_scene', 0)
                        if current_running_scene > 0:
                            break
                except:
                    pass

        # Update each row's status
        for scene_id, row_data in self.scene_rows.items():
            try:
                has_image = find_scene_image(img_dir, scene_id) is not None
                is_running = (scene_id == current_running_scene)
                is_even = row_data.get('bg_even', True)

                if has_image:
                    status_text = "DONE"
                    status_fg = '#00ff88'
                    bg = '#0a2e0a' if is_even else '#0d380d'
                elif is_running:
                    status_text = "RUN"
                    status_fg = '#ffd93d'
                    bg = '#2e2e0a' if is_even else '#38380d'
                else:
                    status_text = "--"
                    status_fg = '#666'
                    bg = '#1a1a2e' if is_even else '#16213e'

                # Update status label
                status_label = row_data.get('status_label')
                if status_label:
                    status_label.config(text=status_text, fg=status_fg, bg=bg)

                # Update row background
                frame = row_data.get('frame')
                if frame:
                    frame.config(bg=bg)
                    for child in frame.winfo_children():
                        try:
                            child.config(bg=bg)
                        except:
                            pass
            except:
                pass

    def _update_detail_panel(self):
        """Update detail panel neu co project duoc chon."""
        if self.selected_project:
            # Update Excel steps status
            self._update_excel_steps(self.selected_project)
            # v1.0.47: Update scene status colors
            self._update_scene_status()

    def _update_excel_steps(self, project_code: str):
        """Update mau cua Excel step labels dua tren trang thai."""
        project_dir = TOOL_DIR / "PROJECTS" / project_code
        excel_path = project_dir / f"{project_code}_prompts.xlsx"

        if not excel_path.exists():
            # Reset all to gray
            for lbl in self.excel_step_labels:
                lbl.config(fg='#666')
            return

        try:
            from modules.excel_manager import PromptWorkbook
            wb = PromptWorkbook(str(excel_path))
            wb.load_or_create()

            # Get step status
            step_ids = ["step_1", "step_2", "step_3", "step_4", "step_5", "step_6", "step_7"]

            if hasattr(wb, 'get_all_step_status'):
                all_status = wb.get_all_step_status()
                status_dict = {}
                for s in all_status:
                    if isinstance(s, dict):
                        status_dict[s.get('step_id', '')] = s.get('status', 'PENDING')

                for i, step_id in enumerate(step_ids):
                    status = status_dict.get(step_id, 'PENDING')
                    if status == 'COMPLETED':
                        self.excel_step_labels[i].config(fg='#00ff88')  # Green
                    elif status == 'IN_PROGRESS':
                        self.excel_step_labels[i].config(fg='#ffd93d')  # Yellow
                    elif status == 'ERROR':
                        self.excel_step_labels[i].config(fg='#ff6b6b')  # Red
                    else:
                        self.excel_step_labels[i].config(fg='#666')  # Gray
            else:
                # Fallback - check if sheets exist
                for i, lbl in enumerate(self.excel_step_labels):
                    lbl.config(fg='#666')

        except Exception:
            for lbl in self.excel_step_labels:
                lbl.config(fg='#666')

    def _load_projects_on_startup(self):
        """Load projects ngay khi mo GUI (truoc khi bat dau chay)."""
        if not self.manager:
            return

        try:
            projects = self.manager.scan_projects()[:12]

            for i, code in enumerate(projects):
                if code not in self.project_rows:
                    self._create_row(code, i)

                # Load basic info
                status = self.manager.quality_checker.get_project_status(code)
                if status:
                    labels = self.project_rows[code]['labels']

                    # Excel status - show OK or %
                    excel_status = getattr(status, 'excel_status', '')
                    fallback_prompts = getattr(status, 'fallback_prompts', 0)

                    if excel_status == "complete":
                        text = "OK*" if fallback_prompts > 0 else "OK"
                        labels['excel'].config(text=text, fg='#00ff88')
                    elif excel_status == "partial":
                        img_prompts = getattr(status, 'img_prompts_count', 0)
                        total = getattr(status, 'total_scenes', 0)
                        if total > 0:
                            pct = int(img_prompts * 100 / total)
                            labels['excel'].config(text=f"{pct}%", fg='#00d9ff')
                        else:
                            labels['excel'].config(text="--", fg='#666')
                    elif excel_status == "fallback":
                        labels['excel'].config(text="FB", fg='#ffd93d')
                    else:
                        labels['excel'].config(text="--", fg='#666')

                    # Tham chieu (Reference images)
                    nv_done = getattr(status, 'characters_with_ref', 0)
                    nv_total = getattr(status, 'characters_count', 0)
                    if nv_total > 0:
                        color = '#00ff88' if nv_done >= nv_total else '#ff6b6b'
                        labels['thamchieu'].config(text=f"{nv_done}/{nv_total}", fg=color)
                    else:
                        labels['thamchieu'].config(text="--", fg='#666')

                    # Scene anh
                    img_done = getattr(status, 'images_done', 0)
                    img_total = getattr(status, 'total_scenes', 0)
                    if img_total > 0:
                        color = '#00ff88' if img_done >= img_total else '#00d9ff'
                        labels['images'].config(text=f"{img_done}/{img_total}", fg=color)
                    else:
                        labels['images'].config(text="--", fg='#666')

                    labels['conlai'].config(text="--", fg='#888')

        except Exception as e:
            print(f"Error loading projects: {e}")

    def _start(self):
        # v1.0.346: Không start nếu đang chạy (tránh duplicate từ master RUN)
        if self.running:
            return

        if not self.manager:
            self.manager = VMManager(num_chrome_workers=2)

        self.manager.settings.excel_mode = self.mode_var.get()
        self.manager.settings.video_mode = self.mode_var.get()

        # Auto-detect IPv6 before starting
        # Check if IPv6 is available and working
        ipv6_ok = self._auto_detect_ipv6()

        self.running = True
        ipv6_status = "IPv6" if ipv6_ok else "Direct"
        self.status_var.set(f"Dang chay ({self.mode_var.get().upper()}) - {ipv6_status}")
        self.start_btn.config(bg='#666', state="disabled")

        # Log start
        if LOGGER_AVAILABLE:
            from modules.central_logger import log
            log("main", f"=== STARTED === Mode: {self.mode_var.get()}, IPv6: {ipv6_status}", "INFO")

        def run():
            # v1.0.346: Kill Chrome cũ trước khi start (giống start_all)
            self.manager.kill_all_chrome()

            # 1. Start Excel worker first
            if self.manager.enable_excel:
                self.manager.start_worker("excel", gui_mode=True)
                time.sleep(2)

            # 2. Start Chrome workers (login sẽ xử lý trong mỗi worker)
            for i in range(1, self.manager.num_chrome_workers + 1):
                self.manager.start_worker(f"chrome_{i}", gui_mode=True)
                time.sleep(2)

            # v1.0.335: Reset timers tránh auto-restart/recovery chạy ngay sau start
            self.manager.chrome_last_restart = time.time()
            self.manager._start_time = time.time()

            # 4. Start orchestration (chỉ tạo thread mới nếu thread cũ đã chết)
            self.manager._stop_flag = False
            if self.manager._orch_thread is None or not self.manager._orch_thread.is_alive():
                self.manager._orch_thread = threading.Thread(target=self.manager.orchestrate, daemon=True)
                self.manager._orch_thread.start()

            # v1.0.335: Start watchdog để báo status cho master
            self.manager.start_watchdog()

            # 5. Auto-arrange tat ca cua so sau khi workers da mo
            # v1.0.366: Doi 15s (Chrome can thoi gian mo) + arrange 2 lan
            time.sleep(15)
            self.after(0, self._arrange_windows)
            time.sleep(10)
            self.after(0, self._arrange_windows)

        threading.Thread(target=run, daemon=True).start()

    def _pre_login_chrome(self):
        """
        v1.0.121: Không dùng nữa - login xử lý trong Chrome workers.
        """
        pass

    def _stop(self):
        if self.manager and self.running:
            self.running = False
            self.status_var.set("Dang dung...")

            # Log stop
            if LOGGER_AVAILABLE:
                from modules.central_logger import log
                log("main", "=== STOPPING ===", "INFO")

            # Kill all CMD and Chrome processes
            def stop_and_kill():
                self.manager.stop_all()
                self.manager.kill_all_chrome()
            threading.Thread(target=stop_and_kill, daemon=True).start()
            self.start_btn.config(bg='#00ff88', state="normal")

    def _reset_workers(self):
        """Reset workers: Kill all Chrome + CMD, then restart workers."""
        if not self.manager:
            return

        from tkinter import messagebox
        if not messagebox.askyesno("Reset Workers",
                                   "Reset tat ca workers?\n\n"
                                   "- Tat tat ca CMD (Excel + Chrome)\n"
                                   "- Kill Chrome processes\n"
                                   "- Khoi dong lai tat ca workers\n\n"
                                   "Tiep tuc?"):
            return

        self.status_var.set("Dang reset workers...")
        self.reset_btn.config(bg='#666', state="disabled")

        def do_reset():
            try:
                # Log
                if LOGGER_AVAILABLE:
                    from modules.central_logger import log
                    log("main", "=== RESET WORKERS ===", "INFO")

                # 1. Stop all workers
                self.manager.stop_all()
                time.sleep(2)

                # 2. Kill all Chrome + CMD
                self.manager.kill_all_chrome()
                time.sleep(2)

                # 3. Restart ALL workers (Excel + Chrome)
                self.manager._stop_flag = False  # Cho phép orchestrate tiếp tục
                for wid in self.manager.workers:
                    self.manager.start_worker(wid)
                    time.sleep(2)

                # Update status
                self.status_var.set("Reset xong!")
                messagebox.showinfo("Reset Complete", "Tat ca workers da duoc reset thanh cong!")

            except Exception as e:
                self.status_var.set(f"Loi reset: {str(e)[:40]}")
                messagebox.showerror("Reset Error", f"Loi: {e}")

            finally:
                self.reset_btn.config(bg='#ff6348', state="normal")

        threading.Thread(target=do_reset, daemon=True).start()

    def _toggle_windows(self):
        """Toggle Chrome and CMD windows visibility."""
        if not self.manager:
            return

        if self.windows_visible:
            # Hide both Chrome and CMD
            self.manager.hide_chrome_windows()
            self.manager.hide_cmd_windows()
            self.toggle_btn.config(text="HIEN CHROME", bg='#6c5ce7')
            self.windows_visible = False
        else:
            # Show both Chrome and CMD
            self.manager.show_chrome_windows()
            self.manager.show_cmd_windows()
            self.toggle_btn.config(text="AN CHROME", bg='#00b894')
            self.windows_visible = True


    def _auto_hide_windows(self):
        """Auto-hide Chrome and CMD windows when GUI starts."""
        if self.manager:
            try:
                print("[GUI] Auto-hiding CMD and Chrome windows...")
                self.manager.hide_cmd_windows()
                self.manager.hide_chrome_windows()
                self.toggle_btn.config(text="HIEN CHROME", bg='#6c5ce7')
                self.windows_visible = False
                print("[GUI] Windows hidden successfully")
            except Exception as e:
                print(f"[GUI] Error hiding windows: {e}")

    def _get_git_version(self) -> str:
        """Lay thong tin version tu VERSION.txt (uu tien) hoac git."""
        # Uu tien VERSION.txt (chua thoi gian chinh xac)
        try:
            version_file = TOOL_DIR / "VERSION.txt"
            if version_file.exists():
                with open(version_file, 'r', encoding='utf-8') as f:
                    lines = f.read().strip().split('\n')
                    if len(lines) >= 2:
                        version = lines[0].strip()  # VD: 1.0.37
                        date_time = lines[1].strip()  # VD: 2026-01-27 10:05
                        return f"v{version} | {date_time}"
                    elif len(lines) >= 1:
                        version = lines[0].strip()
                        return f"v{version}"
        except Exception as e:
            print(f"[GUI] Version file error: {e}")

        # Fallback: Thu git neu VERSION.txt khong co
        try:
            import subprocess
            result = subprocess.run(['git', 'rev-parse', '--short', 'HEAD'],
                                  capture_output=True, text=True, cwd=str(TOOL_DIR), timeout=2)
            if result.returncode == 0:
                commit_hash = result.stdout.strip()
                return f"v{commit_hash}"
        except:
            pass

        return "unknown"

    def _get_ipv6_setting(self) -> bool:
        """Doc IPv6 enabled tu settings.yaml. Mac dinh la True."""
        try:
            import yaml
            config_path = TOOL_DIR / "config" / "settings.yaml"
            if config_path.exists():
                with open(config_path, "r", encoding="utf-8") as f:
                    config = yaml.safe_load(f) or {}
                ipv6_cfg = config.get('ipv6_rotation', {})
                # Mac dinh la True neu khong co setting
                return ipv6_cfg.get('enabled', True)
        except:
            pass
        return True  # Mac dinh enabled

    def _auto_detect_ipv6(self) -> bool:
        """
        Auto-detect IPv6 connectivity.
        v1.0.562: Pool mode → test Pool API connection first.
        File mode → Test ping to Google DNS IPv6 for each IP in config/ipv6.txt.
        Returns True if at least one IPv6 works, False if none work.
        Also updates settings.yaml automatically.
        """
        import subprocess
        import yaml

        config_path = TOOL_DIR / "config" / "settings.yaml"

        # v1.0.562: Check Pool mode first
        try:
            if config_path.exists():
                with open(config_path, "r", encoding="utf-8") as f:
                    _cfg = yaml.safe_load(f) or {}
                pool_api_url = _cfg.get('mikrotik', {}).get('pool_api_url', '')
                if pool_api_url:
                    print(f"[IPv6] Pool mode: testing {pool_api_url}...")
                    self.status_var.set("Dang kiem tra IPv6 Pool...")
                    self.update()
                    try:
                        from modules.ipv6_pool_client import IPv6PoolClient
                        client = IPv6PoolClient(api_url=pool_api_url, timeout=5)
                        if client.ping():
                            status = client.get_status()
                            avail = status.get('available', 0) if status else 0
                            print(f"[IPv6] Pool API OK! {avail} IPs available")
                            self._set_ipv6_enabled(True)
                            return True
                        else:
                            print(f"[IPv6] Pool API not available, fallback to file")
                    except Exception as e:
                        print(f"[IPv6] Pool API error: {e}, fallback to file")
        except:
            pass

        # Find IPv6 file (ipv6.txt or ipv6_list.txt)
        ipv6_file = TOOL_DIR / "config" / "ipv6.txt"
        if not ipv6_file.exists():
            ipv6_file = TOOL_DIR / "config" / "ipv6_list.txt"

        if not ipv6_file.exists():
            print("[IPv6] No IPv6 file found, using direct connection")
            self._set_ipv6_enabled(False)
            return False

        # Read IPv6 list
        try:
            with open(ipv6_file, "r", encoding="utf-8") as f:
                ipv6_list = [line.strip() for line in f if line.strip() and not line.startswith('#')]
        except Exception as e:
            print(f"[IPv6] Error reading IPv6 file: {e}")
            self._set_ipv6_enabled(False)
            return False

        if not ipv6_list:
            print("[IPv6] IPv6 file is empty, using direct connection")
            self._set_ipv6_enabled(False)
            return False

        # Test first 3 IPs (enough to detect if IPv6 works)
        test_count = min(3, len(ipv6_list))
        working_count = 0

        print(f"[IPv6] Testing {test_count} IPv6 addresses...")
        self.status_var.set("Dang kiem tra IPv6...")
        self.update()

        for i, ip in enumerate(ipv6_list[:test_count]):
            try:
                # v1.0.375: Dùng curl -6 thay vì ping (nhiều mạng chặn ICMP nhưng HTTP vẫn OK)
                ok = False
                try:
                    result = subprocess.run(
                        'curl -6 --connect-timeout 5 -s -o nul -w "%{http_code}" https://www.google.com',
                        shell=True, capture_output=True, text=True, timeout=8,
                        creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0
                    )
                    ok = result.returncode == 0 and result.stdout.strip().startswith(('2', '3'))
                except Exception:
                    pass

                if not ok:
                    # Fallback: ping
                    result = subprocess.run(
                        ['ping', '-n', '1', '-w', '3000', '2001:4860:4860::8888'],
                        capture_output=True, text=True, timeout=5,
                        creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0
                    )
                    ok = result.returncode == 0

                if ok:
                    working_count += 1
                    print(f"[IPv6] Test {i+1}/{test_count}: OK")
                else:
                    print(f"[IPv6] Test {i+1}/{test_count}: FAIL")
            except Exception as e:
                print(f"[IPv6] Test {i+1}/{test_count}: ERROR - {e}")

        # Decide IPv6 mode
        if working_count > 0:
            print(f"[IPv6] {working_count}/{test_count} tests passed - Using IPv6 Rotation")
            self._set_ipv6_enabled(True)
            return True
        else:
            print("[IPv6] All tests failed - Using Direct Connection")
            self._set_ipv6_enabled(False)
            return False

    def _set_ipv6_enabled(self, enabled: bool):
        """Update IPv6 enabled setting in settings.yaml and manager."""
        import yaml

        config_path = TOOL_DIR / "config" / "settings.yaml"
        try:
            config = {}
            if config_path.exists():
                with open(config_path, "r", encoding="utf-8") as f:
                    config = yaml.safe_load(f) or {}

            if 'ipv6_rotation' not in config:
                config['ipv6_rotation'] = {}
            config['ipv6_rotation']['enabled'] = enabled

            with open(config_path, "w", encoding="utf-8") as f:
                yaml.dump(config, f, default_flow_style=False, allow_unicode=True)

            # Update manager settings if exists
            if hasattr(self, 'manager') and self.manager:
                if hasattr(self.manager, 'settings') and hasattr(self.manager.settings, 'ipv6_rotation'):
                    self.manager.settings.ipv6_rotation['enabled'] = enabled

            status = "BAT" if enabled else "TAT"
            print(f"[IPv6] Settings updated: IPv6 = {status}")

        except Exception as e:
            print(f"[IPv6] Error saving settings: {e}")

    def _run_update(self):
        """Cap nhat code tu GitHub - ho tro ca khi khong co Git."""
        import subprocess
        import urllib.request
        import zipfile
        import shutil

        GITHUB_ZIP_URL = "https://github.com/nguyenvantuong161978-dotcom/ve3-tool-simple/archive/refs/heads/main.zip"
        GITHUB_GIT_URL = "https://github.com/nguyenvantuong161978-dotcom/ve3-tool-simple.git"

        def do_update():
            self.update_btn.config(state="disabled", text="DANG CAP NHAT...", bg='#666')
            self.status_var.set("Dang kiem tra...")

            try:
                # Kiem tra git co san khong
                git_available = False
                try:
                    result = subprocess.run(
                        ["git", "--version"],
                        capture_output=True,
                        text=True,
                        timeout=10
                    )
                    git_available = (result.returncode == 0)
                except:
                    git_available = False

                if git_available:
                    # === DUNG GIT ===
                    self.status_var.set("Dang cap nhat qua Git...")

                    # Kiem tra remote origin - dam bao URL dung
                    result = subprocess.run(
                        ["git", "remote", "get-url", "origin"],
                        cwd=str(TOOL_DIR),
                        capture_output=True,
                        text=True,
                        timeout=10
                    )

                    if result.returncode != 0:
                        # Chua co remote → them moi
                        subprocess.run(
                            ["git", "remote", "add", "origin", GITHUB_GIT_URL],
                            cwd=str(TOOL_DIR),
                            capture_output=True,
                            timeout=10
                        )
                    elif GITHUB_GIT_URL not in result.stdout.strip():
                        # Remote co nhung URL sai → sua lai
                        subprocess.run(
                            ["git", "remote", "set-url", "origin", GITHUB_GIT_URL],
                            cwd=str(TOOL_DIR),
                            capture_output=True,
                            timeout=10
                        )

                    # Fetch va reset
                    cmds = [
                        ["git", "fetch", "origin", "main"],
                        ["git", "checkout", "main"],
                        ["git", "reset", "--hard", "origin/main"]
                    ]

                    for cmd in cmds:
                        result = subprocess.run(
                            cmd,
                            cwd=str(TOOL_DIR),
                            capture_output=True,
                            text=True,
                            timeout=120
                        )
                else:
                    # === KHONG CO GIT - TAI ZIP ===
                    self.status_var.set("Dang tai ZIP tu GitHub...")

                    # Tai file zip
                    zip_path = TOOL_DIR / "update_temp.zip"
                    extract_dir = TOOL_DIR / "update_temp"

                    # Download - bo qua SSL certificate
                    import ssl
                    import time
                    ssl_context = ssl.create_default_context()
                    ssl_context.check_hostname = False
                    ssl_context.verify_mode = ssl.CERT_NONE

                    # Cache-busting: them timestamp vao URL
                    cache_buster = f"?t={int(time.time())}"
                    download_url = GITHUB_ZIP_URL + cache_buster

                    with urllib.request.urlopen(download_url, context=ssl_context) as response:
                        with open(str(zip_path), 'wb') as out_file:
                            out_file.write(response.read())

                    self.status_var.set("Dang giai nen...")

                    # Giai nen
                    with zipfile.ZipFile(str(zip_path), 'r') as zip_ref:
                        zip_ref.extractall(str(extract_dir))

                    # Tim thu muc da giai nen (ve3-tool-simple-main)
                    extracted_folder = extract_dir / "ve3-tool-simple-main"

                    self.status_var.set("Dang cap nhat files...")

                    # Copy files moi (chi copy .py va modules/)
                    files_to_update = [
                        "vm_manager.py",
                        "vm_manager_gui.py",
                        "run_excel_api.py",
                        "run_worker.py",
                        "START.py",
                        "START.bat",
                        "START_SERVER.bat",
                        "requirements.txt",
                        "_run_chrome1.py",
                        "_run_chrome2.py",
                        "google_login.py",
                        "VERSION.txt",
                    ]

                    for f in files_to_update:
                        src = extracted_folder / f
                        dst = TOOL_DIR / f
                        if src.exists():
                            shutil.copy2(str(src), str(dst))

                    # Copy modules folder (including subdirectories)
                    src_modules = extracted_folder / "modules"
                    dst_modules = TOOL_DIR / "modules"
                    if src_modules.exists():
                        for py_file in src_modules.glob("*.py"):
                            shutil.copy2(str(py_file), str(dst_modules / py_file.name))
                        # Copy subdirectories (e.g. modules/topic_prompts/)
                        for sub_dir in src_modules.iterdir():
                            if sub_dir.is_dir():
                                dst_sub = dst_modules / sub_dir.name
                                if dst_sub.exists():
                                    shutil.rmtree(str(dst_sub))
                                shutil.copytree(str(sub_dir), str(dst_sub))

                    # Copy server folder (local proxy server)
                    src_server = extracted_folder / "server"
                    dst_server = TOOL_DIR / "server"
                    if src_server.exists():
                        dst_server.mkdir(exist_ok=True)
                        for py_file in src_server.glob("*.py"):
                            shutil.copy2(str(py_file), str(dst_server / py_file.name))

                    # Copy ipv6 folder (IPv6 Dynamic Pool)
                    src_ipv6 = extracted_folder / "ipv6"
                    dst_ipv6 = TOOL_DIR / "ipv6"
                    if src_ipv6.exists():
                        dst_ipv6.mkdir(exist_ok=True)
                        for py_file in src_ipv6.glob("*.py"):
                            shutil.copy2(str(py_file), str(dst_ipv6 / py_file.name))

                    # Xoa temp files
                    if zip_path.exists():
                        zip_path.unlink()
                    if extract_dir.exists():
                        shutil.rmtree(str(extract_dir))

                # v1.0.438: KHÔNG reset video_mode/excel_mode khi update
                # Giữ nguyên settings mà user đã chọn trên VM
                # (Trước đây force reset về 'small' - gây mất settings)

                # Lay version moi sau khi update
                new_version = self._get_git_version()

                self.status_var.set("Cap nhat xong! Khoi dong lai tool.")
                self.update_btn.config(text="XONG", bg='#00ff88')

                # v1.0.343: Tự restart luôn, không hỏi
                import os
                os.execv(sys.executable, [sys.executable] + sys.argv)

            except Exception as e:
                self.status_var.set(f"Loi: {str(e)[:40]}")
                self.update_btn.config(text="LOI", bg='#e94560')
                print(f"Update error: {e}")

                from tkinter import messagebox
                messagebox.showerror("Loi cap nhat", f"Loi: {e}\n\nThu tai thu cong:\n{GITHUB_ZIP_URL}")
            finally:
                self.after(3000, lambda: self.update_btn.config(state="normal", text="UPDATE", bg='#0984e3'))

        threading.Thread(target=do_update, daemon=True).start()

    def _open_settings(self):
        """Mo cua so Settings."""
        SettingsWindow(self)

    # ================================================================
    # LOCAL PROXY SERVER - Chay/dung server tu GUI
    # ================================================================

    def _toggle_server(self):
        """Bat/tat Local Proxy Server."""
        if self._server_process and self._server_process.poll() is None:
            # Server dang chay → dung
            self._stop_server()
        else:
            # Server chua chay → chay
            self._start_server()

    def _start_server(self):
        """Chay server trong subprocess."""
        import subprocess

        # Check flask installed
        try:
            import flask
        except ImportError:
            from tkinter import messagebox
            messagebox.showerror("Thieu Flask",
                "Chua cai flask!\n\nChay lenh:\n  pip install flask\n\nRoi thu lai.")
            return

        python_exe = sys.executable
        server_script = str(TOOL_DIR / "server" / "app.py")

        try:
            # Chay server trong subprocess rieng
            self._server_process = subprocess.Popen(
                [python_exe, "-u", server_script],
                cwd=str(TOOL_DIR),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                creationflags=subprocess.CREATE_NEW_CONSOLE,
            )

            self.server_btn.config(text="DUNG SERVER", bg='#e94560')
            self.server_status_dot.config(text="SERVER DANG CHAY", fg='#00ff88')
            self.status_var.set("Server dang khoi dong...")

            # Monitor server process
            self._monitor_server()

        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Loi", f"Khong chay duoc server:\n{e}")

    def _stop_server(self):
        """Dung server."""
        if self._server_process:
            try:
                self._server_process.terminate()
                self._server_process.wait(timeout=5)
            except Exception:
                try:
                    self._server_process.kill()
                except Exception:
                    pass
            self._server_process = None

        self.server_btn.config(text="CHAY SERVER", bg='#fd79a8')
        self.server_status_dot.config(text="", fg='#888')
        self.status_var.set("Server da dung")

    def _monitor_server(self):
        """Kiem tra server con chay khong."""
        if self._server_process and self._server_process.poll() is not None:
            # Server da tat
            exit_code = self._server_process.returncode
            self._server_process = None
            self.server_btn.config(text="CHAY SERVER", bg='#fd79a8')
            self.server_status_dot.config(text=f"DA TAT (code={exit_code})", fg='#e94560')
            return
        if self._server_process:
            self.after(3000, self._monitor_server)

    def _setup_vm(self):
        """Setup SMB share + IPv6/Proxy cho may ao."""
        import tkinter.messagebox as msgbox
        import yaml
        import subprocess
        import threading

        # Default settings
        default_ip = "192.168.88.254"
        default_share = "D"
        default_user = "smbuser"
        default_pass = "159753"

        # Tao popup
        popup = tk.Toplevel(self)
        popup.title("Setup VM")
        popup.geometry("550x750")
        popup.configure(bg='#1a1a2e')
        popup.transient(self)
        popup.grab_set()

        # Center popup
        popup.update_idletasks()
        x = self.winfo_x() + (self.winfo_width() - 550) // 2
        y = self.winfo_y() + (self.winfo_height() - 750) // 2
        popup.geometry(f"+{x}+{y}")

        # Scrollable content
        canvas = tk.Canvas(popup, bg='#1a1a2e', highlightthickness=0)
        scrollbar = tk.Scrollbar(popup, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas, bg='#1a1a2e')

        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw", width=530)
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Mouse wheel scroll
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        popup.bind("<Destroy>", lambda e: canvas.unbind_all("<MouseWheel>") if e.widget == popup else None)

        # ============ SECTION 1: SMB SHARE ============
        tk.Label(scroll_frame, text="1. KET NOI O MANG (SMB SHARE)",
                bg='#1a1a2e', fg='#00ff88', font=("Arial", 11, "bold")).pack(pady=(10, 5))

        # Form frame
        form = tk.Frame(scroll_frame, bg='#1a1a2e')
        form.pack(pady=5, padx=20, fill="x")

        # IP
        tk.Label(form, text="IP May chu:", bg='#1a1a2e', fg='white', font=("Arial", 9)).grid(row=0, column=0, sticky="e", pady=3)
        ip_var = tk.StringVar(value=default_ip)
        tk.Entry(form, textvariable=ip_var, width=22, font=("Arial", 9)).grid(row=0, column=1, pady=3, padx=5)

        # Share name
        tk.Label(form, text="Ten Share:", bg='#1a1a2e', fg='white', font=("Arial", 9)).grid(row=1, column=0, sticky="e", pady=3)
        share_var = tk.StringVar(value=default_share)
        tk.Entry(form, textvariable=share_var, width=22, font=("Arial", 9)).grid(row=1, column=1, pady=3, padx=5)

        # Username
        tk.Label(form, text="Username:", bg='#1a1a2e', fg='white', font=("Arial", 9)).grid(row=2, column=0, sticky="e", pady=3)
        user_var = tk.StringVar(value=default_user)
        tk.Entry(form, textvariable=user_var, width=22, font=("Arial", 9)).grid(row=2, column=1, pady=3, padx=5)

        # Password
        tk.Label(form, text="Password:", bg='#1a1a2e', fg='white', font=("Arial", 9)).grid(row=3, column=0, sticky="e", pady=3)
        pass_var = tk.StringVar(value=default_pass)
        tk.Entry(form, textvariable=pass_var, width=22, font=("Arial", 9), show="*").grid(row=3, column=1, pady=3, padx=5)

        # SMB Status
        smb_status_var = tk.StringVar(value="")
        smb_status_lbl = tk.Label(scroll_frame, textvariable=smb_status_var, bg='#1a1a2e', fg='#ffd93d', font=("Arial", 9))
        smb_status_lbl.pack(pady=3)

        def do_smb_setup():
            """Thuc hien ket noi SMB."""
            ip = ip_var.get().strip()
            share = share_var.get().strip()
            user = user_var.get().strip()
            passwd = pass_var.get().strip()

            if not all([ip, share, user, passwd]):
                smb_status_var.set("Vui long nhap day du thong tin!")
                smb_status_lbl.config(fg='#e94560')
                return

            smb_status_var.set("Dang ket noi...")
            smb_status_lbl.config(fg='#ffd93d')
            popup.update()

            try:
                subprocess.run(['net', 'use', 'Z:', '/delete', '/y'], capture_output=True, text=True)
                cmd = ['net', 'use', 'Z:', f'\\\\{ip}\\{share}', f'/user:{user}', passwd, '/persistent:yes']
                result = subprocess.run(cmd, capture_output=True, text=True)

                if result.returncode == 0:
                    auto_path = TOOL_DIR.parent.parent / "Z:" / "AUTO"
                    try:
                        from pathlib import Path
                        auto_path = Path("Z:\\AUTO")
                        if auto_path.exists():
                            smb_status_var.set("THANH CONG! Z:\\AUTO da san sang")
                            smb_status_lbl.config(fg='#00ff88')
                        else:
                            smb_status_var.set("Da ket noi nhung khong tim thay AUTO")
                            smb_status_lbl.config(fg='#ffd93d')
                    except:
                        smb_status_var.set("Da ket noi Z:")
                        smb_status_lbl.config(fg='#00ff88')
                else:
                    smb_status_var.set(f"LOI: {(result.stderr or result.stdout or 'Unknown')[:40]}")
                    smb_status_lbl.config(fg='#e94560')
            except Exception as e:
                smb_status_var.set(f"LOI: {str(e)[:40]}")
                smb_status_lbl.config(fg='#e94560')

        tk.Button(scroll_frame, text="KET NOI SMB", command=do_smb_setup,
                 bg='#00ff88', fg='#1a1a2e', font=("Arial", 9, "bold"),
                 relief="flat", padx=15, pady=3).pack(pady=5)

        # ============ SEPARATOR ============
        tk.Frame(scroll_frame, bg='#444', height=2).pack(fill="x", padx=20, pady=10)

        # ============ SECTION 2: PROXY PROVIDER ============
        tk.Label(scroll_frame, text="2. PROXY / IP ROTATION",
                bg='#1a1a2e', fg='#00ff88', font=("Arial", 11, "bold")).pack(pady=(5, 5))

        # --- Proxy Type Selector ---
        proxy_type_frame = tk.Frame(scroll_frame, bg='#1a1a2e')
        proxy_type_frame.pack(pady=5, padx=20, fill="x")

        tk.Label(proxy_type_frame, text="Loai Proxy:", bg='#1a1a2e', fg='white',
                 font=("Arial", 10)).pack(side="left", padx=(0, 10))

        # Read current proxy_provider config
        config_path = TOOL_DIR / "config" / "settings.yaml"
        current_proxy_type = "ipv6"  # default
        current_ws_username = ""
        current_ws_password = ""
        current_ws_machine_id = "1"
        current_pool_api_url = ""
        try:
            if config_path.exists():
                with open(config_path, "r", encoding="utf-8") as f:
                    _cfg = yaml.safe_load(f) or {}
                pp_cfg = _cfg.get('proxy_provider', {})
                if pp_cfg:
                    current_proxy_type = pp_cfg.get('type', 'ipv6')
                    ws_cfg = pp_cfg.get('webshare', {})
                    current_ws_username = ws_cfg.get('rotating_username', '')
                    current_ws_password = ws_cfg.get('rotating_password', '')
                    current_ws_machine_id = str(ws_cfg.get('machine_id', 1))
                else:
                    # Backward compat
                    ipv6_cfg = _cfg.get('ipv6_rotation', {})
                    current_proxy_type = "ipv6" if ipv6_cfg.get('enabled', False) else "none"
                # v1.0.562: Pool API URL
                mikrotik_cfg = _cfg.get('mikrotik', {})
                current_pool_api_url = mikrotik_cfg.get('pool_api_url', '')
                if current_pool_api_url and current_proxy_type == 'ipv6':
                    current_proxy_type = 'ipv6_pool'
        except:
            pass

        proxy_type_var = tk.StringVar(value=current_proxy_type)
        proxy_types = [("Khong dung", "none"), ("IPv6 File", "ipv6"), ("IPv6 Pool", "ipv6_pool"), ("Webshare", "webshare")]

        for text, val in proxy_types:
            tk.Radiobutton(proxy_type_frame, text=text, variable=proxy_type_var, value=val,
                          bg='#1a1a2e', fg='white', selectcolor='#1a1a2e', font=("Arial", 9),
                          activebackground='#1a1a2e', activeforeground='white',
                          command=lambda: _on_proxy_type_changed()
                          ).pack(side="left", padx=8)

        # --- Webshare Settings Frame ---
        ws_frame = tk.LabelFrame(scroll_frame, text=" Webshare.io Settings ", bg='#16213e', fg='#ffd93d',
                                  font=("Arial", 9, "bold"), padx=10, pady=8)

        ws_form = tk.Frame(ws_frame, bg='#16213e')
        ws_form.pack(fill="x", pady=5)

        tk.Label(ws_form, text="Username:", bg='#16213e', fg='white', font=("Arial", 9)).grid(row=0, column=0, sticky="e", pady=3)
        ws_username_var = tk.StringVar(value=current_ws_username)
        tk.Entry(ws_form, textvariable=ws_username_var, width=35, font=("Consolas", 9),
                 bg='#0f3460', fg='white', insertbackground='white').grid(row=0, column=1, pady=3, padx=5)

        tk.Label(ws_form, text="Password:", bg='#16213e', fg='white', font=("Arial", 9)).grid(row=1, column=0, sticky="e", pady=3)
        ws_password_var = tk.StringVar(value=current_ws_password)
        tk.Entry(ws_form, textvariable=ws_password_var, width=35, font=("Consolas", 9),
                 bg='#0f3460', fg='white', insertbackground='white').grid(row=1, column=1, pady=3, padx=5)

        tk.Label(ws_form, text="Machine ID:", bg='#16213e', fg='white', font=("Arial", 9)).grid(row=2, column=0, sticky="e", pady=3)
        ws_machine_var = tk.StringVar(value=current_ws_machine_id)
        tk.Entry(ws_form, textvariable=ws_machine_var, width=5, font=("Consolas", 9),
                 bg='#0f3460', fg='white', insertbackground='white').grid(row=2, column=1, pady=3, padx=5, sticky="w")

        tk.Label(ws_frame, text="Webshare.io Rotating Residential - Doi IP bang session ID",
                 bg='#16213e', fg='#888', font=("Arial", 8)).pack(anchor="w")

        # Webshare test status
        ws_status_var = tk.StringVar(value="")
        ws_status_lbl = tk.Label(ws_frame, textvariable=ws_status_var, bg='#16213e', fg='#ffd93d', font=("Arial", 9))
        ws_status_lbl.pack(anchor="w", pady=3)

        def _test_webshare():
            """Test Webshare proxy connectivity."""
            username = ws_username_var.get().strip()
            password = ws_password_var.get().strip()
            if not username or not password:
                ws_status_var.set("Nhap username va password!")
                ws_status_lbl.config(fg='#e94560')
                return
            ws_status_var.set("Dang test...")
            ws_status_lbl.config(fg='#ffd93d')

            def _do_test():
                try:
                    from modules.proxy_providers.webshare_provider import WebshareProvider
                    provider = WebshareProvider(config={'webshare': {
                        'rotating_username': username,
                        'rotating_password': password,
                        'machine_id': int(ws_machine_var.get() or 1),
                    }})
                    ok = provider.test_connectivity()
                    def _show():
                        if ok:
                            ws_status_var.set("OK! Ket noi thanh cong")
                            ws_status_lbl.config(fg='#00ff88')
                        else:
                            ws_status_var.set("THAT BAI! Kiem tra lai thong tin")
                            ws_status_lbl.config(fg='#e94560')
                    popup.after(0, _show)
                except Exception as e:
                    popup.after(0, lambda: (ws_status_var.set(f"LOI: {str(e)[:40]}"), ws_status_lbl.config(fg='#e94560')))

            threading.Thread(target=_do_test, daemon=True).start()

        tk.Button(ws_frame, text="TEST KET NOI", command=_test_webshare,
                 bg='#6c5ce7', fg='white', font=("Arial", 8, "bold"),
                 relief="flat", padx=10, pady=2).pack(anchor="w", pady=3)

        # --- v1.0.562: IPv6 Pool section ---
        pool_frame = tk.LabelFrame(scroll_frame, text=" IPv6 Pool (MikroTik) ", bg='#16213e', fg='#ffd93d',
                                    font=("Arial", 9, "bold"), padx=10, pady=8)

        pool_form = tk.Frame(pool_frame, bg='#16213e')
        pool_form.pack(fill="x", pady=5)

        tk.Label(pool_form, text="Pool API URL:", bg='#16213e', fg='white', font=("Arial", 9)).grid(row=0, column=0, sticky="e", pady=3)
        pool_api_var = tk.StringVar(value=current_pool_api_url)
        tk.Entry(pool_form, textvariable=pool_api_var, width=35, font=("Consolas", 9),
                 bg='#0f3460', fg='white', insertbackground='white').grid(row=0, column=1, pady=3, padx=5)

        tk.Label(pool_frame, text="VD: http://192.168.88.1:8765 - Lay IPv6 tu MikroTik Pool API",
                 bg='#16213e', fg='#888', font=("Arial", 8)).pack(anchor="w")

        # Pool status + stats
        pool_status_var = tk.StringVar(value="")
        pool_status_lbl = tk.Label(pool_frame, textvariable=pool_status_var, bg='#16213e', fg='#ffd93d', font=("Arial", 9))
        pool_status_lbl.pack(anchor="w", pady=3)

        pool_stats_var = tk.StringVar(value="")
        pool_stats_lbl = tk.Label(pool_frame, textvariable=pool_stats_var, bg='#16213e', fg='#888', font=("Consolas", 8),
                                   justify="left")
        pool_stats_lbl.pack(anchor="w", pady=2)

        def _test_pool():
            """Test ket noi Pool API va hien thi thong ke."""
            url = pool_api_var.get().strip()
            if not url:
                pool_status_var.set("Nhap Pool API URL!")
                pool_status_lbl.config(fg='#e94560')
                return
            pool_status_var.set("Dang ket noi...")
            pool_status_lbl.config(fg='#ffd93d')
            popup.update()

            def _do_test():
                try:
                    from modules.ipv6_pool_client import IPv6PoolClient
                    import requests

                    # Test 1: Raw HTTP request to show exact error
                    test_url = f"{url.rstrip('/')}/api/ping"
                    try:
                        resp = requests.get(test_url, timeout=5)
                        if resp.status_code == 200:
                            data = resp.json()
                            if data.get("ok"):
                                # Ping OK - get status
                                client = IPv6PoolClient(api_url=url, timeout=5)
                                status = client.get_status()
                                def _show():
                                    pool_status_var.set("OK! Ket noi Pool thanh cong")
                                    pool_status_lbl.config(fg='#00ff88')
                                    if status:
                                        avail = status.get('available', 0)
                                        in_use = status.get('in_use', 0)
                                        burned = status.get('burned', 0)
                                        total = avail + in_use + burned
                                        stats_text = (
                                            f"Tong: {total} IP | San sang: {avail} | "
                                            f"Dang dung: {in_use} | Da burn: {burned}"
                                        )
                                        pool_stats_var.set(stats_text)
                                popup.after(0, _show)
                                return
                            else:
                                err_msg = f"API tra ve: {data}"
                        else:
                            err_msg = f"HTTP {resp.status_code}"
                    except requests.exceptions.ConnectionError:
                        err_msg = f"Khong ket noi duoc {url} - Kiem tra IP/port va firewall"
                    except requests.exceptions.Timeout:
                        err_msg = f"Timeout sau 5s - Server cham hoac khong phan hoi"
                    except Exception as e:
                        err_msg = str(e)[:80]

                    popup.after(0, lambda: (
                        pool_status_var.set(f"THAT BAI! {err_msg[:60]}"),
                        pool_status_lbl.config(fg='#e94560'),
                        pool_stats_var.set("")
                    ))
                except Exception as e:
                    popup.after(0, lambda: (
                        pool_status_var.set(f"LOI: {str(e)[:50]}"),
                        pool_status_lbl.config(fg='#e94560'),
                        pool_stats_var.set("")
                    ))

            threading.Thread(target=_do_test, daemon=True).start()

        def _rotate_pool_manual():
            """Doi IPv6 thu cong qua Pool API."""
            url = pool_api_var.get().strip()
            if not url:
                pool_status_var.set("Nhap Pool API URL!")
                pool_status_lbl.config(fg='#e94560')
                return
            pool_status_var.set("Dang doi IPv6...")
            pool_status_lbl.config(fg='#ffd93d')
            popup.update()

            def _do_rotate():
                try:
                    from modules.ipv6_pool_client import IPv6PoolClient
                    client = IPv6PoolClient(api_url=url, timeout=5)

                    # Lay IP hien tai tu rotator (neu co)
                    current_ip = None
                    try:
                        from modules.ipv6_rotator import get_ipv6_rotator
                        rotator = get_ipv6_rotator()
                        if rotator:
                            current_ip = rotator.get_current_ipv6()
                    except:
                        pass

                    if current_ip:
                        new_ip = client.rotate_ip(current_ip, reason="manual", worker="vm_gui")
                    else:
                        new_ip = client.get_ip(worker="vm_gui")

                    if new_ip:
                        # Thu set IPv6 len may
                        try:
                            from modules.ipv6_rotator import get_ipv6_rotator
                            rotator = get_ipv6_rotator()
                            if rotator:
                                ok = rotator.set_ipv6(new_ip)
                                if ok:
                                    rotator.current_ipv6 = new_ip
                                    popup.after(0, lambda: (
                                        pool_status_var.set(f"DA DOI: {new_ip}"),
                                        pool_status_lbl.config(fg='#00ff88')
                                    ))
                                    return
                        except:
                            pass
                        popup.after(0, lambda: (
                            pool_status_var.set(f"LAY DUOC: {new_ip} (chua set len may)"),
                            pool_status_lbl.config(fg='#ffd93d')
                        ))
                    else:
                        popup.after(0, lambda: (
                            pool_status_var.set("THAT BAI! Pool het IP hoac loi"),
                            pool_status_lbl.config(fg='#e94560')
                        ))
                except Exception as e:
                    popup.after(0, lambda: (
                        pool_status_var.set(f"LOI: {str(e)[:50]}"),
                        pool_status_lbl.config(fg='#e94560')
                    ))

            threading.Thread(target=_do_rotate, daemon=True).start()

        pool_btn_frame = tk.Frame(pool_frame, bg='#16213e')
        pool_btn_frame.pack(anchor="w", pady=5)

        tk.Button(pool_btn_frame, text="TEST KET NOI", command=_test_pool,
                 bg='#6c5ce7', fg='white', font=("Arial", 8, "bold"),
                 relief="flat", padx=10, pady=2).pack(side="left", padx=(0, 8))

        tk.Button(pool_btn_frame, text="DOI IPv6 THU CONG", command=_rotate_pool_manual,
                 bg='#e17055', fg='white', font=("Arial", 8, "bold"),
                 relief="flat", padx=10, pady=2).pack(side="left", padx=(0, 8))

        # --- IPv6 section (existing - file based) ---
        ipv6_section_frame = tk.Frame(scroll_frame, bg='#1a1a2e')

        # Show/hide based on proxy type
        def _on_proxy_type_changed():
            ptype = proxy_type_var.get()
            if ptype == "webshare":
                ws_frame.pack(fill="x", padx=20, pady=5, after=proxy_type_frame)
                pool_frame.pack_forget()
                ipv6_section_frame.pack_forget()
            elif ptype == "ipv6_pool":
                ws_frame.pack_forget()
                pool_frame.pack(fill="x", padx=20, pady=5, after=proxy_type_frame)
                ipv6_section_frame.pack_forget()
            elif ptype == "ipv6":
                ws_frame.pack_forget()
                pool_frame.pack_forget()
                ipv6_section_frame.pack(fill="x", padx=0, pady=5, after=proxy_type_frame)
            else:
                ws_frame.pack_forget()
                pool_frame.pack_forget()
                ipv6_section_frame.pack_forget()

        # IPv6 content inside ipv6_section_frame
        tk.Label(ipv6_section_frame, text="IPv6 ROTATION",
                bg='#1a1a2e', fg='#00ff88', font=("Arial", 10, "bold")).pack(pady=(5, 5))

        # Read current IPv6 setting
        ipv6_file = TOOL_DIR / "config" / "ipv6.txt"
        current_ipv6_enabled = self._get_ipv6_setting()

        # Radio buttons for IPv6 mode
        ipv6_mode_var = tk.IntVar(value=1 if current_ipv6_enabled else 0)

        mode_frame = tk.Frame(ipv6_section_frame, bg='#1a1a2e')
        mode_frame.pack(pady=5)

        tk.Radiobutton(mode_frame, text="Khong dung IPv6 (Direct)", variable=ipv6_mode_var, value=0,
                      bg='#1a1a2e', fg='white', selectcolor='#1a1a2e', font=("Arial", 9),
                      activebackground='#1a1a2e', activeforeground='white').pack(side="left", padx=10)
        tk.Radiobutton(mode_frame, text="Dung IPv6 Rotation", variable=ipv6_mode_var, value=1,
                      bg='#1a1a2e', fg='#00ff88', selectcolor='#1a1a2e', font=("Arial", 9, "bold"),
                      activebackground='#1a1a2e', activeforeground='#00ff88').pack(side="left", padx=10)

        # IPv6 list frame
        ipv6_frame = tk.Frame(ipv6_section_frame, bg='#1a1a2e')
        ipv6_frame.pack(pady=5, padx=20, fill="x")

        tk.Label(ipv6_frame, text="Danh sach IPv6 (config/ipv6.txt):",
                bg='#1a1a2e', fg='#aaa', font=("Arial", 9)).pack(anchor="w")

        # Text widget for IPv6 list
        ipv6_text = tk.Text(ipv6_frame, height=8, width=50, bg='#2a2a4e', fg='white',
                           font=("Consolas", 9), insertbackground='white')
        ipv6_text.pack(fill="x", pady=5)

        # Load IPv6 list from file
        if ipv6_file.exists():
            with open(ipv6_file, "r", encoding="utf-8") as f:
                ipv6_text.insert("1.0", f.read())

        # IPv6 test status
        ipv6_status_var = tk.StringVar(value="")
        ipv6_status_lbl = tk.Label(ipv6_section_frame, textvariable=ipv6_status_var, bg='#1a1a2e', fg='#ffd93d', font=("Arial", 9))
        ipv6_status_lbl.pack(pady=3)

        # Test results frame
        test_results_var = tk.StringVar(value="")
        test_results_lbl = tk.Label(ipv6_section_frame, textvariable=test_results_var, bg='#1a1a2e', fg='#888', font=("Consolas", 8))
        test_results_lbl.pack(pady=2)

        def test_ipv6():
            """Test IPv6 connectivity for all IPs in list."""
            ipv6_list = [line.strip() for line in ipv6_text.get("1.0", "end").split("\n") if line.strip() and not line.strip().startswith("#")]

            if not ipv6_list:
                ipv6_status_var.set("Khong co IPv6 de test!")
                ipv6_status_lbl.config(fg='#e94560')
                return

            ipv6_status_var.set(f"Dang test {len(ipv6_list)} IPv6...")
            ipv6_status_lbl.config(fg='#ffd93d')
            popup.update()

            def run_test():
                working = []
                failed = []

                for i, ipv6 in enumerate(ipv6_list):
                    try:
                        # Ping Google DNS IPv6
                        cmd = f'ping -n 1 -w 2000 2001:4860:4860::8888'
                        result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=5)
                        if result.returncode == 0 and 'Reply from' in result.stdout:
                            working.append(ipv6)
                        else:
                            failed.append(ipv6)
                    except:
                        failed.append(ipv6)

                    # Update progress
                    popup.after(0, lambda i=i: ipv6_status_var.set(f"Testing... {i+1}/{len(ipv6_list)}"))

                # Update final results
                def show_results():
                    if working:
                        ipv6_status_var.set(f"XONG: {len(working)} OK, {len(failed)} FAIL")
                        ipv6_status_lbl.config(fg='#00ff88')
                        test_results_var.set(f"Working: {', '.join(working[:3])}{'...' if len(working) > 3 else ''}")
                    else:
                        ipv6_status_var.set("KHONG CO IPv6 NAO HOAT DONG!")
                        ipv6_status_lbl.config(fg='#e94560')
                        test_results_var.set("Kiem tra lai danh sach IPv6 hoac ket noi mang")

                popup.after(0, show_results)

            threading.Thread(target=run_test, daemon=True).start()

        # Save status label (dung chung cho ca save)
        save_status_var = tk.StringVar(value="")
        save_status_lbl = tk.Label(scroll_frame, textvariable=save_status_var, bg='#1a1a2e', fg='#ffd93d', font=("Arial", 9))

        def save_proxy_config():
            """Save proxy provider + IPv6 settings to config."""
            try:
                # Save IPv6 list to file
                ipv6_content = ipv6_text.get("1.0", "end").strip()
                with open(ipv6_file, "w", encoding="utf-8") as f:
                    f.write(ipv6_content)

                # Read existing config
                config = {}
                if config_path.exists():
                    with open(config_path, "r", encoding="utf-8") as f:
                        config = yaml.safe_load(f) or {}

                # Save proxy_provider section
                ptype = proxy_type_var.get()
                # Map ipv6_pool back to ipv6 for proxy_provider (pool is IPv6 source, not provider type)
                pp_type = 'ipv6' if ptype == 'ipv6_pool' else ptype
                if 'proxy_provider' not in config:
                    config['proxy_provider'] = {}
                config['proxy_provider']['type'] = pp_type

                # Webshare settings
                config['proxy_provider']['webshare'] = {
                    'rotating_host': 'p.webshare.io',
                    'rotating_port': 80,
                    'rotating_username': ws_username_var.get().strip(),
                    'rotating_password': ws_password_var.get().strip(),
                    'machine_id': int(ws_machine_var.get() or 1),
                }

                # v1.0.562: Pool API URL → mikrotik section
                if 'mikrotik' not in config:
                    config['mikrotik'] = {}
                if ptype == 'ipv6_pool':
                    config['mikrotik']['pool_api_url'] = pool_api_var.get().strip()
                else:
                    config['mikrotik']['pool_api_url'] = ''

                # IPv6 backward compat
                if 'ipv6_rotation' not in config:
                    config['ipv6_rotation'] = {}
                config['ipv6_rotation']['enabled'] = (ptype in ('ipv6', 'ipv6_pool') and
                                                       (ptype == 'ipv6_pool' or ipv6_mode_var.get() == 1))

                with open(config_path, "w", encoding="utf-8") as f:
                    yaml.dump(config, f, default_flow_style=False, allow_unicode=True)

                # Update manager settings
                if hasattr(self, 'manager') and self.manager:
                    if hasattr(self.manager, 'settings'):
                        if hasattr(self.manager.settings, 'config'):
                            self.manager.settings.config = config.copy()

                type_labels = {'none': 'KHONG DUNG', 'ipv6': 'IPv6 FILE', 'ipv6_pool': 'IPv6 POOL', 'webshare': 'WEBSHARE'}
                save_status_var.set(f"DA LUU! Proxy: {type_labels.get(ptype, ptype)}")
                save_status_lbl.config(fg='#00ff88')
                print(f"[GUI] Proxy config saved: type={ptype}")

            except Exception as e:
                save_status_var.set(f"LOI: {str(e)[:40]}")
                save_status_lbl.config(fg='#e94560')

        def fetch_ipv6_from_sheet():
            """Lay IPv6 tu trang tinh THONG TIN."""
            try:
                ipv6_status_var.set("Dang lay IPv6 tu trang tinh...")
                ipv6_status_lbl.config(fg='#ffd93d')
                popup.update()

                from google_login import detect_machine_code, get_channel_ipv6

                machine_code = detect_machine_code()
                if not machine_code:
                    ipv6_status_var.set("LOI: Khong detect duoc ma may!")
                    ipv6_status_lbl.config(fg='#e94560')
                    return

                ipv6_list = get_channel_ipv6(machine_code)
                if not ipv6_list:
                    ipv6_status_var.set(f"Khong tim thay IPv6 cho {machine_code}")
                    ipv6_status_lbl.config(fg='#e94560')
                    return

                # Dien vao text widget
                ipv6_text.delete("1.0", "end")
                ipv6_text.insert("1.0", '\n'.join(ipv6_list))

                ipv6_status_var.set(f"Da lay {len(ipv6_list)} IPv6 cho {machine_code} - BÁM LƯU ĐỂ ÁP DỤNG")
                ipv6_status_lbl.config(fg='#00ff88')

            except Exception as e:
                ipv6_status_var.set(f"LOI: {str(e)[:50]}")
                ipv6_status_lbl.config(fg='#e94560')

        # Buttons for IPv6
        ipv6_btn_frame = tk.Frame(ipv6_section_frame, bg='#1a1a2e')
        ipv6_btn_frame.pack(pady=5)

        tk.Button(ipv6_btn_frame, text="LAY TU TRANG TINH", command=fetch_ipv6_from_sheet,
                 bg='#e17055', fg='white', font=("Arial", 9, "bold"),
                 relief="flat", padx=12, pady=3).pack(side="left", padx=5)

        tk.Button(ipv6_btn_frame, text="TEST IPv6", command=test_ipv6,
                 bg='#6c5ce7', fg='white', font=("Arial", 9, "bold"),
                 relief="flat", padx=15, pady=3).pack(side="left", padx=5)

        # --- Trigger initial show/hide ---
        _on_proxy_type_changed()

        # --- Bottom buttons (LUU + DONG) - outside proxy sections ---
        tk.Frame(scroll_frame, bg='#444', height=2).pack(fill="x", padx=20, pady=10)

        save_status_lbl.pack(pady=3)

        bottom_btn_frame = tk.Frame(scroll_frame, bg='#1a1a2e')
        bottom_btn_frame.pack(pady=10)

        tk.Button(bottom_btn_frame, text="LUU CAU HINH", command=save_proxy_config,
                 bg='#00ff88', fg='#1a1a2e', font=("Arial", 10, "bold"),
                 relief="flat", padx=20, pady=5).pack(side="left", padx=10)

        tk.Button(bottom_btn_frame, text="DONG", command=popup.destroy,
                 bg='#e94560', fg='white', font=("Arial", 10, "bold"),
                 relief="flat", padx=20, pady=5).pack(side="left", padx=10)

    def _position_tool_window(self):
        """Dat cua so VE3 tool vao goc trai tren man hinh."""
        try:
            import ctypes
            user32 = ctypes.windll.user32
            screen_h = user32.GetSystemMetrics(1)
            # Tool chiem ~42% chieu cao (phan con lai cho CMDs)
            tool_h = int(screen_h * 0.42)
            self._tool_h = tool_h
            self.geometry(f"700x{tool_h}+0+0")
        except Exception as e:
            print(f"[GUI] _position_tool_window error: {e}")

    def _arrange_windows(self):
        """Sap xep tat ca cua so theo Layout A: Tool+CMDs trai, Chrome phai."""
        if not self.manager:
            return
        try:
            import ctypes
            user32 = ctypes.windll.user32
            screen_h = user32.GetSystemMetrics(1)
            tool_h = getattr(self, '_tool_h', int(screen_h * 0.42))
            self.manager.arrange_all_windows(
                tool_x=0, tool_y=0, tool_w=700, tool_h=tool_h
            )
        except Exception as e:
            print(f"[GUI] _arrange_windows error: {e}")

    def _draw_progress(self, canvas: tk.Canvas, pct: float):
        """Ve thanh progress tren Canvas (pct = 0.0 den 1.0)."""
        try:
            canvas.delete("all")
            w = canvas.winfo_width() or 110
            h = canvas.winfo_height() or 10
            # Background
            canvas.create_rectangle(0, 0, w, h, fill='#2a2a4e', outline='')
            # Progress fill
            if pct > 0:
                fill_w = max(1, int(w * min(1.0, pct)))
                color = '#00ff88' if pct >= 1.0 else '#00d9ff'
                canvas.create_rectangle(0, 0, fill_w, h, fill=color, outline='')
        except:
            pass

    def _clear_log(self):
        """Xoa noi dung log."""
        try:
            self.log_text.config(state="normal")
            self.log_text.delete('1.0', tk.END)
            self.log_text.config(state="disabled")
        except:
            pass

    def _update_log(self):
        """Cap nhat log area tu central logger."""
        if not LOGGER_AVAILABLE:
            return
        try:
            lines = tail_log(20)
            if not lines:
                return
            content = '\n'.join(lines)
            self.log_text.config(state="normal")
            self.log_text.delete('1.0', tk.END)
            self.log_text.insert('1.0', content)
            self.log_text.config(state="disabled")
            self.log_text.see(tk.END)
        except:
            pass

    def _on_new_log(self, line: str):
        """Callback khi co log moi - xu ly trong _update_log theo lich."""
        pass

    def _update_loop(self):
        self._update_workers()
        self._update_projects()
        self._auto_arrange_on_new_window()
        self.after(1000, self._update_loop)

    def _auto_arrange_on_new_window(self):
        """Tu dong sap xep khi phat hien Chrome hoac CMD moi mo (HWND moi)."""
        if not self.manager or not self.running:
            return
        try:
            current_chrome = set(self.manager.get_chrome_windows())
            current_cmd = set(self.manager.get_cmd_windows())

            new_chrome = current_chrome - self._known_chrome_hwnds
            new_cmd = current_cmd - self._known_cmd_hwnds

            self._known_chrome_hwnds = current_chrome
            self._known_cmd_hwnds = current_cmd

            if new_chrome or new_cmd:
                # Co cua so moi → sap xep nhieu lan (Chrome can thoi gian load)
                self.after(3000, self._arrange_windows)
                self.after(8000, self._arrange_windows)
        except Exception:
            pass

    def _update_workers(self):
        if not self.manager:
            return

        active_project = ""

        for wid in ["excel", "chrome_1", "chrome_2"]:
            try:
                status = self.manager.get_worker_status(wid)
                if status:
                    proj = status.get('current_project', '') or ''
                    state = status.get('state', 'idle')
                    step = status.get('current_step', 0)
                    step_name = status.get('step_name', '')
                    is_active = proj and state.lower() in ['working', 'busy']

                    if is_active:
                        # ACTIVE state
                        self.worker_vars[f"{wid}_project"].set(proj)
                        self.worker_labels[wid]['project'].config(fg='#00d9ff')
                        self.worker_labels[wid]['name'].config(fg='#00ff88')
                        self.worker_labels[wid]['badge'].config(text="▶ RUNNING", fg='#00ff88', bg='#16213e')
                        if wid in self.worker_dots:
                            self.worker_dots[wid].config(fg='#00ff88')
                        if proj:
                            active_project = proj

                        # Status text & progress
                        if wid == "excel" and step > 0:
                            short_name = step_name[:22] if step_name else ""
                            self.worker_vars[f"{wid}_status"].set(f"Step {step}/7 {short_name}")
                            pct = step / 7.0
                        else:
                            current_scene = status.get('current_scene', 0)
                            total_scenes = status.get('total_scenes', 0)
                            completed = status.get('completed_count', 0)
                            if current_scene and current_scene > 0 and total_scenes > 0:
                                self.worker_vars[f"{wid}_status"].set(f"S{current_scene} ({completed}/{total_scenes})")
                                pct = completed / total_scenes
                            elif completed > 0:
                                self.worker_vars[f"{wid}_status"].set(f"{completed} done")
                                pct = 0.0
                            else:
                                self.worker_vars[f"{wid}_status"].set("Dang khoi dong...")
                                pct = 0.0

                        # Draw progress bar
                        self._draw_progress(self.worker_progress[wid], pct)
                        self.worker_labels[wid]['pct'].config(text=f"{int(pct*100)}%")

                    else:
                        # IDLE state
                        self.worker_vars[f"{wid}_project"].set("")
                        self.worker_vars[f"{wid}_status"].set("")
                        self.worker_labels[wid]['name'].config(fg='#555')
                        self.worker_labels[wid]['badge'].config(text="■ IDLE", fg='#444', bg='#1a1a2e')
                        if wid in self.worker_dots:
                            self.worker_dots[wid].config(fg='#555')
                        self._draw_progress(self.worker_progress[wid], 0.0)
                        self.worker_labels[wid]['pct'].config(text="")
                else:
                    # No status yet
                    self.worker_labels[wid]['badge'].config(text="■ IDLE", fg='#444', bg='#1a1a2e')
                    if wid in self.worker_dots:
                        self.worker_dots[wid].config(fg='#555')
            except:
                pass

        # Update title bar with active project
        if active_project:
            self.title_project_var.set(f"▶ {active_project}")
        else:
            self.title_project_var.set("")

    def _update_projects(self):
        if not self.manager:
            return

        try:
            projects = self.manager.scan_projects()[:12]
            current_projects = set(projects)

            # v1.0.86: Remove rows for projects no longer in scan results (real-time update)
            old_projects = set(self.project_rows.keys())
            removed = old_projects - current_projects
            for code in removed:
                if code in self.project_rows:
                    row_data = self.project_rows[code]
                    row_data['row'].destroy()  # Remove Tkinter widget
                    del self.project_rows[code]

            for i, code in enumerate(projects):
                if code not in self.project_rows:
                    self._create_row(code, i)

                status = self.manager.quality_checker.get_project_status(code)
                if status:
                    labels = self.project_rows[code]['labels']

                    # Excel - show OK or %
                    excel_status = getattr(status, 'excel_status', '')
                    fallback_prompts = getattr(status, 'fallback_prompts', 0)

                    if excel_status == "complete":
                        text = "OK*" if fallback_prompts > 0 else "OK"
                        labels['excel'].config(text=text, fg='#00ff88')
                    elif excel_status == "partial":
                        img_prompts = getattr(status, 'img_prompts_count', 0)
                        total = getattr(status, 'total_scenes', 0)
                        if total > 0:
                            pct = int(img_prompts * 100 / total)
                            labels['excel'].config(text=f"{pct}%", fg='#00d9ff')
                        else:
                            labels['excel'].config(text="--", fg='#666')
                    elif excel_status == "fallback":
                        labels['excel'].config(text="FB", fg='#ffd93d')
                    else:
                        labels['excel'].config(text="--", fg='#666')

                    # Tham chieu (Reference images)
                    nv_done = getattr(status, 'characters_with_ref', 0)
                    nv_total = getattr(status, 'characters_count', 0)
                    if nv_total > 0:
                        color = '#00ff88' if nv_done >= nv_total else '#ff6b6b'
                        labels['thamchieu'].config(text=f"{nv_done}/{nv_total}", fg=color)
                    else:
                        labels['thamchieu'].config(text="--", fg='#666')

                    # Scene anh
                    img_done = getattr(status, 'images_done', 0)
                    img_total = getattr(status, 'total_scenes', 0)
                    if img_total > 0:
                        color = '#00ff88' if img_done >= img_total else '#00d9ff'
                        labels['images'].config(text=f"{img_done}/{img_total}", fg=color)
                    else:
                        labels['images'].config(text="--", fg='#666')

                    # Con lai (Remaining time for active project)
                    PROJECT_TIMEOUT = 6 * 3600
                    remaining_txt = "--"
                    remaining_fg = '#888'
                    for wid in ["chrome_1", "chrome_2"]:
                        try:
                            ws = self.manager.get_worker_status(wid)
                            if ws and ws.get('current_project') == code:
                                elapsed = ws.get('project_elapsed_seconds', 0)
                                if elapsed > 0:
                                    remaining = max(0, PROJECT_TIMEOUT - elapsed)
                                    rh = int(remaining // 3600)
                                    rm = int((remaining % 3600) // 60)
                                    if remaining <= 0:
                                        remaining_txt = "TIMEOUT!"
                                        remaining_fg = '#e94560'
                                    elif rh > 0:
                                        remaining_txt = f"{rh}h{rm:02d}m"
                                        remaining_fg = '#ffd93d'
                                    else:
                                        remaining_txt = f"{rm}m"
                                        remaining_fg = '#ffd93d'
                                break
                        except:
                            pass
                    labels['conlai'].config(text=remaining_txt, fg=remaining_fg)
        except Exception as e:
            pass

    def _create_row(self, code: str, i: int):
        bg = '#1a1a2e' if i % 2 == 0 else '#16213e'

        row = tk.Frame(self.projects_frame, bg=bg, cursor="hand2")
        row.pack(fill="x", pady=1)
        row.bind("<Button-1>", lambda e, c=code: self._select_project(c))

        labels = {}

        labels['code'] = tk.Label(row, text=code, width=10, bg=bg, fg='white', font=("Consolas", 10, "bold"), anchor="w")
        labels['code'].pack(side="left", padx=2)
        labels['code'].bind("<Button-1>", lambda e, c=code: self._select_project(c))

        labels['excel'] = tk.Label(row, text="--", width=5, bg=bg, fg='#666', font=("Consolas", 10), cursor="hand2")
        labels['excel'].pack(side="left", padx=2)
        labels['excel'].bind("<Button-1>", lambda e, c=code: self._show_excel_detail(c))

        labels['thamchieu'] = tk.Label(row, text="--", width=10, bg=bg, fg='#666', font=("Consolas", 10), cursor="hand2")
        labels['thamchieu'].pack(side="left", padx=2)
        labels['thamchieu'].bind("<Button-1>", lambda e, c=code: self._show_nv_detail(c))

        labels['images'] = tk.Label(row, text="--", width=9, bg=bg, fg='#666', font=("Consolas", 10))
        labels['images'].pack(side="left", padx=2)

        labels['conlai'] = tk.Label(row, text="--", width=10, bg=bg, fg='#888', font=("Consolas", 9))
        labels['conlai'].pack(side="left", padx=2)

        # Nut XONG cho moi project
        xong_btn = tk.Button(row, text="XONG", width=4,
                            command=lambda c=code: self._force_complete_by_code(c),
                            bg='#ff6600', fg='white',
                            font=("Arial", 8, "bold"),
                            padx=2, pady=0)
        xong_btn.pack(side="left", padx=2)
        labels['xong_btn'] = xong_btn

        self.project_rows[code] = {'row': row, 'labels': labels, 'bg': bg}

    def _show_detail_popup(self, code: str):
        """Double click - mo popup chi tiet."""
        ProjectDetail(self, code)

    def destroy(self):
        # Dung server neu dang chay
        if hasattr(self, '_server_process') and self._server_process:
            try:
                self._server_process.terminate()
                self._server_process.wait(timeout=3)
            except Exception:
                try:
                    self._server_process.kill()
                except Exception:
                    pass

        # Kill all processes khi dong GUI
        if self.manager:
            try:
                self.manager.stop_all()
                self.manager.kill_all_chrome()
            except:
                pass

        # Remove callback when closing
        if LOGGER_AVAILABLE:
            try:
                remove_callback(self._on_new_log)
            except:
                pass
        super().destroy()


# ================================================================================
# MAIN
# ================================================================================

if __name__ == "__main__":
    if not VM_AVAILABLE:
        print("Loi: Khong tim thay vm_manager.py!")
        input("Nhan Enter de thoat...")
        sys.exit(1)

    app = SimpleGUI()
    app.mainloop()
